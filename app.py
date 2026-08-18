from __future__ import annotations
from contextlib import contextmanager
from datetime import date, datetime, timedelta, timezone
from hashlib import sha256
from html import escape
from io import BytesIO
import gc
import hmac
import json
import os
import re
import time
from pathlib import Path
from typing import Any
from urllib.parse import urlencode, urljoin
from zoneinfo import ZoneInfo
import pandas as pd
import pyarrow as pa
import pyarrow.dataset as ds
import pyarrow.parquet as pq
import requests
import streamlit as st
from requests.auth import HTTPBasicAuth, HTTPDigestAuth
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

try:
    import fcntl
except ImportError:  # pragma: no cover - Streamlit Cloud runs on Linux.
    fcntl = None


# =============================================================================
# Configuration
# =============================================================================

APP_TITLE = "AtlasFlow"
APP_DIR = Path(__file__).resolve().parent
SNAPSHOT_DIR = APP_DIR / ".atlasflow_cache"
RAW_SNAPSHOT_FILE = SNAPSHOT_DIR / "reportdata_raw.parquet"
METADATA_SNAPSHOT_FILE = SNAPSHOT_DIR / "reportdata_metadata.json"
ODATA_ENDPOINT = "https://online.marorka.com/Odata/v1/ODataService.svc/ReportData"
REPORTPIVOTS_ENDPOINT = "https://online.marorka.com/Odata/v1/ODataService.svc/ReportPivots"
SHIPPIVOTS_ENDPOINT = "https://online.marorka.com/Odata/v1/ODataService.svc/ShipPivots"

SOURCE_CONFIGS = {
    "reportdata": {
        "label": "ReportData",
        "endpoint": ODATA_ENDPOINT,
        "snapshot_file": SNAPSHOT_DIR / "reportdata_raw.parquet",
        "metadata_file": SNAPSHOT_DIR / "reportdata_metadata.json",
        "datetime_candidates": ["StartDateTimeGMT", "DateTime", "dateTime", "Timestamp"],
    },
    "reportpivots": {
        "label": "ReportPivots",
        "endpoint": REPORTPIVOTS_ENDPOINT,
        "snapshot_file": SNAPSHOT_DIR / "reportpivots_raw.parquet",
        "metadata_file": SNAPSHOT_DIR / "reportpivots_metadata.json",
        "datetime_candidates": ["DateTime", "StartDateTimeGMT", "ReportDateTime", "Timestamp"],
    },
    "shippivots": {
        "label": "ShipPivots",
        "endpoint": SHIPPIVOTS_ENDPOINT,
        "snapshot_file": SNAPSHOT_DIR / "shippivots_raw.parquet",
        "metadata_file": SNAPSHOT_DIR / "shippivots_metadata.json",
        "datetime_candidates": ["DateTime", "StartDateTimeGMT", "Timestamp"],
    },
}
MAX_ODATA_PAGES = 1000
MAX_CONSECUTIVE_EMPTY_ODATA_PAGES = 2
API_CACHE_TTL_SECONDS = 21600  # 6 hours
API_FULL_START_DATE = date(2026, 1, 1)
TABLE_PREVIEW_ROW_LIMIT = 1000
DISPLAY_DATETIME_FORMAT = "%d/%m/%Y %H:%M"


# Persistent, user-ready multi-source snapshot settings.
ATLAS_SNAPSHOT_SCHEMA_VERSION = "2026-07-15-multisource-prepared-incremental-v1"
ATLAS_SNAPSHOT_GENERATIONS_TO_KEEP = 2
ATLAS_PREPARE_VERSION = "atlasflow_dynamic_pivot_v3_oil_stats_prepared_v1"
API_REQUEST_TIMEOUT_SECONDS = 60
API_REQUEST_MAX_ATTEMPTS = 3

DEFAULT_SOURCE_CHUNK_DAYS = {
    "reportdata": 31,
    "reportpivots": 31,
    "shippivots": 7,
}
DEFAULT_SOURCE_OVERLAP_DAYS = {
    "reportdata": 14,
    "reportpivots": 14,
    "shippivots": 10,
}
DEFAULT_SOURCE_FULL_REFRESH_MAX_MINUTES = {
    "reportdata": 240,
    "reportpivots": 240,
    "shippivots": 360,
}
DEFAULT_SOURCE_INCREMENTAL_REFRESH_MAX_MINUTES = {
    "reportdata": 45,
    "reportpivots": 60,
    "shippivots": 90,
}

EXCLUDED_REPORT_TYPES = [
    "Intake Report",
    "Fuel Change Report",
]

# ReportData is intentionally loaded in the same compact mode as the
# Performance KPIs app. Pulling every ValueDescription from ReportData is too
# broad for Streamlit Cloud and makes the source slow/fragile. These aliases
# are the KPI/calculation values that AtlasFlow needs now; we can expand this
# whitelist later in controlled groups.
PERFORMANCE_KPI_VALUE_ALIASES = {
    "Engine Distance [nm]": [
        "Engine Distance [nm]",
    ],
    "Distance Over Ground [nm]": [
        "Distance Over Ground [nm]",
    ],
    "Steaming Time Since Last Report [hh:mm]": [
        "Steaming Time Since Last Report [hh:mm]",
        "Steaming Time Since Last Report",
    ],
    "ME Load [%MCR]": [
        "ME Load [%MCR]",
        "ME Load [% MCR]",
    ],
    "Power from Torque Meter [kW]": [
        "Power from Torque Meter [kW]",
        "Total Shaft Power [kW] (kW)",
        "Total Shaft Power [kW]",
    ],
    "GPS Speed [kn]": [
        "GPS Speed [kn]",
        "GPS Speed",
        "Speed Over Ground [kn]",
        "Speed Over Ground",
    ],
    "Log Speed [kn]": [
        "Log Speed [kn]",
        "Log Speed",
        "Speed Through Water [kn]",
        "Speed Through Water",
    ],
    "Main Engine - HSHFO": ["Main Engine - HSHFO"],
    "Main Engine - HSLFO": ["Main Engine - HSLFO"],
    "Main Engine - MGO": ["Main Engine - MGO"],
    "Main Engine - ULSHFO": ["Main Engine - ULSHFO"],
    "Main Engine - ULSLFO": ["Main Engine - ULSLFO"],
    "Main Engine - VLSHFO": ["Main Engine - VLSHFO"],
    "Main Engine - VLSLFO": ["Main Engine - VLSLFO"],
    "Boiler - HSHFO": ["Boiler - HSHFO"],
    "Boiler - HSLFO": ["Boiler - HSLFO"],
    "Boiler - MGO": ["Boiler - MGO"],
    "Boiler - ULSHFO": ["Boiler - ULSHFO"],
    "Boiler - ULSLFO": ["Boiler - ULSLFO"],
    "Boiler - VLSHFO": ["Boiler - VLSHFO"],
    "Boiler - VLSLFO": ["Boiler - VLSLFO"],

    # Additional bunker/fuel consumption ValueDescriptions for ME/DG/Auxiliary
    # analysis. These are included in the ReportData API whitelist so bunker
    # consumption fields can be selected/exported and later used for derived
    # calculations without broad-loading all ReportData.
    "Main Engine Total Consumed": [
        "Main Engine Total Consumed",
        "ME Total Consumed",
        "Main Engine Consumption",
        "ME Consumption",
        "MEConsumed",
    ],
    "Diesel Generator Total Consumed": [
        "Diesel Generator Total Consumed",
        "DG Total Consumed",
        "DG Totals Consumed",
        "DGTotalsConsumed",
        "DGTotalConsumed",
        "Generator Total Consumed",
    ],
    "Auxiliary Engine Total Consumed": [
        "Auxiliary Engine Total Consumed",
        "Aux Engine Total Consumed",
        "Aux Total Consumed",
        "AuxConsumed",
    ],
    "Total Fuel Consumed": [
        "Total Fuel Consumed",
        "Total Consumed",
        "Total Consumption",
        "Bunker Consumption",
        "Fuel Consumption",
    ],
    "Diesel Generator - HSHFO": ["Diesel Generator - HSHFO", "DG - HSHFO", "Generator - HSHFO"],
    "Diesel Generator - HSLFO": ["Diesel Generator - HSLFO", "DG - HSLFO", "Generator - HSLFO"],
    "Diesel Generator - MGO": ["Diesel Generator - MGO", "DG - MGO", "Generator - MGO", "DG - MGO/MDO"],
    "Diesel Generator - ULSHFO": ["Diesel Generator - ULSHFO", "DG - ULSHFO", "Generator - ULSHFO"],
    "Diesel Generator - ULSLFO": ["Diesel Generator - ULSLFO", "DG - ULSLFO", "Generator - ULSLFO"],
    "Diesel Generator - VLSHFO": ["Diesel Generator - VLSHFO", "DG - VLSHFO", "Generator - VLSHFO"],
    "Diesel Generator - VLSLFO": ["Diesel Generator - VLSLFO", "DG - VLSLFO", "Generator - VLSLFO"],
    "Auxiliary Engine - HSHFO": ["Auxiliary Engine - HSHFO", "Aux Engine - HSHFO", "Aux - HSHFO"],
    "Auxiliary Engine - HSLFO": ["Auxiliary Engine - HSLFO", "Aux Engine - HSLFO", "Aux - HSLFO"],
    "Auxiliary Engine - MGO": ["Auxiliary Engine - MGO", "Aux Engine - MGO", "Aux - MGO", "Aux - MGO/MDO"],
    "Auxiliary Engine - ULSHFO": ["Auxiliary Engine - ULSHFO", "Aux Engine - ULSHFO", "Aux - ULSHFO"],
    "Auxiliary Engine - ULSLFO": ["Auxiliary Engine - ULSLFO", "Aux Engine - ULSLFO", "Aux - ULSLFO"],
    "Auxiliary Engine - VLSHFO": ["Auxiliary Engine - VLSHFO", "Aux Engine - VLSHFO", "Aux - VLSHFO"],
    "Auxiliary Engine - VLSLFO": ["Auxiliary Engine - VLSLFO", "Aux Engine - VLSLFO", "Aux - VLSLFO"],

    # Lub oil ROB/received and DG running hours for consumption aggregation.
    # These are pulled from ReportData so AtlasFlow can expose oil consumption
    # totals as selectable derived variables, not as engineered KPI metrics.
    "MELO ROB [ltr]": ["MELO ROB [ltr]"],
    "MELO Received [ltr]": ["MELO Received [ltr]"],
    "Cylinder Oil 1 ROB [ltr]": ["Cylinder Oil 1 ROB [ltr]"],
    "Cylinder Oil 1 Received [ltr]": ["Cylinder Oil 1 Received [ltr]"],
    "Cylinder Oil 2 ROB [ltr]": ["Cylinder Oil 2 ROB [ltr]"],
    "Cylinder Oil 2 Received [ltr]": ["Cylinder Oil 2 Received [ltr]"],
    "GELO ROB [ltr]": ["GELO ROB [ltr]", "GELO Grade ROB [ltr]"],
    "GELO Received [ltr]": ["GELO Received [ltr]"],
    "DG1 Running Hours [hh:mm]": ["DG1 Running Hours [hh:mm]"],
    "DG2 Running Hours [hh:mm]": ["DG2 Running Hours [hh:mm]"],
    "DG3 Running Hours [hh:mm]": ["DG3 Running Hours [hh:mm]"],
    "DG4 Running Hours [hh:mm]": ["DG4 Running Hours [hh:mm]"],
}

# Cargo / voyage values discovered in the live Marorka V1 ReportData scan.
# Keep these in a dedicated controlled group so AtlasFlow can support the
# Voyage Analysis workspace without broad-loading every ReportData variable.
CARGO_VALUE_ALIASES = {
    "Cargo Weight [tons]": ["Cargo Weight [tons]"],
    "Cargo Weight Added [MT]": ["Cargo Weight Added [MT]", "Cargo Weight Added [tons]"],
    "Cargo Weight Removed [MT]": ["Cargo Weight Removed [MT]", "Cargo Weight Removed [tons]"],
    "20ft Full Units": ["20ft Full Units"],
    "20ft Full Units Weight [tons]": ["20ft Full Units Weight [tons]"],
    "20ft Empty Units": ["20ft Empty Units"],
    "20ft Empty Units Weight [tons]": ["20ft Empty Units Weight [tons]"],
    "40ft Full Units": ["40ft Full Units"],
    "40ft Full Units Weight [tons]": ["40ft Full Units Weight [tons]"],
    "40ft Empty Units": ["40ft Empty Units"],
    "40ft Empty Units Weight [tons]": ["40ft Empty Units Weight [tons]"],
    "20ft Reefer Units": ["20ft Reefer Units"],
    "40ft Reefer Units": ["40ft Reefer Units"],
    "20ft DG Units": ["20ft DG Units"],
    "40ft DG Units": ["40ft DG Units"],
    "TEU Loaded Units": ["TEU Loaded Units"],
    "TEU Loaded Weight [tons]": ["TEU Loaded Weight [tons]"],
    "TEU Discharged Units": ["TEU Discharged Units"],
    "TEU Discharged Weight [tons]": ["TEU Discharged Weight [tons]"],
    "FEU Loaded Units": ["FEU Loaded Units"],
    "FEU Loaded Weight [tons]": ["FEU Loaded Weight [tons]"],
    "FEU Discharged Units": ["FEU Discharged Units"],
    "FEU Discharged Weight [tons]": ["FEU Discharged Weight [tons]"],
    "Reefers Loaded Units": ["Reefers Loaded Units"],
    "Reefers Loaded Weight [tons]": ["Reefers Loaded Weight [tons]"],
    "Reefers Discharged Units": ["Reefers Discharged Units"],
    "Reefers Discharged Weight [tons]": ["Reefers Discharged Weight [tons]"],
    "Reefer Units Weight [tons]": ["Reefer Units Weight [tons]"],
    "DG Units Weight [tons]": ["DG Units Weight [tons]"],
    "Total Number Full Units (20 and 40ft)": ["Total Number Full Units (20 and 40ft)"],
    "Total Number Empty Units (20 and 40ft)": ["Total Number Empty Units (20 and 40ft)"],
    "Total Number Reefer Units (20 and 40ft)": ["Total Number Reefer Units (20 and 40ft)"],
    "Total Number DG Units (20 and 40ft)": ["Total Number DG Units (20 and 40ft)"],
    "Total Number of 20ft Units (Full and Empty)": ["Total Number of 20ft Units (Full and Empty)"],
    "Total Number of 40ft Units (Full and Empty)": ["Total Number of 40ft Units (Full and Empty)"],
    "Total Full Units Weight (20 and 40ft) [tons]": ["Total Full Units Weight (20 and 40ft) [tons]"],
    "Total Empty Units Weight (20 and 40ft) [tons]": ["Total Empty Units Weight (20 and 40ft) [tons]"],
    "Total 20ft Units Weight (Full and Empty)": ["Total 20ft Units Weight (Full and Empty)"],
    "Total 40 ft Units Weight (Full and Empty)": ["Total 40 ft Units Weight (Full and Empty)"],
    "Total Units Weight (All Categories)": ["Total Units Weight (All Categories)"],
    "Draft Forward [m] (m)": ["Draft Forward [m] (m)", "Draft Forward [m]"],
    "Draft Aft [m] (m)": ["Draft Aft [m] (m)", "Draft Aft [m]"],
    "Observed Draft Forward [m]": ["Observed Draft Forward [m]"],
    "Observed Draft Aft [m]": ["Observed Draft Aft [m]"],
    "Observed Mean Draft [m]": ["Observed Mean Draft [m]"],
    "Calculated Draft Forward [m]": ["Calculated Draft Forward [m]"],
    "Calculated Draft Aft [m]": ["Calculated Draft Aft [m]"],
    "Calculated Mean Draft [m]": ["Calculated Mean Draft [m]"],
    "Ballast Amount [tons]": ["Ballast Amount [tons]"],
    "Dead Load [tons]": ["Dead Load [tons]"],
    "Air Draft [m]": ["Air Draft [m]"],
    "Cargo Operations Completed During Port Stay": ["Cargo Operations Completed During Port Stay"],
    "Commenced Cargo Operation Time [dd:mm:yyyy hh:mm]": ["Commenced Cargo Operation Time [dd:mm:yyyy hh:mm]"],
    "Completed Cargo Operation Time [dd:mm:yyyy hh:mm]": ["Completed Cargo Operation Time [dd:mm:yyyy hh:mm]"],
    "Cargo Checked: Bridges": ["Cargo Checked: Bridges"],
    "Cargo Checked: Lashings": ["Cargo Checked: Lashings"],
    "Reefer Energy [kWh]": ["Reefer Energy [kWh]"],
    "Total Reefer Power Draw (kW)": ["Total Reefer Power Draw (kW)"],
    "Average Power per Reefer [kW]": ["Average Power per Reefer [kW]"],

    # Voyage route context. Marorka ReportPivots exposes UN/LOCODEs only, while
    # ReportData also carries the user-facing full port names on Departure/Arrival reports.
    "Departure port name": ["Departure port name"],
    "Arrival port name": ["Arrival port name"],
    "Departure port UN/LOCODE": ["Departure port UN/LOCODE"],
    "Arrival port UN/LOCODE": ["Arrival port UN/LOCODE"],
}

REPORTDATA_VALUE_WHITELIST = sorted(
    {
        alias
        for alias_group in (PERFORMANCE_KPI_VALUE_ALIASES, CARGO_VALUE_ALIASES)
        for aliases in alias_group.values()
        for alias in aliases
    },
    key=str.casefold,
)
REPORTDATA_VALUE_WHITELIST_KEYS = {
    re.sub(r"[^a-z0-9]+", "", value.lower()) for value in REPORTDATA_VALUE_WHITELIST
}

SOURCE_COLUMNS = [
    "ReportId",
    "ShipName",
    "ReportType",
    "StartDateTimeGMT",
    "EndDateTimeGMT",
    "LapTime",
    "StateName",
    "ValueDescription",
    "ReportedValue",
]

PIVOT_IDENTITY_COLUMNS = [
    "ReportId",
    "ShipName",
    "ReportType",
    "StartDateTimeGMT",
    "EndDateTimeGMT",
    "LapTime",
    "StateName",
]

DEFAULT_DISPLAY_IDENTITY_COLUMNS = [
    "ShipName",
    "ReportType",
    "StartDateTimeGMT",
    "EndDateTimeGMT",
    "LapTime",
    "StateName",
]

# Derived calculation setup. These columns are calculated from the same Marorka
# source values used in the Performance KPIs app, but are exposed here as normal
# AtlasFlow variables that can be selected, displayed, filtered, summarized, and exported.
DERIVED_VALUE_ALIASES = {
    "Engine Distance [nm]": [
        "Engine Distance [nm]",
    ],
    "Distance Over Ground [nm]": [
        "Distance Over Ground [nm]",
    ],
    "Steaming Time Since Last Report [hh:mm]": [
        "Steaming Time Since Last Report [hh:mm]",
        "Steaming Time Since Last Report",
    ],
    "ME Load [%MCR]": [
        "ME Load [%MCR]",
        "ME Load [% MCR]",
    ],
    "Power from Torque Meter [kW]": [
        "Power from Torque Meter [kW]",
        "Total Shaft Power [kW] (kW)",
        "Total Shaft Power [kW]",
    ],
    "Main Engine - HSHFO": ["Main Engine - HSHFO"],
    "Main Engine - HSLFO": ["Main Engine - HSLFO"],
    "Main Engine - MGO": ["Main Engine - MGO"],
    "Main Engine - ULSHFO": ["Main Engine - ULSHFO"],
    "Main Engine - ULSLFO": ["Main Engine - ULSLFO"],
    "Main Engine - VLSHFO": ["Main Engine - VLSHFO"],
    "Main Engine - VLSLFO": ["Main Engine - VLSLFO"],
    "Boiler - HSHFO": ["Boiler - HSHFO"],
    "Boiler - HSLFO": ["Boiler - HSLFO"],
    "Boiler - MGO": ["Boiler - MGO"],
    "Boiler - ULSHFO": ["Boiler - ULSHFO"],
    "Boiler - ULSLFO": ["Boiler - ULSLFO"],
    "Boiler - VLSHFO": ["Boiler - VLSHFO"],
    "Boiler - VLSLFO": ["Boiler - VLSLFO"],

    # Additional bunker/fuel consumption ValueDescriptions for ME/DG/Auxiliary
    # analysis. These are included in the ReportData API whitelist so bunker
    # consumption fields can be selected/exported and later used for derived
    # calculations without broad-loading all ReportData.
    "Main Engine Total Consumed": [
        "Main Engine Total Consumed",
        "ME Total Consumed",
        "Main Engine Consumption",
        "ME Consumption",
        "MEConsumed",
    ],
    "Diesel Generator Total Consumed": [
        "Diesel Generator Total Consumed",
        "DG Total Consumed",
        "DG Totals Consumed",
        "DGTotalsConsumed",
        "DGTotalConsumed",
        "Generator Total Consumed",
    ],
    "Auxiliary Engine Total Consumed": [
        "Auxiliary Engine Total Consumed",
        "Aux Engine Total Consumed",
        "Aux Total Consumed",
        "AuxConsumed",
    ],
    "Total Fuel Consumed": [
        "Total Fuel Consumed",
        "Total Consumed",
        "Total Consumption",
        "Bunker Consumption",
        "Fuel Consumption",
    ],
    "Diesel Generator - HSHFO": ["Diesel Generator - HSHFO", "DG - HSHFO", "Generator - HSHFO"],
    "Diesel Generator - HSLFO": ["Diesel Generator - HSLFO", "DG - HSLFO", "Generator - HSLFO"],
    "Diesel Generator - MGO": ["Diesel Generator - MGO", "DG - MGO", "Generator - MGO", "DG - MGO/MDO"],
    "Diesel Generator - ULSHFO": ["Diesel Generator - ULSHFO", "DG - ULSHFO", "Generator - ULSHFO"],
    "Diesel Generator - ULSLFO": ["Diesel Generator - ULSLFO", "DG - ULSLFO", "Generator - ULSLFO"],
    "Diesel Generator - VLSHFO": ["Diesel Generator - VLSHFO", "DG - VLSHFO", "Generator - VLSHFO"],
    "Diesel Generator - VLSLFO": ["Diesel Generator - VLSLFO", "DG - VLSLFO", "Generator - VLSLFO"],
    "Auxiliary Engine - HSHFO": ["Auxiliary Engine - HSHFO", "Aux Engine - HSHFO", "Aux - HSHFO"],
    "Auxiliary Engine - HSLFO": ["Auxiliary Engine - HSLFO", "Aux Engine - HSLFO", "Aux - HSLFO"],
    "Auxiliary Engine - MGO": ["Auxiliary Engine - MGO", "Aux Engine - MGO", "Aux - MGO", "Aux - MGO/MDO"],
    "Auxiliary Engine - ULSHFO": ["Auxiliary Engine - ULSHFO", "Aux Engine - ULSHFO", "Aux - ULSHFO"],
    "Auxiliary Engine - ULSLFO": ["Auxiliary Engine - ULSLFO", "Aux Engine - ULSLFO", "Aux - ULSLFO"],
    "Auxiliary Engine - VLSHFO": ["Auxiliary Engine - VLSHFO", "Aux Engine - VLSHFO", "Aux - VLSHFO"],
    "Auxiliary Engine - VLSLFO": ["Auxiliary Engine - VLSLFO", "Aux Engine - VLSLFO", "Aux - VLSLFO"],

    # Lub oil ROB/received and DG running hours for consumption aggregation.
    # These are pulled from ReportData so AtlasFlow can expose oil consumption
    # totals as selectable derived variables, not as engineered KPI metrics.
    "MELO ROB [ltr]": ["MELO ROB [ltr]"],
    "MELO Received [ltr]": ["MELO Received [ltr]"],
    "Cylinder Oil 1 ROB [ltr]": ["Cylinder Oil 1 ROB [ltr]"],
    "Cylinder Oil 1 Received [ltr]": ["Cylinder Oil 1 Received [ltr]"],
    "Cylinder Oil 2 ROB [ltr]": ["Cylinder Oil 2 ROB [ltr]"],
    "Cylinder Oil 2 Received [ltr]": ["Cylinder Oil 2 Received [ltr]"],
    "GELO ROB [ltr]": ["GELO ROB [ltr]", "GELO Grade ROB [ltr]"],
    "GELO Received [ltr]": ["GELO Received [ltr]"],
    "DG1 Running Hours [hh:mm]": ["DG1 Running Hours [hh:mm]"],
    "DG2 Running Hours [hh:mm]": ["DG2 Running Hours [hh:mm]"],
    "DG3 Running Hours [hh:mm]": ["DG3 Running Hours [hh:mm]"],
    "DG4 Running Hours [hh:mm]": ["DG4 Running Hours [hh:mm]"],
}

ME_FUEL_COLUMNS = [
    "Main Engine - HSHFO",
    "Main Engine - HSLFO",
    "Main Engine - MGO",
    "Main Engine - ULSHFO",
    "Main Engine - ULSLFO",
    "Main Engine - VLSHFO",
    "Main Engine - VLSLFO",
]

BOILER_FUEL_COLUMNS = [
    "Boiler - HSHFO",
    "Boiler - HSLFO",
    "Boiler - MGO",
    "Boiler - ULSHFO",
    "Boiler - ULSLFO",
    "Boiler - VLSHFO",
    "Boiler - VLSLFO",
]

DG_FUEL_COLUMNS = [
    "Diesel Generator - HSHFO",
    "Diesel Generator - HSLFO",
    "Diesel Generator - MGO",
    "Diesel Generator - ULSHFO",
    "Diesel Generator - ULSLFO",
    "Diesel Generator - VLSHFO",
    "Diesel Generator - VLSLFO",
]

AUXILIARY_FUEL_COLUMNS = [
    "Auxiliary Engine - HSHFO",
    "Auxiliary Engine - HSLFO",
    "Auxiliary Engine - MGO",
    "Auxiliary Engine - ULSHFO",
    "Auxiliary Engine - ULSLFO",
    "Auxiliary Engine - VLSHFO",
    "Auxiliary Engine - VLSLFO",
]

DERIVED_VARIABLES = [
    "Calculated Slip",
    "ME Consumption Total",
    "DG Consumption Total",
    "Auxiliary Engine Consumption Total",
    "Boiler Sum",
    "Total Fuel Consumption",
    "Consumption ME 24 Hours [MT]",
    "SFOC [gr/Kwh]",
    "MELO Consumption Total [ltr]",
    "CYLO Consumption Total [ltr]",
    "GELO Consumption Total [ltr]",
    "Total DG Running Hours [hh:mm]",
]

VESSEL_GROUPS = {
    "Fleet 1": ["ATETI", "CMA CGM THALASSA", "CZECH", "DOLPHIN II", "GSL CHRISTEL ELISABETH", "GSL VINIA", "ORCA I", "MYNY", "SYDNEY EXPRESS"],
    "Fleet 2": ["AGIOS DIMITRIOS", "ELENI T", "MAIRA", "MELINA", "NEWYORKER", "NIKOLAS", "TORRANCE"],
    "Fleet 3": ["BREMERHAVEN EXPRESS", "CMA CGM ALCAZAR", "GSL ALICE", "GSL CHATEAU D'IF", "GSL ELEFTHERIA", "GSL MAREN", "GSL MELINA", "ISTANBUL EXPRESS"],
    "Fleet 4": ["ANTHEA Y", "COLOMBIA EXPRESS", "COSTA RICA EXPRESS", "JAMAICA EXPRESS", "MEXICO EXPRESS", "NICARAGUA EXPRESS", "PANAMA EXPRESS", "ZIM NORFOLK", "ZIM XIAMEN"],
    "Fleet 9": ["CMA CGM AMERICA", "CMA CGM SAMBHAR", "GSL ELENI", "GSL GRANIA", "GSL KALLIOPI", "GSL NINGBO", "MSC QINGDAO", "MSC TIANJIN"],
    "Fleet 10": ["CAPTAIN THANASIS I", "CMA CGM JAMAICA", "GSL CHRISTEN", "GSL NICOLETTA", "GSL VALERIE", "JULIE", "KUMASI", "MANET"],
    "Fleet 11": ["ATHENA", "EPAMINONDAS", "IAN H", "MARIANNA I", "MSC ROMA", "TINA I"],
    "Fleet 12": ["GSL DOROTHEA", "GSL KITHIRA", "GSL MARIA", "GSL MELITA", "GSL SYROS", "GSL TEGEA", "GSL TINOS", "GSL TRIPOLI"],
    "Fleet 14": ["GSL CHLOE", "GSL ELIZABETH", "GSL MAMITSA", "GSL MERCER", "GSL ROSSI", "GSL SUSAN", "TONSBERG"],
    "Fleet 15": ["GSL ALEXANDRA", "GSL ARCADIA", "GSL EFFIE", "GSL LYDIA", "GSL MYNY", "GSL SOFIA", "GSL VIOLETTA", "KOSTAS K", "MARIA Y"],
}

VESSEL_OPTIONS = sorted({v for vessels in VESSEL_GROUPS.values() for v in vessels})

st.set_page_config(page_title=APP_TITLE, layout="wide")


# =============================================================================
# Styling
# =============================================================================


def apply_custom_css() -> None:
    st.markdown(
        """
        <style>
        :root {
            --atlas-topbar-h: 67px;
            --atlas-sidebar-w: 324px;
            --atlas-ink: #0B1F33;
            --atlas-muted: #24364F;
            --atlas-soft: #64748B;
            --atlas-teal: #006B68;
            --atlas-teal-bright: #0AAEA6;
            --atlas-line: #D9E6E5;
            --atlas-bg: #FAFCFC;
            --atlas-chip: #DDF4F2;
        }

        html,
        body,
        .stApp {
            background: var(--atlas-bg) !important;
            color: var(--atlas-ink) !important;
            font-family: "Segoe UI", "Inter", "Aptos", Arial, sans-serif !important;
        }

        .stApp {
            background:
                linear-gradient(90deg, rgba(0, 107, 104, 0.035), transparent 24rem),
                linear-gradient(180deg, #FFFFFF 0%, #FBFDFD 46%, #F4FAF9 100%) !important;
        }

        header[data-testid="stHeader"] {
            left: 0 !important;
            right: 0 !important;
            width: 100vw !important;
            height: var(--atlas-topbar-h) !important;
            background: rgba(255, 255, 255, 0.98) !important;
            border-bottom: 1px solid rgba(15, 23, 42, 0.10) !important;
            box-shadow: 0 1px 12px rgba(15, 23, 42, 0.04) !important;
            z-index: 999990 !important;
        }

        header[data-testid="stHeader"] > div {
            height: var(--atlas-topbar-h) !important;
            background: transparent !important;
        }

        div[data-testid="stToolbar"] {
            top: 0.55rem !important;
            right: 1.55rem !important;
            z-index: 999995 !important;
        }

        div[data-testid="stDecoration"] {
            display: none !important;
        }

        /* Keep Streamlit's native sidebar behavior, but present it as the
           familiar navigation menu control instead of a directional chevron. */
        button[data-testid="stSidebarCollapsedControl"],
        button[data-testid="stExpandSidebarButton"],
        button[data-testid="stSidebarCollapseButton"],
        [data-testid="stSidebarCollapsedControl"] button,
        [data-testid="stExpandSidebarButton"] button,
        [data-testid="stSidebarCollapseButton"] button,
        header button[aria-label*="sidebar" i],
        header button[title*="sidebar" i] {
            width: 2.65rem !important;
            height: 2.65rem !important;
            min-width: 2.65rem !important;
            margin: 0.42rem 0 0 0.42rem !important;
            border: 1px solid transparent !important;
            border-radius: 6px !important;
            background: transparent !important;
            color: transparent !important;
            font-size: 0 !important;
            box-shadow: none !important;
        }

        button[data-testid="stSidebarCollapsedControl"]:hover,
        button[data-testid="stExpandSidebarButton"]:hover,
        button[data-testid="stSidebarCollapseButton"]:hover,
        [data-testid="stSidebarCollapsedControl"] button:hover,
        [data-testid="stExpandSidebarButton"] button:hover,
        [data-testid="stSidebarCollapseButton"] button:hover,
        header button[aria-label*="sidebar" i]:hover,
        header button[title*="sidebar" i]:hover {
            background: #EAF6F5 !important;
            border-color: #C9E8E5 !important;
        }

        button[data-testid="stSidebarCollapsedControl"] > *,
        button[data-testid="stExpandSidebarButton"] > *,
        button[data-testid="stSidebarCollapseButton"] > *,
        [data-testid="stSidebarCollapsedControl"] button > *,
        [data-testid="stExpandSidebarButton"] button > *,
        [data-testid="stSidebarCollapseButton"] button > *,
        header button[aria-label*="sidebar" i] > *,
        header button[title*="sidebar" i] > * {
            display: none !important;
        }

        button[data-testid="stSidebarCollapsedControl"]::after,
        button[data-testid="stExpandSidebarButton"]::after,
        button[data-testid="stSidebarCollapseButton"]::after,
        [data-testid="stSidebarCollapsedControl"] button::after,
        [data-testid="stExpandSidebarButton"] button::after,
        [data-testid="stSidebarCollapseButton"] button::after,
        header button[aria-label*="sidebar" i]::after,
        header button[title*="sidebar" i]::after {
            content: "\\2630";
            color: var(--atlas-teal) !important;
            font-family: Arial, sans-serif !important;
            font-size: 1.35rem !important;
            font-weight: 600 !important;
            line-height: 1 !important;
        }

        .atlas-topbar-brand {
            position: fixed;
            top: 0;
            left: 0;
            height: var(--atlas-topbar-h);
            display: flex;
            align-items: center;
            gap: 0.95rem;
            /* Reserve a clear gutter for the native sidebar menu control. */
            padding-left: 4.5rem;
            z-index: 999996;
            pointer-events: none;
        }

        .atlas-menu-lines {
            width: 24px;
            height: 24px;
            position: relative;
            flex: 0 0 24px;
        }

        .atlas-menu-lines::before,
        .atlas-menu-lines::after,
        .atlas-menu-lines span {
            content: "";
            position: absolute;
            left: 2px;
            width: 18px;
            height: 2px;
            border-radius: 999px;
            background: var(--atlas-teal);
        }

        .atlas-menu-lines::before { top: 6px; }
        .atlas-menu-lines span { top: 11px; }
        .atlas-menu-lines::after { top: 16px; }

        .atlas-logo-mark {
            width: 34px;
            height: 34px;
            position: relative;
            flex: 0 0 34px;
        }

        .atlas-logo-mark::before,
        .atlas-logo-mark::after {
            content: "";
            position: absolute;
            border-radius: 18px 18px 18px 6px;
            transform: rotate(34deg);
            box-shadow: 0 4px 12px rgba(0, 107, 104, 0.18);
        }

        .atlas-logo-mark::before {
            width: 20px;
            height: 31px;
            left: 4px;
            top: 2px;
            background: linear-gradient(145deg, #013F43 0%, #008C86 62%, #19BFB5 100%);
        }

        .atlas-logo-mark::after {
            width: 18px;
            height: 24px;
            left: 15px;
            top: 10px;
            background: linear-gradient(145deg, #0FB5AD 0%, #006B68 100%);
            opacity: 0.94;
        }

        .atlas-brand-word {
            color: #07515A;
            font-size: 1.85rem;
            font-weight: 400;
            line-height: 1;
            letter-spacing: 0;
        }

        @media (min-width: 769px) {
            section[data-testid="stSidebar"] {
                width: var(--atlas-sidebar-w) !important;
                min-width: var(--atlas-sidebar-w) !important;
                top: var(--atlas-topbar-h) !important;
                height: calc(100vh - var(--atlas-topbar-h)) !important;
                background:
                    radial-gradient(circle at 18px 8px, rgba(35, 209, 199, 0.18), transparent 16rem),
                    linear-gradient(180deg, #006A66 0%, #004743 46%, #003C39 100%) !important;
                border-right: 1px solid rgba(255, 255, 255, 0.16) !important;
                box-shadow: none !important;
            }

            section[data-testid="stSidebar"] > div {
                height: calc(100vh - var(--atlas-topbar-h)) !important;
                padding: 1.25rem 1.25rem 6rem 1.25rem !important;
            }

            .block-container {
                max-width: none !important;
                padding: 3.55rem 1.9rem 3rem 2.35rem !important;
            }

            /* A collapsed sidebar should create a focused wide-canvas view, not
               leave an empty sidebar-sized gutter beside the content. */
            section[data-testid="stSidebar"][aria-expanded="false"] {
                width: 0 !important;
                min-width: 0 !important;
            }

            section[data-testid="stSidebar"][aria-expanded="false"] ~ div[data-testid="stMain"] .block-container {
                max-width: 1440px !important;
                margin-left: auto !important;
                margin-right: auto !important;
                padding: 3.9rem clamp(2.4rem, 5vw, 5.75rem) 3.25rem !important;
            }

            section[data-testid="stSidebar"][aria-expanded="false"] ~ div[data-testid="stMain"] .atlas-metric-grid {
                gap: 1.55rem;
            }

            section[data-testid="stSidebar"][aria-expanded="false"] ~ div[data-testid="stMain"] .atlas-tabbar {
                flex-wrap: nowrap;
                justify-content: space-between;
                gap: 0.7rem;
                overflow-x: auto;
                scrollbar-width: none;
            }

            section[data-testid="stSidebar"][aria-expanded="false"] ~ div[data-testid="stMain"] .atlas-tabbar::-webkit-scrollbar {
                display: none;
            }
        }

        @media (max-width: 768px) {
            .atlas-brand-word { font-size: 1.35rem; }
            .atlas-logo-mark { width: 28px; height: 28px; }
            .block-container { padding: 5.2rem 1rem 2rem 1rem !important; }
        }

        section[data-testid="stSidebar"] [data-testid="stVerticalBlock"] {
            gap: 0.72rem !important;
        }

        section[data-testid="stSidebar"] h1,
        section[data-testid="stSidebar"] h2,
        section[data-testid="stSidebar"] h3,
        section[data-testid="stSidebar"] p,
        section[data-testid="stSidebar"] span,
        section[data-testid="stSidebar"] label,
        section[data-testid="stSidebar"] label *,
        section[data-testid="stSidebar"] [data-testid="stWidgetLabel"],
        section[data-testid="stSidebar"] [data-testid="stWidgetLabel"] *,
        section[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] *,
        section[data-testid="stSidebar"] [data-testid="stCaptionContainer"] * {
            color: #FFFFFF !important;
            -webkit-text-fill-color: #FFFFFF !important;
            font-weight: 760 !important;
            letter-spacing: 0 !important;
        }

        section[data-testid="stSidebar"] h3 {
            font-size: 1.05rem !important;
            margin-top: 0.25rem !important;
            margin-bottom: 0.1rem !important;
        }

        section[data-testid="stSidebar"] [data-testid="stCaptionContainer"] *,
        section[data-testid="stSidebar"] small {
            color: #D7FFFA !important;
            -webkit-text-fill-color: #D7FFFA !important;
        }

        section[data-testid="stSidebar"] div[data-baseweb="select"] > div,
        section[data-testid="stSidebar"] div[data-baseweb="input"],
        section[data-testid="stSidebar"] [data-testid="stMultiSelect"] div[data-baseweb="select"] > div,
        section[data-testid="stSidebar"] [data-testid="stSelectbox"] div[data-baseweb="select"] > div {
            min-height: 44px !important;
            border-radius: 9px !important;
            border: 1px solid rgba(8, 72, 70, 0.12) !important;
            background: #FFFFFF !important;
            box-shadow: none !important;
        }

        section[data-testid="stSidebar"] div[data-baseweb="select"] > div *,
        section[data-testid="stSidebar"] div[data-baseweb="input"] *,
        section[data-testid="stSidebar"] input {
            color: var(--atlas-ink) !important;
            -webkit-text-fill-color: var(--atlas-ink) !important;
        }

        section[data-testid="stSidebar"] div[data-testid="stButton"] button,
        section[data-testid="stSidebar"] div[data-testid="stButton"] button *,
        section[data-testid="stSidebar"] .stButton button,
        section[data-testid="stSidebar"] .stButton button * {
            color: #FFFFFF !important;
            -webkit-text-fill-color: #FFFFFF !important;
        }

        section[data-testid="stSidebar"] div[data-baseweb="select"] svg,
        section[data-testid="stSidebar"] div[data-baseweb="input"] svg {
            color: #334155 !important;
            fill: #334155 !important;
        }

        section[data-testid="stSidebar"] [data-baseweb="tag"] {
            background: var(--atlas-chip) !important;
            border: 1px solid rgba(0, 107, 104, 0.22) !important;
            border-radius: 999px !important;
            box-shadow: none !important;
        }

        section[data-testid="stSidebar"] [data-baseweb="tag"] *,
        section[data-testid="stSidebar"] [data-baseweb="tag"] span,
        section[data-testid="stSidebar"] [data-baseweb="tag"] svg {
            color: #12313E !important;
            -webkit-text-fill-color: #12313E !important;
            fill: #12313E !important;
        }

        section[data-testid="stSidebar"] [data-testid="stExpander"] {
            border-radius: 8px !important;
            border: 1px solid rgba(255, 255, 255, 0.72) !important;
            background: rgba(255, 255, 255, 0.07) !important;
            box-shadow: none !important;
        }

        section[data-testid="stSidebar"] [data-testid="stExpander"] summary,
        section[data-testid="stSidebar"] [data-testid="stExpander"] summary * {
            color: #FFFFFF !important;
            -webkit-text-fill-color: #FFFFFF !important;
            font-weight: 760 !important;
        }



        /* Sidebar confirmation/warning cards need dark text on the light card.
           The general sidebar rule forces labels/text to white, so alerts must
           be overridden explicitly for readability. */
        section[data-testid="stSidebar"] div[data-testid="stAlert"],
        section[data-testid="stSidebar"] div[data-testid="stAlert"] > div,
        section[data-testid="stSidebar"] div[data-testid="stAlert"] [role="alert"] {
            background: rgba(255, 255, 255, 0.96) !important;
            border: 1px solid rgba(0, 107, 104, 0.28) !important;
            border-radius: 9px !important;
            box-shadow: 0 2px 10px rgba(15, 23, 42, 0.10) !important;
        }

        section[data-testid="stSidebar"] div[data-testid="stAlert"] *,
        section[data-testid="stSidebar"] div[data-testid="stAlert"] p,
        section[data-testid="stSidebar"] div[data-testid="stAlert"] span,
        section[data-testid="stSidebar"] div[data-testid="stAlert"] div {
            color: var(--atlas-ink) !important;
            -webkit-text-fill-color: var(--atlas-ink) !important;
            font-weight: 650 !important;
        }

        section[data-testid="stSidebar"] div[data-testid="stAlert"] svg {
            color: var(--atlas-teal) !important;
            fill: var(--atlas-teal) !important;
        }

        section[data-testid="stSidebar"] [data-testid="stTooltipHoverTarget"],
        section[data-testid="stSidebar"] [data-testid="stTooltipIcon"],
        section[data-testid="stSidebar"] button[aria-label="Help"],
        section[data-testid="stSidebar"] button[title="Help"] {
            position: relative !important;
            width: 18px !important;
            height: 18px !important;
            min-width: 18px !important;
            min-height: 18px !important;
            border-radius: 999px !important;
            background: #063F3C !important;
            border: 1.5px solid #BDF7EF !important;
            box-shadow: 0 0 0 2px rgba(255, 255, 255, 0.08) !important;
            opacity: 1 !important;
            display: inline-flex !important;
            align-items: center !important;
            justify-content: center !important;
            overflow: hidden !important;
        }

        section[data-testid="stSidebar"] [data-testid="stTooltipHoverTarget"] svg,
        section[data-testid="stSidebar"] [data-testid="stTooltipIcon"] svg,
        section[data-testid="stSidebar"] button[aria-label="Help"] svg,
        section[data-testid="stSidebar"] button[title="Help"] svg {
            display: none !important;
        }

        section[data-testid="stSidebar"] [data-testid="stTooltipHoverTarget"]::after,
        section[data-testid="stSidebar"] [data-testid="stTooltipIcon"]::after,
        section[data-testid="stSidebar"] button[aria-label="Help"]::after,
        section[data-testid="stSidebar"] button[title="Help"]::after {
            content: "?" !important;
            color: #FFFFFF !important;
            -webkit-text-fill-color: #FFFFFF !important;
            font: 800 12px/1 Arial, sans-serif !important;
        }

        section[data-testid="stSidebar"] [data-testid="stTooltipHoverTarget"]:hover,
        section[data-testid="stSidebar"] [data-testid="stTooltipIcon"]:hover,
        section[data-testid="stSidebar"] button[aria-label="Help"]:hover,
        section[data-testid="stSidebar"] button[title="Help"]:hover {
            background: var(--atlas-teal-bright) !important;
            border-color: #FFFFFF !important;
        }

        section[data-testid="stSidebar"] [data-testid="stTooltipIcon"] [data-testid="stTooltipHoverTarget"] {
            background: transparent !important;
            border: 0 !important;
            box-shadow: none !important;
            overflow: visible !important;
        }

        section[data-testid="stSidebar"] [data-testid="stTooltipIcon"] [data-testid="stTooltipHoverTarget"]::after,
        section[data-testid="stSidebar"] [data-testid="stTooltipIcon"] button[aria-label^="Help"]::after {
            content: none !important;
        }

        section[data-testid="stSidebar"] [data-testid="stTooltipIcon"] button[aria-label^="Help"] {
            background: transparent !important;
            border: 0 !important;
            box-shadow: none !important;
            padding: 0 !important;
            width: 18px !important;
            height: 18px !important;
        }

        section[data-testid="stSidebar"] [data-testid="stTooltipIcon"] button[aria-label^="Help"] svg {
            display: block !important;
            width: 13px !important;
            height: 13px !important;
            color: #FFFFFF !important;
            stroke: #FFFFFF !important;
            fill: none !important;
        }

        section[data-testid="stSidebar"] [data-testid="stTooltipIcon"]::after {
            content: none !important;
        }

        section[data-testid="stSidebar"] [data-testid="stTooltipIcon"] [data-testid="stTooltipHoverTarget"] {
            position: absolute !important;
            inset: 0 !important;
            width: 100% !important;
            height: 100% !important;
            display: flex !important;
            align-items: center !important;
            justify-content: center !important;
        }

        section[data-testid="stSidebar"] [data-testid="stTooltipIcon"] button[aria-label^="Help"] {
            display: inline-flex !important;
            align-items: center !important;
            justify-content: center !important;
            width: 100% !important;
            height: 100% !important;
        }

        .dashboard-hero {
            box-sizing: border-box !important;
            min-height: 164px !important;
            margin: 0 0 0.82rem 0 !important;
            padding: 1.45rem 1.65rem 0.1rem 1.65rem !important;
            border-radius: 13px !important;
            border: 1px solid rgba(15, 23, 42, 0.11) !important;
            background: rgba(255, 255, 255, 0.93) !important;
            box-shadow: 0 2px 12px rgba(15, 23, 42, 0.12) !important;
            backdrop-filter: blur(4px) !important;
        }

        .eyebrow {
            color: var(--atlas-teal) !important;
            -webkit-text-fill-color: var(--atlas-teal) !important;
            font-size: 0.75rem !important;
            font-weight: 750 !important;
            letter-spacing: 0.14em !important;
            text-transform: uppercase !important;
            margin-bottom: 0.82rem !important;
        }

        .dashboard-title {
            color: var(--atlas-ink) !important;
            -webkit-text-fill-color: var(--atlas-ink) !important;
            font-family: "Segoe UI", "Inter", "Aptos Display", sans-serif !important;
            font-size: clamp(2.9rem, 3.6vw, 3.7rem) !important;
            font-weight: 400 !important;
            letter-spacing: 0 !important;
            line-height: 1.04 !important;
            margin: 0 !important;
            padding: 0 !important;
        }

        .dashboard-subtitle {
            color: var(--atlas-muted) !important;
            -webkit-text-fill-color: var(--atlas-muted) !important;
            font-size: 0.96rem !important;
            font-weight: 500 !important;
            margin-top: 0.75rem !important;
        }

        .api-load-caption,
        .atlas-pill {
            display: inline-flex !important;
            align-items: center !important;
            gap: 0.52rem !important;
            min-height: 37px !important;
            margin: 0 0 1.15rem 0 !important;
            padding: 0.4rem 0.9rem !important;
            border-radius: 999px !important;
            border: 1px solid rgba(15, 23, 42, 0.12) !important;
            background: rgba(255, 255, 255, 0.96) !important;
            box-shadow: 0 1px 8px rgba(15, 23, 42, 0.06) !important;
            color: var(--atlas-muted) !important;
            -webkit-text-fill-color: var(--atlas-muted) !important;
            font-size: 0.82rem !important;
            font-weight: 700 !important;
        }

        .api-load-caption strong,
        .api-load-caption span,
        .atlas-pill span {
            color: var(--atlas-ink) !important;
            -webkit-text-fill-color: var(--atlas-ink) !important;
            font-weight: 800 !important;
        }

        .api-load-clock {
            width: 17px;
            height: 17px;
            color: var(--atlas-teal);
            -webkit-text-fill-color: var(--atlas-teal);
            display: inline-flex;
            align-items: center;
            justify-content: center;
        }

        .api-load-clock svg {
            width: 17px;
            height: 17px;
            stroke: currentColor;
            stroke-width: 2.1;
            fill: none;
            stroke-linecap: round;
            stroke-linejoin: round;
        }

        div[data-testid="stTabs"] [data-baseweb="tab-list"] {
            gap: 1.35rem !important;
            border-bottom: 1px solid var(--atlas-line) !important;
        }

        button[data-baseweb="tab"] {
            height: 42px !important;
            padding: 0 0.45rem !important;
            color: #334155 !important;
            -webkit-text-fill-color: #334155 !important;
            font-size: 0.94rem !important;
            font-weight: 500 !important;
            letter-spacing: 0 !important;
        }

        button[data-baseweb="tab"][aria-selected="true"] {
            color: var(--atlas-teal) !important;
            -webkit-text-fill-color: var(--atlas-teal) !important;
            font-weight: 700 !important;
        }

        div[data-baseweb="tab-highlight"] {
            background-color: var(--atlas-teal) !important;
            height: 2px !important;
        }

        .section-title {
            color: var(--atlas-ink) !important;
            -webkit-text-fill-color: var(--atlas-ink) !important;
            font-size: 1.58rem !important;
            line-height: 1.22 !important;
            font-weight: 400 !important;
            letter-spacing: 0 !important;
            margin: 0.55rem 0 0.45rem 0 !important;
        }

        .stApp [data-testid="stMarkdownContainer"] p,
        .stApp [data-testid="stCaptionContainer"] p {
            color: var(--atlas-muted) !important;
            -webkit-text-fill-color: var(--atlas-muted) !important;
        }

        .stApp section[data-testid="stSidebar"] [data-testid="stWidgetLabel"] *,
        .stApp section[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p,
        .stApp section[data-testid="stSidebar"] [data-testid="stCaptionContainer"] p {
            color: #FFFFFF !important;
            -webkit-text-fill-color: #FFFFFF !important;
        }

        .atlas-tabbar {
            display: flex;
            flex-wrap: wrap;
            align-items: flex-end;
            gap: 1.35rem;
            border-bottom: 1px solid var(--atlas-line);
            padding: 0 0 0.05rem 0;
            margin: 0.35rem 0 1rem 0;
        }

        .atlas-tabbar.compact {
            gap: 1.65rem;
            margin-top: 0.2rem;
            margin-bottom: 1.1rem;
        }

        .atlas-tablink {
            position: relative;
            display: inline-flex;
            align-items: center;
            min-height: 42px;
            padding: 0 0.45rem 0.58rem 0.45rem;
            color: #334155 !important;
            -webkit-text-fill-color: #334155 !important;
            font-size: 0.94rem !important;
            font-weight: 500 !important;
            line-height: 1.15 !important;
            text-decoration: none !important;
            border: 0 !important;
            background: transparent !important;
        }

        .atlas-tablink:hover {
            color: var(--atlas-teal) !important;
            -webkit-text-fill-color: var(--atlas-teal) !important;
            text-decoration: none !important;
        }

        .atlas-tablink.active {
            color: var(--atlas-teal) !important;
            -webkit-text-fill-color: var(--atlas-teal) !important;
            font-weight: 700 !important;
        }

        .atlas-tablink.active::after {
            content: "";
            position: absolute;
            left: 0.25rem;
            right: 0.25rem;
            bottom: -0.05rem;
            height: 2px;
            border-radius: 999px;
            background: var(--atlas-teal);
        }


        /* Native Streamlit radio widgets used as text-only tab bars.
           This avoids HTML links, so tab clicks never open a browser tab. */
        div[data-testid="stRadio"] > div[role="radiogroup"] {
            display: flex !important;
            flex-direction: row !important;
            flex-wrap: wrap !important;
            align-items: flex-end !important;
            gap: 1.35rem !important;
            border-bottom: 1px solid var(--atlas-line) !important;
            padding: 0 0 0.05rem 0 !important;
            margin: 0.35rem 0 1rem 0 !important;
        }

        div[data-testid="stRadio"] > div[role="radiogroup"] label {
            position: relative !important;
            min-height: 42px !important;
            padding: 0 0.45rem 0.58rem 0.45rem !important;
            margin: 0 !important;
            background: transparent !important;
            border: 0 !important;
            box-shadow: none !important;
            cursor: pointer !important;
            display: inline-flex !important;
            align-items: center !important;
        }

        div[data-testid="stRadio"] > div[role="radiogroup"] label div[data-baseweb="radio"],
        div[data-testid="stRadio"] > div[role="radiogroup"] label > div:has(input[type="radio"]),
        div[data-testid="stRadio"] > div[role="radiogroup"] label input[type="radio"] {
            display: none !important;
        }

        div[data-testid="stRadio"] > div[role="radiogroup"] label p,
        div[data-testid="stRadio"] > div[role="radiogroup"] label span,
        div[data-testid="stRadio"] > div[role="radiogroup"] label [data-testid="stMarkdownContainer"] * {
            color: #334155 !important;
            -webkit-text-fill-color: #334155 !important;
            font-size: 0.94rem !important;
            font-weight: 500 !important;
            line-height: 1.15 !important;
            text-decoration: none !important;
        }

        div[data-testid="stRadio"] > div[role="radiogroup"] label:hover p,
        div[data-testid="stRadio"] > div[role="radiogroup"] label:hover span,
        div[data-testid="stRadio"] > div[role="radiogroup"] label:hover [data-testid="stMarkdownContainer"] * {
            color: var(--atlas-teal) !important;
            -webkit-text-fill-color: var(--atlas-teal) !important;
        }

        div[data-testid="stRadio"] > div[role="radiogroup"] label:has(input[type="radio"]:checked) p,
        div[data-testid="stRadio"] > div[role="radiogroup"] label:has(input[type="radio"]:checked) span,
        div[data-testid="stRadio"] > div[role="radiogroup"] label:has(input[type="radio"]:checked) [data-testid="stMarkdownContainer"] * {
            color: var(--atlas-teal) !important;
            -webkit-text-fill-color: var(--atlas-teal) !important;
            font-weight: 700 !important;
        }

        div[data-testid="stRadio"] > div[role="radiogroup"] label:has(input[type="radio"]:checked)::after {
            content: "";
            position: absolute;
            left: 0.25rem;
            right: 0.25rem;
            bottom: -0.05rem;
            height: 2px;
            border-radius: 999px;
            background: var(--atlas-teal);
        }

        .atlas-metric-grid {
            display: grid;
            grid-template-columns: repeat(4, minmax(0, 1fr));
            gap: 1.35rem;
            margin: 0 0 0.4rem 0;
        }
        .atlas-metric-grid-three { grid-template-columns: repeat(3, minmax(0, 1fr)); }

        .atlas-metric-card {
            display: grid;
            grid-template-columns: 58px minmax(0, 1fr);
            align-items: center;
            gap: 1rem;
            min-height: 95px;
            padding: 1rem 1.05rem;
            background: #FFFFFF;
            border: 1px solid rgba(15, 23, 42, 0.11);
            border-radius: 9px;
            box-shadow: 0 3px 13px rgba(15, 23, 42, 0.08);
        }

        .atlas-metric-icon {
            width: 55px;
            height: 55px;
            border-radius: 8px;
            background: linear-gradient(135deg, #006F6A 0%, var(--atlas-teal-bright) 100%);
            color: #FFFFFF;
            display: inline-flex;
            align-items: center;
            justify-content: center;
        }

        .atlas-metric-icon svg {
            width: 32px;
            height: 32px;
            stroke: #FFFFFF;
            stroke-width: 2.2;
            fill: none;
            stroke-linecap: round;
            stroke-linejoin: round;
        }

        .atlas-metric-label {
            color: var(--atlas-muted);
            font-size: 0.88rem;
            line-height: 1.2;
            font-weight: 600;
            margin-bottom: 0.35rem;
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
        }

        .atlas-metric-value {
            color: var(--atlas-ink);
            font-size: 1.85rem;
            line-height: 1;
            font-weight: 400;
            letter-spacing: 0;
        }

        .atlas-table-frame,
        div[data-testid="stDataFrame"] {
            border: 1px solid rgba(15, 23, 42, 0.10) !important;
            border-radius: 7px !important;
            background: #FFFFFF !important;
            overflow: hidden;
            box-shadow: 0 2px 10px rgba(15, 23, 42, 0.08) !important;
        }

        .atlas-table-frame { margin-top: 0.4rem; }
        .atlas-table-scroll { width: 100%; overflow: auto; max-height: 560px; }

        .atlas-preview-table {
            width: 100%;
            border-collapse: collapse;
            table-layout: fixed;
            font-size: 0.88rem;
            color: var(--atlas-ink);
        }

        .atlas-preview-table thead th {
            position: sticky;
            top: 0;
            z-index: 1;
            text-align: left;
            padding: 0.62rem 1rem;
            color: #FFFFFF;
            background: linear-gradient(180deg, #006F6A 0%, #005D59 100%);
            border-right: 1px solid rgba(255, 255, 255, 0.20);
            font-weight: 750;
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
        }

        .atlas-preview-table tbody td {
            padding: 0.48rem 1rem;
            border-top: 1px solid #E7EDEF;
            border-right: 1px solid #E7EDEF;
            background: #FFFFFF;
            color: #12233D;
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
        }

        .atlas-preview-table tbody tr:nth-child(even) td { background: #FCFDFD; }
        .atlas-preview-table th:last-child,
        .atlas-preview-table td:last-child { border-right: 0; }

        .atlas-table-empty {
            margin-top: 1rem;
            padding: 1.1rem 1.2rem;
            border: 1px solid rgba(15, 23, 42, 0.10);
            border-radius: 9px;
            background: #FFFFFF;
            color: var(--atlas-muted);
            box-shadow: 0 2px 10px rgba(15, 23, 42, 0.06);
        }

        div[data-testid="stMetric"] {
            background: #FFFFFF !important;
            border: 1px solid rgba(15, 23, 42, 0.10) !important;
            border-radius: 9px !important;
            box-shadow: 0 2px 10px rgba(15, 23, 42, 0.08) !important;
        }

        [data-baseweb="popover"],
        [data-baseweb="menu"],
        [role="listbox"] {
            background: #FFFFFF !important;
            color: var(--atlas-ink) !important;
        }

        [role="option"],
        [role="option"] *,
        [data-baseweb="menu"] * {
            color: var(--atlas-ink) !important;
            -webkit-text-fill-color: var(--atlas-ink) !important;
        }

        div[data-testid="stAlert"],
        div[data-testid="stAlert"] > div,
        div[data-testid="stAlert"] [role="alert"] {
            background: rgba(255, 255, 255, 0.92) !important;
            border: 1px solid rgba(15, 118, 110, 0.18) !important;
            color: var(--atlas-ink) !important;
            border-radius: 9px !important;
            box-shadow: none !important;
        }

        /* Final high-specificity sidebar alert override: Streamlit's markdown
           container inside the alert was still inheriting the global sidebar
           white-text rule. Keep the alert card light and force readable text. */
        section[data-testid="stSidebar"] div[data-testid="stAlert"] [data-testid="stMarkdownContainer"],
        section[data-testid="stSidebar"] div[data-testid="stAlert"] [data-testid="stMarkdownContainer"] *,
        section[data-testid="stSidebar"] div[data-testid="stAlert"] [data-testid="stAlertContent"],
        section[data-testid="stSidebar"] div[data-testid="stAlert"] [data-testid="stAlertContent"] *,
        section[data-testid="stSidebar"] div[data-testid="stAlert"] [role="alert"],
        section[data-testid="stSidebar"] div[data-testid="stAlert"] [role="alert"] *,
        section[data-testid="stSidebar"] div[data-testid="stAlert"] p,
        section[data-testid="stSidebar"] div[data-testid="stAlert"] span {
            color: #0B1F33 !important;
            -webkit-text-fill-color: #0B1F33 !important;
            opacity: 1 !important;
            font-weight: 650 !important;
            text-shadow: none !important;
        }

        section[data-testid="stSidebar"] div[data-testid="stAlert"] {
            background: #FFFFFF !important;
        }

        .stButton button,
        .stDownloadButton button,
        div[data-testid="stButton"] button,
        div[data-testid="stDownloadButton"] button,
        button[kind="primary"],
        button[kind="secondary"] {
            border-radius: 8px !important;
            background: linear-gradient(135deg, #006F6A, var(--atlas-teal-bright)) !important;
            border: 1px solid rgba(0, 107, 104, 0.32) !important;
            box-shadow: none !important;
        }

        .stButton button,
        .stButton button *,
        .stDownloadButton button,
        .stDownloadButton button *,
        div[data-testid="stButton"] button,
        div[data-testid="stButton"] button *,
        div[data-testid="stDownloadButton"] button,
        div[data-testid="stDownloadButton"] button *,
        button[kind="primary"],
        button[kind="primary"] *,
        button[kind="secondary"],
        button[kind="secondary"] * {
            color: #FFFFFF !important;
            -webkit-text-fill-color: #FFFFFF !important;
            font-weight: 600 !important;
            opacity: 1 !important;
            text-shadow: none !important;
        }

        .stButton button:hover,
        .stDownloadButton button:hover,
        div[data-testid="stButton"] button:hover,
        div[data-testid="stDownloadButton"] button:hover,
        button[kind="primary"]:hover,
        button[kind="secondary"]:hover {
            color: #FFFFFF !important;
            -webkit-text-fill-color: #FFFFFF !important;
            filter: brightness(1.04);
        }

        @media (max-width: 1100px) {
            .atlas-metric-grid, .atlas-metric-grid-three { grid-template-columns: repeat(2, minmax(0, 1fr)); }
        }

        @media (max-width: 640px) {
            .atlas-metric-grid, .atlas-metric-grid-three { grid-template-columns: 1fr; }
            .dashboard-title { font-size: 2.55rem !important; }
            .dashboard-hero { min-height: 150px !important; }
        }
        </style>
        <div class="atlas-topbar-brand" aria-hidden="true">
            <div class="atlas-logo-mark"></div>
            <div class="atlas-brand-word">Atlas Flow</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_header(selected_group: str, selected_vessels: list[str], selected_variables: list[str]) -> None:
    vessel_text = "All selected vessels" if len(selected_vessels) != 1 else selected_vessels[0]
    variable_text = "No variables selected" if not selected_variables else f"{len(selected_variables):,} selected variables"
    st.markdown(
        f"""
        <div class="dashboard-hero">
            <div class="eyebrow">Marorka API Explorer</div>
            <h1 class="dashboard-title">Atlas Flow</h1>
            <div class="dashboard-subtitle">
                {escape(selected_group)} | {escape(vessel_text)} | {escape(variable_text)}
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_api_load_caption(metadata: dict[str, Any] | None) -> None:
    metadata = metadata or {}
    last_load = metadata.get("loaded_at_local") or metadata.get("loaded_at_utc") or "-"
    last_load_display = str(last_load).replace(" EEST", "").replace(" EET", "")
    st.markdown(
        f"""
        <div class="api-load-caption">
            <span class="api-load-clock" aria-hidden="true">
                <svg viewBox="0 0 24 24">
                    <circle cx="12" cy="12" r="9"></circle>
                    <path d="M12 7v5l3 2"></path>
                </svg>
            </span>
            <strong>Last API load:</strong> <span>{escape(last_load_display)} LT</span>
        </div>
        """,
        unsafe_allow_html=True,
    )


def slugify_tab_label(label: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", str(label).lower()).strip("-") or "tab"


def current_query_params_dict() -> dict[str, str]:
    try:
        items = st.query_params.to_dict()
    except Exception:
        try:
            raw_items = st.experimental_get_query_params()
            items = {key: str(value[0]) if isinstance(value, list) and value else str(value) for key, value in raw_items.items()}
        except Exception:
            items = {}
    return {str(key): str(value) for key, value in items.items()}


def get_tab_selection(param_name: str, options: list[str], default: str) -> str:
    slug_to_option = {slugify_tab_label(option): option for option in options}
    raw_value = get_query_param(param_name, slugify_tab_label(default)).strip().lower()
    if raw_value in slug_to_option:
        return slug_to_option[raw_value]
    if raw_value in options:
        return raw_value
    return default if default in options else options[0]


def render_text_tab_bar(
    options: list[str],
    selected: str,
    *,
    param_name: str,
    css_class: str = "",
    reset_params: list[str] | None = None,
) -> str:
    """Render a native Streamlit text-tab selector.

    Older AtlasFlow batches used HTML anchor links for the tab strip. Those
    looked correct, but browsers treated them as links and sometimes opened the
    app in a new tab. This version uses st.radio under the hood, styled as
    text-only tabs, so clicking a tab only updates Streamlit session state.
    """
    if not options:
        return selected

    key_suffix = slugify_tab_label(css_class) if css_class else "main"
    state_key = f"atlas_tab_{param_name}_{key_suffix}"
    if selected not in options:
        selected = options[0]
    if st.session_state.get(state_key) not in options:
        st.session_state[state_key] = selected

    previous_value = st.session_state.get(state_key, selected)
    choice = st.radio(
        " ",
        options=options,
        horizontal=True,
        label_visibility="collapsed",
        key=state_key,
    )

    if choice != previous_value:
        for reset_param in reset_params or []:
            reset_key = f"atlas_tab_{reset_param}_compact"
            st.session_state.pop(reset_key, None)
            if reset_param == "preview":
                st.session_state["atlas_reportdata_preview_mode"] = "Clean Dataset"
        st.session_state[state_key] = choice

    return choice


# =============================================================================
# Secrets/auth/API helpers
# =============================================================================


class MarorkaConfigError(RuntimeError):
    pass


def read_secret(name: str, default: str = "") -> str:
    try:
        value = st.secrets.get(name, os.getenv(name, default))
    except Exception:
        value = os.getenv(name, default)
    return str(value).strip() if value is not None else default


def app_timezone() -> ZoneInfo:
    timezone_name = read_secret("APP_TIMEZONE", "Europe/Athens")
    try:
        return ZoneInfo(timezone_name)
    except Exception:
        return ZoneInfo("Europe/Athens")


def local_time_label(dt_utc: datetime | None = None) -> str:
    dt_utc = dt_utc or datetime.now(timezone.utc)
    if dt_utc.tzinfo is None:
        dt_utc = dt_utc.replace(tzinfo=timezone.utc)
    local_dt = dt_utc.astimezone(app_timezone())
    return local_dt.strftime("%d-%m-%Y %H:%M:%S %Z")


def get_query_param(name: str, default: str = "") -> str:
    try:
        value = st.query_params.get(name, default)
    except Exception:
        value = st.experimental_get_query_params().get(name, [default])
    if isinstance(value, list):
        return str(value[0]) if value else default
    return str(value) if value is not None else default


def is_warmup_request() -> bool:
    return get_query_param("warmup", "0") == "1"


def warmup_token_is_valid() -> bool:
    expected_token = read_secret("WARMUP_TOKEN")
    provided_token = get_query_param("token", "")
    return bool(expected_token) and hmac.compare_digest(provided_token, expected_token)


def require_dashboard_password() -> None:
    dashboard_password = read_secret("DASHBOARD_PASSWORD")
    if not dashboard_password:
        return

    if st.session_state.get("dashboard_authenticated"):
        return

    apply_custom_css()
    st.markdown(
        """
        <div class="dashboard-hero">
            <div class="eyebrow">Secure access</div>
            <h1 class="dashboard-title">Atlas Flow</h1>
            <div class="dashboard-subtitle">Enter your dashboard password to continue.</div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    entered_password = st.text_input("Password", type="password")

    if st.button("Sign in", type="primary"):
        if hmac.compare_digest(entered_password, dashboard_password):
            st.session_state["dashboard_authenticated"] = True
            st.rerun()
        st.error("Invalid password.")

    st.stop()


def request_auth(username: str, password: str, auth_method: str) -> Any:
    method = auth_method.lower()
    if method == "basic":
        return HTTPBasicAuth(username, password)
    if method == "digest":
        return HTTPDigestAuth(username, password)
    if method == "bearer":
        return None
    if method in {"none", "anonymous", ""}:
        return None
    raise MarorkaConfigError("Unsupported MARORKA_AUTH_METHOD. Use basic, digest, bearer, or none.")


def request_headers(token: str, auth_method: str) -> dict[str, str]:
    headers = {"Accept": "application/json"}
    if auth_method.lower() == "bearer":
        if not token:
            raise MarorkaConfigError("MARORKA_TOKEN is required for bearer auth.")
        headers["Authorization"] = f"Bearer {token}"
    return headers




RETRYABLE_HTTP_STATUSES = {500, 502, 503, 504}
RETRYABLE_REQUEST_EXCEPTIONS = (
    requests.ConnectionError,
    requests.ReadTimeout,
    requests.Timeout,
    requests.exceptions.ChunkedEncodingError,
)


def request_with_retry(
    session: requests.Session,
    url: str,
    *,
    auth: Any,
    timeout: int = 90,
    max_attempts: int = 5,
    base_sleep_seconds: float = 2.0,
) -> requests.Response:
    """GET an OData page with retry/backoff for transient Marorka disconnects."""
    last_error: Exception | None = None

    for attempt in range(1, max_attempts + 1):
        try:
            response = session.get(url, auth=auth, timeout=timeout)
            if response.status_code in RETRYABLE_HTTP_STATUSES and attempt < max_attempts:
                time.sleep(base_sleep_seconds * (2 ** (attempt - 1)))
                continue
            return response
        except RETRYABLE_REQUEST_EXCEPTIONS as exc:
            last_error = exc
            if attempt >= max_attempts:
                raise
            time.sleep(base_sleep_seconds * (2 ** (attempt - 1)))

    if last_error is not None:
        raise last_error
    raise requests.RequestException("Marorka API request failed before a response was received.")


def odata_quote(value: str) -> str:
    return str(value).replace("'", "''")


def build_reportdata_value_filter() -> str:
    value_filters = [
        f"ValueDescription eq '{odata_quote(value)}'"
        for value in REPORTDATA_VALUE_WHITELIST
    ]
    return "(" + " or ".join(value_filters) + ")"


def build_odata_url(start_date: date) -> str:
    # Keep the OData request simple. The Marorka OData V1 ReportData endpoint
    # rejects long ValueDescription OR filters with 404. We therefore request
    # the date window only, then apply the KPI/consumption ValueDescription
    # whitelist locally inside compact_odata_rows() page-by-page before writing
    # to the Parquet snapshot. This preserves the same final dataset while
    # avoiding invalid/oversized OData URLs.
    start_text = start_date.strftime("%Y-%m-%d")
    params = {
        "$filter": f"StartDateTimeGMT gt DateTime'{start_text}'",
        "$select": ",".join(SOURCE_COLUMNS),
    }
    return f"{ODATA_ENDPOINT}?{urlencode(params)}"


def extract_odata_page(payload: Any) -> tuple[list[dict[str, Any]], str | None]:
    if isinstance(payload, list):
        return payload, None

    if not isinstance(payload, dict):
        raise ValueError("Could not parse OData response payload.")

    rows = payload.get("value")
    next_link = payload.get("@odata.nextLink") or payload.get("odata.nextLink")

    if rows is None and isinstance(payload.get("d"), dict):
        data = payload["d"]
        rows = data.get("results")
        next_link = next_link or data.get("__next")

    if rows is None:
        raise ValueError("Could not find OData rows in the API response.")

    return rows, next_link


def should_continue_odata_paging(
    *,
    current_url: str,
    next_link: str | None,
    seen_urls: set[str],
    consecutive_empty_pages: int,
) -> tuple[bool, str | None, str | None]:
    """Return whether OData paging should continue plus next URL and stop reason.

    AtlasFlow follows the OData nextLink until the feed is exhausted, but still
    protects Streamlit Cloud from pagination loops or abnormal empty-page runs.
    """
    if not next_link:
        return False, None, "end_of_feed"

    if consecutive_empty_pages >= MAX_CONSECUTIVE_EMPTY_ODATA_PAGES:
        return False, None, "consecutive_empty_pages"

    resolved_next_url = urljoin(current_url, next_link)
    if resolved_next_url in seen_urls:
        return False, None, "repeated_next_link"

    return True, resolved_next_url, None


def rows_to_dataframe(rows: list[dict[str, Any]]) -> pd.DataFrame:
    df = pd.DataFrame(rows)
    if "__metadata" in df.columns:
        df = df.drop(columns=["__metadata"])
    for column in SOURCE_COLUMNS:
        if column not in df.columns:
            df[column] = pd.NA
    return df[SOURCE_COLUMNS]


def compact_odata_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    compact_rows: list[dict[str, Any]] = []
    for row in rows:
        value_description = row.get("ValueDescription")
        if value_description is None:
            continue
        if re.sub(r"[^a-z0-9]+", "", str(value_description).lower()) not in REPORTDATA_VALUE_WHITELIST_KEYS:
            continue
        if row.get("ReportType") in EXCLUDED_REPORT_TYPES:
            continue
        compact_rows.append({column: row.get(column) for column in SOURCE_COLUMNS})
    return compact_rows


def fetch_report_data(
    username: str,
    password: str,
    token: str,
    auth_method: str,
    start_date: date,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    started_at = time.perf_counter()
    next_url = build_odata_url(start_date)
    kept_rows: list[dict[str, Any]] = []
    seen_urls: set[str] = set()
    pages = 0
    total_bytes = 0
    scanned_rows = 0
    consecutive_empty_pages = 0
    paging_stop_reason = "max_page_limit"
    first_url = next_url
    auth = request_auth(username, password, auth_method)
    headers = request_headers(token, auth_method)

    with requests.Session() as session:
        session.headers.update(headers)
        for _ in range(MAX_ODATA_PAGES):
            if next_url in seen_urls:
                paging_stop_reason = "repeated_current_url"
                break
            seen_urls.add(next_url)

            response = request_with_retry(session, next_url, auth=auth, timeout=90)
            total_bytes += len(response.content)
            response.raise_for_status()
            pages += 1

            page_rows, next_link = extract_odata_page(response.json())
            scanned_rows += len(page_rows)
            kept_rows.extend(compact_odata_rows(page_rows))
            consecutive_empty_pages = consecutive_empty_pages + 1 if len(page_rows) == 0 else 0

            should_continue, resolved_next_url, stop_reason = should_continue_odata_paging(
                current_url=next_url,
                next_link=next_link,
                seen_urls=seen_urls,
                consecutive_empty_pages=consecutive_empty_pages,
            )
            if not should_continue:
                paging_stop_reason = stop_reason
                break
            next_url = resolved_next_url or next_url

    loaded_at_utc = datetime.now(timezone.utc)
    metadata = {
        "loaded_at_utc": loaded_at_utc.strftime("%d-%m-%Y %H:%M:%S UTC"),
        "loaded_at_local": local_time_label(loaded_at_utc),
        "rows": len(kept_rows),
        "kept_rows": len(kept_rows),
        "scanned_rows": scanned_rows,
        "discarded_rows": max(scanned_rows - len(kept_rows), 0),
        "pages": pages,
        "downloaded_mb": round(total_bytes / 1024 / 1024, 2),
        "fetch_seconds": round(time.perf_counter() - started_at, 2),
        "first_url": first_url,
        "hit_page_limit": pages >= MAX_ODATA_PAGES and paging_stop_reason == "max_page_limit",
        "paging_stop_reason": paging_stop_reason,
        "max_pages": MAX_ODATA_PAGES,
    }
    return rows_to_dataframe(kept_rows), metadata


@st.cache_data(ttl=API_CACHE_TTL_SECONDS, show_spinner=False)
def cached_fetch_report_data(
    username: str,
    password: str,
    token: str,
    auth_method: str,
    start_date: date,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    return fetch_report_data(username, password, token, auth_method, start_date)


# =============================================================================
# Multi-source wide OData helpers
# =============================================================================


def build_wide_odata_url(endpoint: str, start_date: date, datetime_column: str = "DateTime") -> str:
    start_text = start_date.strftime("%Y-%m-%d")
    params = {"$filter": f"{datetime_column} gt DateTime'{start_text}'"}
    return f"{endpoint}?{urlencode(params)}"


def fetch_wide_odata_source(
    source_key: str,
    username: str,
    password: str,
    token: str,
    auth_method: str,
    start_date: date,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    config = SOURCE_CONFIGS[source_key]
    endpoint = str(config["endpoint"])
    datetime_column = str(config.get("datetime_candidates", ["DateTime"])[0])
    next_url = build_wide_odata_url(endpoint, start_date, datetime_column)
    first_url = next_url
    seen_urls: set[str] = set()
    rows: list[dict[str, Any]] = []
    pages = 0
    total_bytes = 0
    consecutive_empty_pages = 0
    paging_stop_reason = "max_page_limit"
    started_at = time.perf_counter()
    auth = request_auth(username, password, auth_method)
    headers = request_headers(token, auth_method)

    with requests.Session() as session:
        session.headers.update(headers)
        for _ in range(MAX_ODATA_PAGES):
            if next_url in seen_urls:
                paging_stop_reason = "repeated_current_url"
                break
            seen_urls.add(next_url)
            response = request_with_retry(session, next_url, auth=auth, timeout=90)
            total_bytes += len(response.content)
            response.raise_for_status()
            pages += 1
            page_rows, next_link = extract_odata_page(response.json())
            rows.extend(page_rows)
            consecutive_empty_pages = consecutive_empty_pages + 1 if len(page_rows) == 0 else 0
            should_continue, resolved_next_url, stop_reason = should_continue_odata_paging(
                current_url=next_url,
                next_link=next_link,
                seen_urls=seen_urls,
                consecutive_empty_pages=consecutive_empty_pages,
            )
            if not should_continue:
                paging_stop_reason = stop_reason
                break
            next_url = resolved_next_url or next_url

    df = pd.DataFrame(rows)
    if "__metadata" in df.columns:
        df = df.drop(columns=["__metadata"])

    loaded_at_utc = datetime.now(timezone.utc)
    metadata = {
        "source": config["label"],
        "endpoint": endpoint,
        "loaded_at_utc": loaded_at_utc.strftime("%d-%m-%Y %H:%M:%S UTC"),
        "loaded_at_local": local_time_label(loaded_at_utc),
        "loaded_start_date": start_date.isoformat(),
        "rows": int(len(df)),
        "columns": int(len(df.columns)),
        "pages": pages,
        "downloaded_mb": round(total_bytes / 1024 / 1024, 2),
        "fetch_seconds": round(time.perf_counter() - started_at, 2),
        "first_url": first_url,
        "hit_page_limit": pages >= MAX_ODATA_PAGES and paging_stop_reason == "max_page_limit",
        "paging_stop_reason": paging_stop_reason,
        "max_pages": MAX_ODATA_PAGES,
    }
    return df, metadata


@st.cache_data(ttl=API_CACHE_TTL_SECONDS, show_spinner=False)
def cached_fetch_wide_odata_source(
    source_key: str,
    username: str,
    password: str,
    token: str,
    auth_method: str,
    start_date: date,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    return fetch_wide_odata_source(source_key, username, password, token, auth_method, start_date)


def source_signature(source_key: str, username: str, auth_method: str, start_date: date) -> dict[str, Any]:
    config = SOURCE_CONFIGS[source_key]
    return {
        "source": source_key,
        "endpoint": str(config["endpoint"]),
        "username_hash": sha256(username.encode("utf-8")).hexdigest()[:12],
        "auth_method": auth_method.lower(),
        "start_date": start_date.isoformat(),
    }


def save_source_snapshot(source_key: str, df: pd.DataFrame, metadata: dict[str, Any], signature: dict[str, Any]) -> None:
    try:
        config = SOURCE_CONFIGS[source_key]
        SNAPSHOT_DIR.mkdir(parents=True, exist_ok=True)
        df.to_parquet(config["snapshot_file"], index=False)
        payload = {
            "metadata": metadata,
            "signature": signature,
            "saved_at_utc": datetime.now(timezone.utc).strftime("%d-%m-%Y %H:%M:%S UTC"),
        }
        Path(config["metadata_file"]).write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")
    except Exception:
        return


def load_source_snapshot(
    source_key: str,
    requested_signature: dict[str, Any],
    requested_start_date: date,
) -> tuple[pd.DataFrame, dict[str, Any], dict[str, Any]] | None:
    try:
        config = SOURCE_CONFIGS[source_key]
        snapshot_file = Path(config["snapshot_file"])
        metadata_file = Path(config["metadata_file"])
        if not snapshot_file.is_file() or not metadata_file.is_file():
            return None
        payload = json.loads(metadata_file.read_text(encoding="utf-8"))
        metadata = payload.get("metadata") or {}
        signature = payload.get("signature") or {}
        if not raw_data_covers_request(signature, metadata, requested_signature, requested_start_date):
            return None
        df = pd.read_parquet(snapshot_file)
        if not isinstance(df, pd.DataFrame):
            return None
        # Reject broken placeholder snapshots that can be created after an interrupted warmup.
        # If API metadata says rows exist, a one-column NoData parquet is not a usable source snapshot.
        if list(df.columns) == ["NoData"] and int(metadata.get("rows", 0) or 0) > 0:
            return None
        metadata = metadata.copy()
        metadata["loaded_from_snapshot"] = True
        metadata.setdefault("snapshot_saved_at_utc", payload.get("saved_at_utc", "-"))
        return df, metadata, signature
    except Exception:
        return None


def parse_wide_source_datetimes(df: pd.DataFrame) -> pd.DataFrame:
    parsed_df = df.copy()
    for column in parsed_df.columns:
        if "date" in str(column).lower() or "time" in str(column).lower():
            parsed = parse_datetime_series(parsed_df[column])
            if parsed.notna().any():
                parsed_df[column] = parsed
    return parsed_df


def detect_datetime_column(df: pd.DataFrame, candidates: list[str]) -> str | None:
    for column in candidates:
        if column in df.columns:
            return column
    for column in df.columns:
        lower = str(column).lower()
        if "datetime" in lower or lower in {"date", "timestamp"}:
            return column
    return None


def filter_wide_source_data(
    df: pd.DataFrame,
    source_key: str,
    selected_vessels: list[str],
    selected_start: date,
    selected_end: date,
) -> pd.DataFrame:
    if df.empty:
        return df
    filtered = parse_wide_source_datetimes(df)
    if "ShipName" in filtered.columns and selected_vessels:
        filtered = filtered[match_selected_vessels(filtered["ShipName"], selected_vessels)].copy()
    datetime_column = detect_datetime_column(filtered, list(SOURCE_CONFIGS[source_key].get("datetime_candidates", [])))
    if datetime_column and datetime_column in filtered.columns:
        values = pd.to_datetime(filtered[datetime_column], errors="coerce", utc=True)
        start_timestamp = pd.Timestamp(selected_start, tz="UTC")
        end_timestamp = pd.Timestamp(selected_end + timedelta(days=1), tz="UTC")
        filtered = filtered[values.ge(start_timestamp) & values.lt(end_timestamp)].copy()
    return filtered


def load_or_fetch_source(
    source_key: str,
    username: str,
    password: str,
    token: str,
    auth_method: str,
    start_date: date,
    refresh: bool,
    auto_fetch: bool = False,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    sig = source_signature(source_key, username, auth_method, start_date)
    state_df_key = f"loaded_{source_key}_df"
    state_meta_key = f"loaded_{source_key}_metadata"
    state_sig_key = f"loaded_{source_key}_signature"
    df = st.session_state.get(state_df_key)
    metadata = st.session_state.get(state_meta_key)
    current_signature = st.session_state.get(state_sig_key)

    needs_load = (
        refresh
        or not isinstance(df, pd.DataFrame)
        or not isinstance(metadata, dict)
        or not raw_data_covers_request(current_signature, metadata, sig, start_date)
    )

    if needs_load and not refresh:
        snapshot = load_source_snapshot(source_key, sig, start_date)
        if snapshot is not None:
            df, metadata, snapshot_sig = snapshot
            st.session_state[state_df_key] = df
            st.session_state[state_meta_key] = metadata
            st.session_state[state_sig_key] = snapshot_sig
            needs_load = False

    if needs_load and not (refresh or auto_fetch):
        config = SOURCE_CONFIGS[source_key]
        empty_metadata = {
            "source": config["label"],
            "endpoint": str(config["endpoint"]),
            "loaded_at_utc": "-",
            "loaded_at_local": "No stored snapshot yet",
            "loaded_from_snapshot": False,
            "rows": 0,
            "columns": 0,
            "pages": 0,
            "first_url": "-",
            "needs_warmup": True,
        }
        return pd.DataFrame(), empty_metadata

    if needs_load:
        if refresh:
            cached_fetch_wide_odata_source.clear()
        df, metadata = fetch_wide_odata_source(source_key, username, password, token, auth_method, start_date)
        save_source_snapshot(source_key, df, metadata, sig)
        st.session_state[state_df_key] = df
        st.session_state[state_meta_key] = metadata
        st.session_state[state_sig_key] = sig

    return st.session_state[state_df_key], st.session_state[state_meta_key]


def render_wide_source_tab(source_label: str, df: pd.DataFrame, metadata: dict[str, Any], source_key: str, selected_vessels: list[str], selected_start: date, selected_end: date) -> pd.DataFrame:
    st.markdown(f'<div class="section-title">{escape(source_label)} Dataset</div>', unsafe_allow_html=True)
    render_api_load_caption(metadata)
    if list(df.columns) == ["NoData"] and int(metadata.get("rows", 0) or 0) > 0:
        st.error(
            f"{source_label} snapshot metadata shows {int(metadata.get('rows', 0)):,} API rows, "
            "but the stored parquet contains only a placeholder column. "
            f"Run the {source_key} warmup again with the latest app version."
        )
        return pd.DataFrame()

    filtered_df = filter_wide_source_data(df, source_key, selected_vessels, selected_start, selected_end)
    render_metric_cards(
        [
            ("Rows", f"{len(filtered_df):,}", "table_eye"),
            ("Columns", f"{len(filtered_df.columns):,}", "checked_columns"),
            ("API Rows", f"{metadata.get('rows', len(df)):,}", "database_rows"),
            ("API Pages", f"{metadata.get('pages', 0):,}", "numeric"),
        ]
    )

    default_columns = [c for c in ["ShipName", "DateTime", "State", "StateName", "GPSSpeed", "LogSpeed", "MEConsumed", "ShaftPower"] if c in filtered_df.columns]
    if not default_columns:
        default_columns = list(filtered_df.columns[: min(12, len(filtered_df.columns))])
    selected_columns = st.multiselect(
        f"{source_label} columns to preview/export",
        options=list(filtered_df.columns),
        default=default_columns,
        key=f"{source_key}_preview_columns",
    )
    if not selected_columns:
        selected_columns = default_columns
    output = filtered_df[selected_columns].copy() if selected_columns else filtered_df.copy()
    render_preview_table(output)
    if len(output) > TABLE_PREVIEW_ROW_LIMIT:
        st.caption(f"Showing first {TABLE_PREVIEW_ROW_LIMIT:,} of {len(output):,} rows. Export includes all filtered rows/columns selected above.")
    return output


def dataframe_memory_mb(df: Any) -> float:
    if not isinstance(df, pd.DataFrame):
        return 0.0
    try:
        return float(df.memory_usage(deep=True).sum()) / 1024 / 1024
    except Exception:
        return 0.0


def clear_wide_source_state(source_key: str) -> None:
    """Release wide-source DataFrames from this browser session.

    Snapshots remain on disk, so reopening a source reloads from Parquet rather
    than calling the API. This is the main Streamlit Cloud memory safeguard.
    """
    for suffix in ["df", "metadata", "signature"]:
        st.session_state.pop(f"loaded_{source_key}_{suffix}", None)


def clear_inactive_wide_sources(active_sources: set[str]) -> None:
    for source_key in ["reportpivots", "shippivots"]:
        if source_key not in active_sources:
            clear_wide_source_state(source_key)
    gc.collect()


def clear_stale_export_bytes(current_signature: str | None = None) -> None:
    """Remove large Excel byte buffers when they no longer match the current view."""
    if current_signature is None or st.session_state.get("atlas_export_signature") != current_signature:
        st.session_state.pop("atlas_export_bytes", None)
    if current_signature is None or st.session_state.get("atlas_multisource_export_signature") != current_signature:
        st.session_state.pop("atlas_multisource_export_bytes", None)
    gc.collect()


def current_memory_audit_rows(extra_frames: dict[str, pd.DataFrame] | None = None) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for key, value in st.session_state.items():
        if isinstance(value, pd.DataFrame):
            rows.append({"Object": f"session_state.{key}", "Rows": len(value), "Columns": len(value.columns), "Memory MB": round(dataframe_memory_mb(value), 2)})
        elif isinstance(value, (bytes, bytearray)):
            rows.append({"Object": f"session_state.{key}", "Rows": "-", "Columns": "-", "Memory MB": round(len(value) / 1024 / 1024, 2)})
    for name, frame in (extra_frames or {}).items():
        if isinstance(frame, pd.DataFrame):
            rows.append({"Object": name, "Rows": len(frame), "Columns": len(frame.columns), "Memory MB": round(dataframe_memory_mb(frame), 2)})
    if not rows:
        return pd.DataFrame(columns=["Object", "Rows", "Columns", "Memory MB"])
    return pd.DataFrame(rows).sort_values("Memory MB", ascending=False)


def wide_source_selected_columns(source_key: str, filtered_df: pd.DataFrame) -> list[str]:
    if filtered_df.empty:
        return []
    default_columns = [
        c for c in ["ShipName", "DateTime", "State", "StateName", "GPSSpeed", "LogSpeed", "MEConsumed", "ShaftPower"]
        if c in filtered_df.columns
    ]
    if not default_columns:
        default_columns = list(filtered_df.columns[: min(12, len(filtered_df.columns))])
    previous = st.session_state.get(f"{source_key}_preview_columns", default_columns)
    if not isinstance(previous, list):
        previous = default_columns
    selected_columns = [column for column in previous if column in filtered_df.columns]
    return selected_columns or default_columns


def load_wide_source_for_view(
    source_key: str,
    username: str,
    password: str,
    token: str,
    auth_method: str,
    start_date: date,
    refresh: bool,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    """Load one wide source only when the user opens it.

    This keeps ReportPivots and ShipPivots out of memory during normal Custom
    Analytics use. If no snapshot exists, the UI instructs the user to run the
    per-source warmup rather than live-loading a huge API in the app session.
    """
    return load_or_fetch_source(
        source_key,
        username,
        password,
        token,
        auth_method,
        start_date,
        refresh=refresh,
        auto_fetch=False,
    )


def build_wide_source_output_for_export(
    source_key: str,
    source_df: pd.DataFrame,
    selected_vessels: list[str],
    selected_start: date,
    selected_end: date,
) -> pd.DataFrame:
    filtered_df = filter_wide_source_data(source_df, source_key, selected_vessels, selected_start, selected_end)
    selected_columns = wide_source_selected_columns(source_key, filtered_df)
    if not selected_columns:
        return filtered_df.copy()
    return filtered_df[selected_columns].copy()


def to_multisource_excel_bytes(
    reportdata_df: pd.DataFrame,
    reportdata_summary_df: pd.DataFrame | None,
    reportpivots_df: pd.DataFrame,
    shippivots_df: pd.DataFrame,
) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        write_table_sheet(writer, reportdata_df, "ReportData Clean", "ReportDataClean")
        if reportdata_summary_df is not None and not reportdata_summary_df.empty:
            write_table_sheet(writer, reportdata_summary_df, "ReportData Summary", "ReportDataSummary")
        if reportpivots_df is not None and not reportpivots_df.empty:
            write_table_sheet(writer, reportpivots_df, "ReportPivots", "ReportPivotsData")
        if shippivots_df is not None and not shippivots_df.empty:
            write_table_sheet(writer, shippivots_df, "ShipPivots", "ShipPivotsData")
    return output.getvalue()


# =============================================================================
# Transform helpers
# =============================================================================


def normalize_text(value: Any) -> str:
    text = str(value).lower()
    return re.sub(r"[^a-z0-9]+", "", text)


def parse_datetime_series(series: pd.Series) -> pd.Series:
    parsed = pd.to_datetime(series, errors="coerce", utc=True)
    missing_mask = parsed.isna()

    if missing_mask.any():
        date_text = series.astype("string")
        dotnet_millis = date_text.str.extract(r"/Date\((-?\d+)").iloc[:, 0]
        dotnet_parsed = pd.to_datetime(
            pd.to_numeric(dotnet_millis, errors="coerce"),
            errors="coerce",
            unit="ms",
            utc=True,
        )
        parsed = parsed.mask(missing_mask, dotnet_parsed)

    return parsed


def parse_numeric_value(value: Any) -> Any:
    if pd.isna(value):
        return pd.NA
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return pd.NA

    duration_match = re.fullmatch(r"(-?\d+):([0-5]?\d)(?::([0-5]?\d))?", text)
    if duration_match:
        hours = int(duration_match.group(1))
        sign = -1 if hours < 0 else 1
        minutes = int(duration_match.group(2))
        seconds = int(duration_match.group(3) or 0)
        return sign * (abs(hours) + minutes / 60 + seconds / 3600)

    numeric_text = text.replace(" ", "")
    if re.fullmatch(r"-?\d+,\d+", numeric_text):
        numeric_text = numeric_text.replace(",", ".")
    else:
        numeric_text = numeric_text.replace(",", "")

    numeric_text = re.sub(r"[^0-9.\-]", "", numeric_text)
    if numeric_text in {"", "-", ".", "-."}:
        return pd.NA

    try:
        return float(numeric_text)
    except ValueError:
        return pd.NA


def parse_numeric_series(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series.map(parse_numeric_value), errors="coerce")


def match_selected_vessels(raw_ship_names: pd.Series, selected_vessels: list[str]) -> pd.Series:
    selected_keys = {normalize_text(vessel) for vessel in selected_vessels}
    return raw_ship_names.map(normalize_text).isin(selected_keys)


@st.cache_data(ttl=API_CACHE_TTL_SECONDS, show_spinner=False)
def cached_prepare_long_data(raw_df: pd.DataFrame) -> pd.DataFrame:
    missing_columns = sorted(set(SOURCE_COLUMNS).difference(raw_df.columns))
    if missing_columns:
        raise ValueError(f"Missing expected API columns: {', '.join(missing_columns)}")

    df = raw_df.copy()
    df["StartDateTimeGMT"] = parse_datetime_series(df["StartDateTimeGMT"])
    df["EndDateTimeGMT"] = parse_datetime_series(df["EndDateTimeGMT"])
    df["LapTime"] = parse_numeric_series(df["LapTime"])
    df["ParsedValue"] = parse_numeric_series(df["ReportedValue"])
    df = df[df["ValueDescription"].notna() & ~df["ReportType"].isin(EXCLUDED_REPORT_TYPES)].copy()
    return df


def available_variables(df: pd.DataFrame) -> list[str]:
    if df.empty or "ValueDescription" not in df.columns:
        return []
    return sorted(df["ValueDescription"].dropna().astype(str).unique().tolist(), key=str.casefold)


def available_report_types(df: pd.DataFrame) -> list[str]:
    if df.empty or "ReportType" not in df.columns:
        return []
    return sorted(df["ReportType"].dropna().astype(str).unique().tolist(), key=str.casefold)


def dataframe_date_window(df: pd.DataFrame) -> tuple[date, date]:
    if df.empty or "StartDateTimeGMT" not in df.columns:
        today = date.today()
        return today, today
    dates = pd.to_datetime(df["StartDateTimeGMT"], errors="coerce", utc=True).dt.date.dropna()
    if dates.empty:
        today = date.today()
        return today, today
    return max(dates.min(), API_FULL_START_DATE), min(dates.max(), date.today())


def filter_long_data(
    df: pd.DataFrame,
    selected_vessels: list[str],
    selected_report_types: list[str],
    selected_start: date,
    selected_end: date,
) -> pd.DataFrame:
    if df.empty:
        return df

    start_timestamp = pd.Timestamp(selected_start, tz="UTC")
    end_timestamp = pd.Timestamp(selected_end + timedelta(days=1), tz="UTC")
    start_values = df["StartDateTimeGMT"]
    if not pd.api.types.is_datetime64_any_dtype(start_values):
        start_values = pd.to_datetime(start_values, errors="coerce", utc=True)

    filtered = df[
        match_selected_vessels(df["ShipName"], selected_vessels)
        & start_values.ge(start_timestamp)
        & start_values.lt(end_timestamp)
    ].copy()

    if selected_report_types:
        filtered = filtered[filtered["ReportType"].astype("string").isin(selected_report_types)].copy()

    return filtered


def safe_divide(numerator: Any, denominator: Any) -> Any:
    numerator = pd.to_numeric(numerator, errors="coerce")
    denominator = pd.to_numeric(denominator, errors="coerce")
    denominator = denominator.mask(denominator == 0)
    return numerator / denominator


def sum_numeric_columns(df: pd.DataFrame, columns: list[str]) -> pd.Series:
    available_columns = [column for column in columns if column in df.columns]
    if not available_columns:
        return pd.Series(pd.NA, index=df.index, dtype="Float64")
    return df[available_columns].apply(pd.to_numeric, errors="coerce").sum(axis=1, min_count=1)


def calculation_alias_to_column() -> dict[str, str]:
    return {
        normalize_text(alias): column
        for column, aliases in DERIVED_VALUE_ALIASES.items()
        for alias in aliases
    }


def calculate_rob_consumption(
    df: pd.DataFrame,
    rob_column: str,
    received_column: str,
) -> pd.Series:
    """Calculate row-level consumption from ROB movement inside the current sample.

    The first report per vessel has no previous ROB inside the selected sample, so it
    remains blank. Negative values are treated as blank because they usually indicate
    correction/noise or a missing receipt entry rather than real consumption.
    """
    if rob_column not in df.columns:
        return pd.Series(pd.NA, index=df.index, dtype="Float64")

    work = df[["ShipName", "StartDateTimeGMT", "EndDateTimeGMT", rob_column]].copy()
    if received_column in df.columns:
        work[received_column] = pd.to_numeric(df[received_column], errors="coerce").fillna(0)
    else:
        work[received_column] = 0.0

    work[rob_column] = pd.to_numeric(work[rob_column], errors="coerce")
    work["_original_index"] = df.index
    work["_sort_date"] = pd.to_datetime(work["EndDateTimeGMT"], errors="coerce", utc=True)
    fallback_dates = pd.to_datetime(work["StartDateTimeGMT"], errors="coerce", utc=True)
    work["_sort_date"] = work["_sort_date"].fillna(fallback_dates)
    work = work.sort_values(["ShipName", "_sort_date", "_original_index"])
    previous_rob = work.groupby("ShipName", dropna=False)[rob_column].shift(1)
    consumption = previous_rob + work[received_column] - work[rob_column]
    consumption = consumption.where(consumption >= 0)
    consumption = consumption.where(previous_rob.notna() & work[rob_column].notna())
    result = pd.Series(pd.NA, index=df.index, dtype="Float64")
    result.loc[work["_original_index"]] = pd.to_numeric(consumption, errors="coerce").to_numpy()
    return result


def build_calculation_source_table(filtered_long_df: pd.DataFrame) -> pd.DataFrame:
    if filtered_long_df.empty:
        return pd.DataFrame(columns=PIVOT_IDENTITY_COLUMNS)

    alias_to_column = calculation_alias_to_column()
    source_long = filtered_long_df.copy()
    source_long["_value_key"] = source_long["ValueDescription"].map(normalize_text)
    source_long = source_long[
        source_long["_value_key"].isin(alias_to_column)
        & source_long["ParsedValue"].notna()
    ].copy()

    if source_long.empty:
        return filtered_long_df[PIVOT_IDENTITY_COLUMNS].drop_duplicates().copy()

    source_long["_canonical_column"] = source_long["_value_key"].map(alias_to_column)
    source_long["_source_order"] = range(len(source_long))
    source_long = source_long.sort_values("_source_order")
    source_long = source_long.drop_duplicates(
        [*PIVOT_IDENTITY_COLUMNS, "_canonical_column"],
        keep="last",
    )

    source_table = (
        source_long
        .pivot(index=PIVOT_IDENTITY_COLUMNS, columns="_canonical_column", values="ParsedValue")
        .reset_index()
    )
    source_table.columns.name = None
    return source_table


def add_performance_calculations(pivot_df: pd.DataFrame, source_table: pd.DataFrame) -> pd.DataFrame:
    if pivot_df.empty:
        for column in DERIVED_VARIABLES:
            if column not in pivot_df.columns:
                pivot_df[column] = pd.NA
        return pivot_df

    df = pivot_df.copy()
    if not source_table.empty:
        calculation_columns = [
            column for column in source_table.columns
            if column not in PIVOT_IDENTITY_COLUMNS and column not in df.columns
        ]
        if calculation_columns:
            df = df.merge(
                source_table[[*PIVOT_IDENTITY_COLUMNS, *calculation_columns]],
                on=PIVOT_IDENTITY_COLUMNS,
                how="left",
            )

    def numeric_column(column: str) -> pd.Series:
        if column not in df.columns:
            return pd.Series(pd.NA, index=df.index, dtype="Float64")
        return pd.to_numeric(df[column], errors="coerce")

    lap_time = numeric_column("LapTime")
    engine_distance = numeric_column("Engine Distance [nm]")
    distance_over_ground = numeric_column("Distance Over Ground [nm]")
    power = numeric_column("Power from Torque Meter [kW]")

    df["Calculated Slip"] = (1 - safe_divide(distance_over_ground, engine_distance)).round(3)

    me_sum = sum_numeric_columns(df, ME_FUEL_COLUMNS)
    official_me_total = numeric_column("Main Engine Total Consumed")
    df["ME Consumption Total"] = me_sum.fillna(official_me_total).round(3)

    dg_sum = sum_numeric_columns(df, DG_FUEL_COLUMNS)
    official_dg_total = numeric_column("Diesel Generator Total Consumed")
    df["DG Consumption Total"] = dg_sum.fillna(official_dg_total).round(3)

    aux_sum = sum_numeric_columns(df, AUXILIARY_FUEL_COLUMNS)
    official_aux_total = numeric_column("Auxiliary Engine Total Consumed")
    df["Auxiliary Engine Consumption Total"] = aux_sum.fillna(official_aux_total).round(3)

    df["Boiler Sum"] = sum_numeric_columns(df, BOILER_FUEL_COLUMNS).round(3)

    calculated_total = pd.concat(
        [
            pd.to_numeric(df["ME Consumption Total"], errors="coerce"),
            pd.to_numeric(df["DG Consumption Total"], errors="coerce"),
            pd.to_numeric(df["Auxiliary Engine Consumption Total"], errors="coerce"),
            pd.to_numeric(df["Boiler Sum"], errors="coerce"),
        ],
        axis=1,
    ).sum(axis=1, min_count=1)
    official_total = numeric_column("Total Fuel Consumed")
    df["Total Fuel Consumption"] = calculated_total.fillna(official_total).round(3)

    df["Consumption ME 24 Hours [MT]"] = safe_divide(df["ME Consumption Total"] * 24, lap_time).round(3)

    df["SFOC [gr/Kwh]"] = (
        safe_divide(df["Consumption ME 24 Hours [MT]"], power) / 0.000024
    ).round(3).fillna(0)

    # Oil consumption aggregation variables. These are row-level consumption
    # movements calculated inside the current selected sample, so Summary Analysis
    # can later Sum them by vessel/fleet/month/report type.
    df["MELO Consumption Total [ltr]"] = calculate_rob_consumption(
        df,
        "MELO ROB [ltr]",
        "MELO Received [ltr]",
    ).round(3)
    cylinder_oil_1_consumption = calculate_rob_consumption(
        df,
        "Cylinder Oil 1 ROB [ltr]",
        "Cylinder Oil 1 Received [ltr]",
    )
    cylinder_oil_2_consumption = calculate_rob_consumption(
        df,
        "Cylinder Oil 2 ROB [ltr]",
        "Cylinder Oil 2 Received [ltr]",
    )
    df["CYLO Consumption Total [ltr]"] = pd.concat(
        [cylinder_oil_1_consumption, cylinder_oil_2_consumption],
        axis=1,
    ).sum(axis=1, min_count=1).round(3)
    df["GELO Consumption Total [ltr]"] = calculate_rob_consumption(
        df,
        "GELO ROB [ltr]",
        "GELO Received [ltr]",
    ).round(3)
    df["Total DG Running Hours [hh:mm]"] = sum_numeric_columns(
        df,
        [
            "DG1 Running Hours [hh:mm]",
            "DG2 Running Hours [hh:mm]",
            "DG3 Running Hours [hh:mm]",
            "DG4 Running Hours [hh:mm]",
        ],
    ).round(3)

    return df


def render_lubricating_oil_workspace(filtered_long_df: pd.DataFrame) -> None:
    """Present AtlasFlow's existing ROB-based oil calculations as an oil workflow."""
    st.markdown('<div class="section-title">Lub Oil Analysis</div>', unsafe_allow_html=True)
    st.caption("Consumption is calculated from ROB movement and received quantities for the selected vessels and period.")

    oil_df = build_pivot_table(filtered_long_df, tuple())
    if oil_df.empty:
        st.info("No report data is available for the current vessel and date selection.")
        return

    def metric_text(value: Any, suffix: str) -> str:
        numeric = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
        return "-" if pd.isna(numeric) else f"{float(numeric):,.2f}{suffix}"

    consumption_columns = [
        "MELO Consumption Total [ltr]",
        "CYLO Consumption Total [ltr]",
        "GELO Consumption Total [ltr]",
    ]
    for column in consumption_columns:
        if column not in oil_df.columns:
            oil_df[column] = pd.NA

    steaming_hours = pd.to_numeric(oil_df.get("Steaming Time Since Last Report [hh:mm]"), errors="coerce").sum(min_count=1)
    dg_hours = pd.to_numeric(oil_df.get("Total DG Running Hours [hh:mm]"), errors="coerce").sum(min_count=1)
    melo_total = pd.to_numeric(oil_df["MELO Consumption Total [ltr]"], errors="coerce").sum(min_count=1)
    cylo_total = pd.to_numeric(oil_df["CYLO Consumption Total [ltr]"], errors="coerce").sum(min_count=1)
    gelo_total = pd.to_numeric(oil_df["GELO Consumption Total [ltr]"], errors="coerce").sum(min_count=1)
    melo_daily = safe_divide(pd.Series([melo_total * 24]), pd.Series([steaming_hours])).iloc[0]
    gelo_daily = safe_divide(pd.Series([gelo_total * 24]), pd.Series([dg_hours])).iloc[0]
    power_values = pd.to_numeric(oil_df["Power from Torque Meter [kW]"], errors="coerce") if "Power from Torque Meter [kW]" in oil_df.columns else pd.Series(pd.NA, index=oil_df.index, dtype="Float64")
    lap_values = pd.to_numeric(oil_df["LapTime"], errors="coerce") if "LapTime" in oil_df.columns else pd.Series(pd.NA, index=oil_df.index, dtype="Float64")
    torque_energy = (power_values * lap_values).sum(min_count=1)
    cylo_sloc = safe_divide(pd.Series([cylo_total * 0.93 * 1000]), pd.Series([torque_energy])).iloc[0]

    render_metric_cards([
        ("MELO Consumption", metric_text(melo_total, " ltr"), "oil_barrel"),
        ("Cylinder Oil Consumption", metric_text(cylo_total, " ltr"), "oil_barrel"),
        ("GELO Consumption", metric_text(gelo_total, " ltr"), "oil_barrel"),
    ], "atlas-metric-grid-three")
    render_metric_cards([
        ("MELO Consumption [ltr/running day]", metric_text(melo_daily, " ltr/day"), "oil_drop"),
        ("CYLO SLOC [g/kWh]", metric_text(cylo_sloc, " g/kWh"), "oil_drop"),
        ("GELO Consumption [ltr/DG running day]", metric_text(gelo_daily, " ltr/day"), "oil_drop"),
    ], "atlas-metric-grid-three")
    support_columns = st.columns(3)
    support_columns[0].caption(f"Total MELO: {metric_text(melo_total, ' ltr')} | Running hours: {metric_text(steaming_hours, ' hrs')}")
    support_columns[1].caption(f"Total CYLO: {metric_text(cylo_total, ' ltr')} | Torque energy: {metric_text(torque_energy, ' kWh')}")
    support_columns[2].caption(f"Total GELO: {metric_text(gelo_total, ' ltr')} | DG running hours: {metric_text(dg_hours, ' hrs')}")

    work = oil_df.copy()
    work["EndDateTimeGMT"] = pd.to_datetime(work.get("EndDateTimeGMT"), errors="coerce", utc=True)
    work["Month"] = work["EndDateTimeGMT"].dt.tz_localize(None).dt.to_period("M").astype("string")
    def numeric_oil_column(column: str) -> pd.Series:
        if column in work.columns:
            return pd.to_numeric(work[column], errors="coerce")
        return pd.Series(pd.NA, index=work.index, dtype="Float64")

    work["_atlas_steaming_hours"] = numeric_oil_column("Steaming Time Since Last Report [hh:mm]")
    work["_atlas_dg_hours"] = numeric_oil_column("Total DG Running Hours [hh:mm]")
    work["_atlas_torque_energy"] = (
        numeric_oil_column("Power from Torque Meter [kW]")
        * numeric_oil_column("LapTime")
    )

    kpi_columns = [
        "MELO [ltr/running day]",
        "CYLO SLOC [g/kWh]",
        "GELO [ltr/DG running day]",
    ]

    def summarize_oil_kpis(frame: pd.DataFrame, group_column: str) -> pd.DataFrame:
        totals = frame.groupby(group_column, as_index=False)[
            consumption_columns + ["_atlas_steaming_hours", "_atlas_dg_hours", "_atlas_torque_energy"]
        ].sum(min_count=1)
        totals["MELO [ltr/running day]"] = safe_divide(
            totals["MELO Consumption Total [ltr]"] * 24,
            totals["_atlas_steaming_hours"],
        )
        totals["CYLO SLOC [g/kWh]"] = safe_divide(
            totals["CYLO Consumption Total [ltr]"] * 0.93 * 1000,
            totals["_atlas_torque_energy"],
        )
        totals["GELO [ltr/DG running day]"] = safe_divide(
            totals["GELO Consumption Total [ltr]"] * 24,
            totals["_atlas_dg_hours"],
        )
        return totals.replace([float("inf"), float("-inf")], pd.NA)

    monthly_consumption = work.dropna(subset=["Month"]).groupby("Month", as_index=False)[consumption_columns].sum(min_count=1)
    vessel_consumption = work.dropna(subset=["ShipName"]).groupby("ShipName", as_index=False)[consumption_columns].sum(min_count=1)
    consumption_chart_labels = {
        "MELO Consumption Total [ltr]": "MELO Consumption [ltr]",
        "CYLO Consumption Total [ltr]": "CYLO Consumption [ltr]",
        "GELO Consumption Total [ltr]": "GELO Consumption [ltr]",
    }
    monthly_consumption_chart = monthly_consumption.rename(columns=consumption_chart_labels)
    vessel_consumption_chart = vessel_consumption.rename(columns=consumption_chart_labels)
    monthly = summarize_oil_kpis(work.dropna(subset=["Month"]), "Month")
    vessel_summary = summarize_oil_kpis(work.dropna(subset=["ShipName"]), "ShipName")

    def render_oil_kpi_chart(summary: pd.DataFrame, dimension: str, chart_type: str) -> None:
        chart_data = summary[[dimension] + kpi_columns].melt(
            id_vars=dimension,
            value_vars=kpi_columns,
            var_name="KPI",
            value_name="Value",
        ).dropna(subset=["Value"])
        if chart_data.empty:
            st.info("No KPI data is available for this view.")
            return

        colors = ["#FF2D2D", "#006BCE", "#78BAF0"]
        common_x = {
            "field": dimension,
            "type": "ordinal",
            "sort": None,
            "axis": {"title": None, "labelAngle": -90},
        }
        tooltip = [
            {"field": dimension, "type": "nominal", "title": dimension},
            {"field": "KPI", "type": "nominal", "title": "KPI"},
            {"field": "Value", "type": "quantitative", "format": ",.2f"},
        ]

        if chart_type == "line":
            # Independent, hidden y-scales keep the three units legible without
            # implying that ltr/day and g/kWh share a common measurement scale.
            layers = []
            for kpi, color in zip(kpi_columns, colors):
                layers.append({
                    "transform": [{"filter": f"datum.KPI === '{kpi}'"}],
                    "mark": {"type": "line", "point": True, "strokeWidth": 2.2, "color": color},
                    "encoding": {
                        "x": common_x,
                        "y": {"field": "Value", "type": "quantitative", "axis": None},
                        "tooltip": tooltip,
                    },
                })
            spec: dict[str, Any] = {
                "layer": layers,
                "resolve": {"scale": {"y": "independent"}},
                "config": {"view": {"stroke": "transparent"}},
            }
        else:
            layers = []
            for kpi, color in zip(kpi_columns, colors):
                layers.append({
                    "transform": [{"filter": f"datum.KPI === '{kpi}'"}],
                    "mark": {"type": "bar", "color": color},
                    "encoding": {
                        "x": common_x,
                        "xOffset": {
                            "field": "KPI",
                            "type": "nominal",
                            "scale": {"domain": kpi_columns},
                        },
                        "y": {"field": "Value", "type": "quantitative", "axis": None},
                        "tooltip": tooltip,
                    },
                })
            spec = {
                "layer": layers,
                "resolve": {"scale": {"y": "independent"}},
                "config": {"view": {"stroke": "transparent"}},
            }

        st.vega_lite_chart(chart_data, spec, use_container_width=True)
        st.markdown(
            '<div style="display:flex;justify-content:center;gap:1.35rem;flex-wrap:wrap;'
            'font-size:0.82rem;color:#4B617A;margin-top:-0.2rem;">'
            '<span><b style="color:#FF2D2D;">&#9679;</b> MELO [ltr/running day]</span>'
            '<span><b style="color:#006BCE;">&#9679;</b> CYLO SLOC [g/kWh]</span>'
            '<span><b style="color:#78BAF0;">&#9679;</b> GELO [ltr/DG running day]</span>'
            '</div>',
            unsafe_allow_html=True,
        )

    chart_left, chart_right = st.columns(2)
    with chart_left:
        st.markdown('<div class="section-title">Consumption by Month</div>', unsafe_allow_html=True)
        if not monthly_consumption.empty:
            st.line_chart(monthly_consumption_chart.set_index("Month"), use_container_width=True)
    with chart_right:
        st.markdown('<div class="section-title">Consumption by Vessel</div>', unsafe_allow_html=True)
        if not vessel_consumption.empty:
            st.bar_chart(vessel_consumption_chart.set_index("ShipName"), use_container_width=True)

    kpi_chart_left, kpi_chart_right = st.columns(2)
    with kpi_chart_left:
        st.markdown('<div class="section-title">KPI Trends by Month</div>', unsafe_allow_html=True)
        if not monthly.empty:
            render_oil_kpi_chart(monthly, "Month", "line")
    with kpi_chart_right:
        st.markdown('<div class="section-title">KPI Comparison by Vessel</div>', unsafe_allow_html=True)
        if not vessel_summary.empty:
            render_oil_kpi_chart(vessel_summary, "ShipName", "bar")

    st.markdown('<div class="section-title">Vessel Comparison</div>', unsafe_allow_html=True)
    st.dataframe(
        format_display_dataframe(vessel_consumption),
        use_container_width=True,
        hide_index=True,
    )

    detail_columns = [column for column in [
        "ShipName", "EndDateTimeGMT", "ReportId", "ReportType", "StateName",
        "MELO ROB [ltr]", "MELO Received [ltr]", "MELO Consumption Total [ltr]",
        "Cylinder Oil 1 ROB [ltr]", "Cylinder Oil 1 Received [ltr]",
        "Cylinder Oil 2 ROB [ltr]", "Cylinder Oil 2 Received [ltr]", "CYLO Consumption Total [ltr]",
        "GELO ROB [ltr]", "GELO Received [ltr]", "GELO Consumption Total [ltr]",
        "Steaming Time Since Last Report [hh:mm]", "Total DG Running Hours [hh:mm]",
    ] if column in work.columns]
    detail_df = work[detail_columns].sort_values("EndDateTimeGMT", ascending=False) if detail_columns else work
    st.markdown('<div class="section-title">Report-Level Oil Detail</div>', unsafe_allow_html=True)
    st.dataframe(format_display_dataframe(detail_df.head(TABLE_PREVIEW_ROW_LIMIT)), use_container_width=True, hide_index=True)

    export_signature = sha256(f"{len(detail_df)}|{detail_df.get('EndDateTimeGMT', pd.Series(dtype='object')).max()}".encode("utf-8")).hexdigest()
    if st.session_state.get("atlas_lub_oil_export_signature") != export_signature:
        st.session_state.pop("atlas_lub_oil_export_bytes", None)
    if st.button("Prepare lub oil Excel", type="primary"):
        with st.spinner("Preparing lub oil Excel..."):
            st.session_state["atlas_lub_oil_export_bytes"] = to_excel_bytes(detail_df)
            st.session_state["atlas_lub_oil_export_signature"] = export_signature
    if st.session_state.get("atlas_lub_oil_export_signature") == export_signature and "atlas_lub_oil_export_bytes" in st.session_state:
        st.download_button("Download lub oil Excel", st.session_state["atlas_lub_oil_export_bytes"], "atlasflow_lub_oil.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@st.cache_data(show_spinner=False)
def build_pivot_table(filtered_long_df: pd.DataFrame, selected_variables: tuple[str, ...]) -> pd.DataFrame:
    if filtered_long_df.empty:
        return pd.DataFrame(columns=PIVOT_IDENTITY_COLUMNS + list(selected_variables))

    calculation_source_table = build_calculation_source_table(filtered_long_df)

    api_selected_variables = [
        variable for variable in selected_variables
        if variable not in DERIVED_VARIABLES
    ]

    if not api_selected_variables:
        pivot_df = (
            filtered_long_df[PIVOT_IDENTITY_COLUMNS]
            .drop_duplicates()
            .sort_values(["ShipName", "EndDateTimeGMT"], ascending=[True, False])
            .reset_index(drop=True)
        )
    else:
        selected_long = filtered_long_df[
            filtered_long_df["ValueDescription"].astype("string").isin(api_selected_variables)
        ].copy()

        if selected_long.empty:
            pivot_df = filtered_long_df[PIVOT_IDENTITY_COLUMNS].drop_duplicates().copy()
            for variable in api_selected_variables:
                pivot_df[variable] = pd.NA
        else:
            selected_long["ValueDescription"] = selected_long["ValueDescription"].astype(str)
            selected_long["_source_order"] = range(len(selected_long))
            selected_long = selected_long.sort_values("_source_order")
            selected_long = selected_long.drop_duplicates(
                [*PIVOT_IDENTITY_COLUMNS, "ValueDescription"],
                keep="last",
            )

            pivot_df = (
                selected_long
                .pivot(index=PIVOT_IDENTITY_COLUMNS, columns="ValueDescription", values="ParsedValue")
                .reset_index()
            )
            pivot_df.columns.name = None

        for variable in api_selected_variables:
            if variable not in pivot_df.columns:
                pivot_df[variable] = pd.NA

    pivot_df = add_performance_calculations(pivot_df, calculation_source_table)

    for variable in selected_variables:
        if variable not in pivot_df.columns:
            pivot_df[variable] = pd.NA

    return pivot_df.sort_values(["ShipName", "EndDateTimeGMT"], ascending=[True, False]).reset_index(drop=True)


# =============================================================================
# Display/export helpers
# =============================================================================


def format_display_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Return a display-safe copy without inserting text into nullable numeric columns.

    Some selected variables can be completely empty for the active vessel/period.
    Pandas keeps those columns as nullable numeric dtypes, where a global
    ``fillna("-")`` raises a TypeError. Convert each display column to strings
    first, then replace missing values column by column.
    """
    display_df = df.copy()
    identity_columns = {
        "ReportId",
        "ShipName",
        "ReportType",
        "StartDateTimeGMT",
        "EndDateTimeGMT",
        "StateName",
    }

    for column in display_df.columns:
        series = display_df[column]

        if column in {"StartDateTimeGMT", "EndDateTimeGMT"} or pd.api.types.is_datetime64_any_dtype(series):
            formatted = pd.to_datetime(series, errors="coerce").dt.strftime(DISPLAY_DATETIME_FORMAT)
            display_df[column] = formatted.astype("string").fillna("-")
            continue

        if column not in identity_columns:
            numeric_values = pd.to_numeric(series, errors="coerce")
            if numeric_values.notna().any():
                display_df[column] = numeric_values.map(
                    lambda value: "-" if pd.isna(value) else f"{value:,.3f}"
                ).astype("string")
                continue

        display_df[column] = series.astype("string").fillna("-")

    return display_df


METRIC_ICON_SVGS = {
    "table_eye": '<svg viewBox="0 0 24 24" aria-hidden="true"><rect x="3" y="5" width="18" height="14" rx="1.8"></rect><path d="M3 10h18M8 5v14M16 5v14"></path><path d="M8.4 12.3c1.1-1.3 2.3-1.9 3.6-1.9s2.5.6 3.6 1.9c-1.1 1.3-2.3 1.9-3.6 1.9s-2.5-.6-3.6-1.9Z"></path><circle cx="12" cy="12.3" r="1.1"></circle></svg>',
    "checked_columns": '<svg viewBox="0 0 24 24" aria-hidden="true"><rect x="4" y="4" width="16" height="16" rx="2"></rect><path d="M9 4v16M15 4v16M7.2 12.2l1.9 1.9 3.8-4.4"></path></svg>',
    "database_rows": '<svg viewBox="0 0 24 24" aria-hidden="true"><ellipse cx="12" cy="5" rx="8" ry="3"></ellipse><path d="M4 5v6c0 1.7 3.6 3 8 3s8-1.3 8-3V5"></path><path d="M4 11v6c0 1.7 3.6 3 8 3s8-1.3 8-3v-6"></path><path d="M8 9h8M8 15h8"></path></svg>',
    "columns_plus": '<svg viewBox="0 0 24 24" aria-hidden="true"><rect x="3" y="4" width="13" height="16" rx="2"></rect><path d="M7.3 4v16M11.7 4v16M18 9v8M14 13h8"></path></svg>',
    "vessel": '<svg viewBox="0 0 24 24" aria-hidden="true"><path d="M4 14h16l-2.2 4.1a3 3 0 0 1-2.6 1.6H8.8a3 3 0 0 1-2.6-1.6L4 14Z"></path><path d="M8 14V9h8v5M12 9V4l3 2.2V9M3.5 21c1.2 0 1.8-.8 3-.8s1.8.8 3 .8 1.8-.8 3-.8 1.8.8 3 .8 1.8-.8 3-.8 1.8.8 3 .8"></path></svg>',
    "api_sources": '<svg viewBox="0 0 24 24" aria-hidden="true"><ellipse cx="12" cy="5" rx="5.5" ry="2.2"></ellipse><path d="M6.5 5v5.2c0 1.2 2.5 2.2 5.5 2.2s5.5-1 5.5-2.2V5M6.5 10.2v4.1c0 1.2 2.5 2.2 5.5 2.2s5.5-1 5.5-2.2v-4.1"></path><path d="M4 18.5h3M17 18.5h3M6 16.5v4M18 16.5v4"></path></svg>',
    "report_rows": '<svg viewBox="0 0 24 24" aria-hidden="true"><path d="M7 3h8l4 4v14H7Z"></path><path d="M15 3v5h4M10 12h6M10 16h6M4 8h3M4 12h3M4 16h3"></path></svg>',
    "time_series_rows": '<svg viewBox="0 0 24 24" aria-hidden="true"><rect x="3" y="5" width="18" height="14" rx="2"></rect><path d="M3 10h18M8 5v14M13 13.5l2.5 1.5V11"></path><circle cx="13" cy="13" r="3.5"></circle></svg>',
    "cargo_weight": '<svg viewBox="0 0 24 24" aria-hidden="true"><path d="M9 7a3 3 0 0 1 6 0"></path><path d="M7 7h10l2 4v7a2 2 0 0 1-2 2H7a2 2 0 0 1-2-2v-7l2-4Z"></path><path d="M8.5 13h7M10 16h4"></path></svg>',
    "cargo_teu": '<svg viewBox="0 0 24 24" aria-hidden="true"><rect x="3.5" y="7" width="17" height="10" rx="1.6"></rect><path d="M7 7v10M10.5 7v10M14 7v10M17.5 7v10M6 20h12M8 4h8"></path></svg>',
    "voyage_duration": '<svg viewBox="0 0 24 24" aria-hidden="true"><circle cx="12" cy="12" r="8"></circle><path d="M12 7v5l3.5 2M5.5 5.5l1.8 1.8M18.5 5.5l-1.8 1.8"></path></svg>',
    "report_count": '<svg viewBox="0 0 24 24" aria-hidden="true"><path d="M7 3h8l4 4v14H7Z"></path><path d="M15 3v5h4M10 12h6M10 16h6M4 7h3M4 11h3M4 15h3"></path></svg>',
    "fuel_total": '<svg viewBox="0 0 24 24" aria-hidden="true"><path d="M7 4h7a2 2 0 0 1 2 2v14H5V6a2 2 0 0 1 2-2Z"></path><path d="M8.5 8h4.5M16 7h2l2 3v7a2 2 0 0 1-2 2h-2M20 10h-2"></path></svg>',
    "fuel_grade": '<svg viewBox="0 0 24 24" aria-hidden="true"><path d="M12 3s6 6.1 6 10.4a6 6 0 0 1-12 0C6 9.1 12 3 12 3Z"></path><path d="M9 13.2h6M9.8 16h4.4"></path></svg>',
    "oil_barrel": '<svg viewBox="0 0 24 24" aria-hidden="true"><path d="M7 3.5h10l1.6 3.2-1.1 2.8 1.1 2.8-1.6 3.2H7l-1.6-3.2 1.1-2.8-1.1-2.8L7 3.5Z"></path><path d="M5.5 6.7h13M6.4 12h11.2M5.5 17.3h13M9.2 3.7v15.8M14.8 3.7v15.8"></path></svg>',
    "oil_drop": '<svg viewBox="0 0 24 24" aria-hidden="true"><path d="M12 3s6 6.1 6 10.4a6 6 0 0 1-12 0C6 9.1 12 3 12 3Z"></path></svg>',
    "average": '<svg viewBox="0 0 24 24" aria-hidden="true"><path d="M4 19V5"></path><path d="M7 17V9M12 17V6M17 17v-5"></path><path d="M4 12h16"></path></svg>',
    "total": '<svg viewBox="0 0 24 24" aria-hidden="true"><path d="M17.5 5H7l6 7-6 7h10.5"></path></svg>',
    "numeric": '<svg viewBox="0 0 24 24" aria-hidden="true"><path d="M4 8h3v8M4 16h6M13 8h4l-4 8h4M20 8v8"></path></svg>',
    "missing": '<svg viewBox="0 0 24 24" aria-hidden="true"><rect x="4" y="4" width="16" height="16" rx="2" stroke-dasharray="3 2"></rect><path d="M10 9.5a2.2 2.2 0 1 1 3.3 1.9c-.8.5-1.3 1-1.3 2.1"></path><path d="M12 17h.01"></path></svg>',
    # Backwards-compatible names used by older cards.
    "table": '<svg viewBox="0 0 24 24" aria-hidden="true"><rect x="3" y="3" width="18" height="18" rx="1.8"></rect><path d="M3 9h18M3 15h18M9 3v18M15 3v18"></path></svg>',
    "list": '<svg viewBox="0 0 24 24" aria-hidden="true"><path d="M8 6h13M8 12h13M8 18h9"></path><path d="M3.5 6h.01M3.5 12h.01M3.5 18h.01"></path></svg>',
    "database": '<svg viewBox="0 0 24 24" aria-hidden="true"><ellipse cx="12" cy="5" rx="8" ry="3"></ellipse><path d="M4 5v6c0 1.7 3.6 3 8 3s8-1.3 8-3V5"></path><path d="M4 11v6c0 1.7 3.6 3 8 3s8-1.3 8-3v-6"></path></svg>',
    "nodes": '<svg viewBox="0 0 24 24" aria-hidden="true"><circle cx="12" cy="4" r="2.5"></circle><circle cx="5" cy="19" r="2.5"></circle><circle cx="19" cy="19" r="2.5"></circle><path d="M10.8 6.2 6.2 16.8M13.2 6.2l4.6 10.6M7.5 19h9"></path></svg>',
}

def render_metric_cards(cards: list[tuple[str, str, str]], grid_class: str = "") -> None:
    card_html = []
    for label, value, icon_name in cards:
        icon_svg = METRIC_ICON_SVGS.get(icon_name, METRIC_ICON_SVGS["table"])
        card_html.append(
            f'<div class="atlas-metric-card"><div class="atlas-metric-icon">{icon_svg}</div>'
            f'<div><div class="atlas-metric-label">{escape(label)}</div>'
            f'<div class="atlas-metric-value">{escape(value)}</div></div></div>'
        )
    st.markdown(f'<div class="atlas-metric-grid {escape(grid_class)}">{"".join(card_html)}</div>', unsafe_allow_html=True)


def render_preview_table(df: pd.DataFrame, row_limit: int = TABLE_PREVIEW_ROW_LIMIT) -> None:
    preview_df = format_display_dataframe(df.head(row_limit))
    if preview_df.empty:
        st.markdown('<div class="atlas-table-empty">No rows to display.</div>', unsafe_allow_html=True)
        return

    columns = [str(column) for column in preview_df.columns]
    header_html = "".join(f"<th>{escape(column)}</th>" for column in columns)
    rows_html: list[str] = []
    for row in preview_df.itertuples(index=False, name=None):
        cell_html = "".join(("<td>" + escape(str(value)) + "</td>") for value in row)
        rows_html.append(f"<tr>{cell_html}</tr>")

    table_html = (
        '<div class="atlas-table-frame"><div class="atlas-table-scroll">'
        f'<table class="atlas-preview-table"><thead><tr>{header_html}</tr></thead>'
        f'<tbody>{"".join(rows_html)}</tbody></table></div></div>'
    )
    st.markdown(table_html, unsafe_allow_html=True)


def make_unique_excel_columns(columns: list[Any]) -> list[str]:
    """Return Excel-table-safe, unique column names while preserving readable labels."""
    safe_columns: list[str] = []
    seen: dict[str, int] = {}

    for position, column in enumerate(columns, start=1):
        label = str(column).strip() if column is not None else ""
        label = re.sub(r"[\x00-\x1f]", "", label)
        if not label or label.lower() in {"nan", "nat", "none"}:
            label = f"Column {position}"
        label = label[:240]

        key = label.casefold()
        count = seen.get(key, 0)
        if count:
            suffix = f"_{count + 1}"
            label = f"{label[:240 - len(suffix)]}{suffix}"
            key = label.casefold()
        seen[key] = count + 1
        safe_columns.append(label)

    return safe_columns


def make_excel_safe_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    safe_df = df.copy()
    safe_df.columns = make_unique_excel_columns(list(safe_df.columns))
    safe_df = safe_df.replace([float("inf"), float("-inf")], pd.NA)
    for column in safe_df.columns:
        if pd.api.types.is_datetime64_any_dtype(safe_df[column]):
            safe_df[column] = pd.to_datetime(safe_df[column], errors="coerce").dt.tz_localize(None)
            continue

        # Wide API exports often hold numbers as strings. Promote only columns whose
        # meaningful values are all numeric, so names, UUIDs, dates, and mixed data stay text.
        column_label = str(column).casefold()
        if pd.api.types.is_numeric_dtype(safe_df[column]) or pd.api.types.is_bool_dtype(safe_df[column]):
            continue
        if any(token in column_label for token in ("date", "time", "timestamp")):
            continue

        text_values = safe_df[column].astype("string").str.strip()
        text_values = text_values.mask(text_values.str.casefold().isin({"", "-", "nan", "nat", "none", "null", "n/a"}))
        numeric_values = pd.to_numeric(text_values.str.replace(",", "", regex=False), errors="coerce")
        non_blank_values = text_values.notna()
        if non_blank_values.any() and numeric_values[non_blank_values].notna().all():
            safe_df[column] = numeric_values
    return safe_df


def add_excel_table(worksheet: Any, table_name: str) -> None:
    if worksheet.max_row < 2 or worksheet.max_column < 1:
        return

    headers = [worksheet.cell(row=1, column=col).value for col in range(1, worksheet.max_column + 1)]
    if any(header is None or str(header).strip() == "" for header in headers):
        return
    if len({str(header).casefold() for header in headers}) != len(headers):
        return

    table_ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
    table = Table(displayName=table_name, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        # Built-in table style kept neutral; explicit teal formatting below
        # gives the export a stable AtlasFlow look in Excel.
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def numeric_excel_column_indexes(df: pd.DataFrame) -> set[int]:
    return {
        index
        for index, column in enumerate(df.columns, start=1)
        if pd.api.types.is_numeric_dtype(df[column]) and not pd.api.types.is_bool_dtype(df[column])
    }


def apply_teal_excel_table_format(worksheet: Any, numeric_columns: set[int]) -> None:
    """Apply AtlasFlow teal styling to exported Excel tables."""
    if worksheet.max_row < 1 or worksheet.max_column < 1:
        return

    header_fill = PatternFill(fill_type="solid", fgColor="006B68")
    even_fill = PatternFill(fill_type="solid", fgColor="EAF7F5")
    odd_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    border_side = Side(style="thin", color="B7DCD8")
    cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = cell_border

    for row_number in range(2, worksheet.max_row + 1):
        fill = even_fill if row_number % 2 == 0 else odd_fill
        for column_index, cell in enumerate(worksheet[row_number], start=1):
            cell.fill = fill
            cell.border = cell_border
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if column_index in numeric_columns and cell.value is not None:
                cell.number_format = "#,##0.00"


def excel_column_display_width(series: pd.Series, column_name: str, is_numeric: bool) -> int:
    if is_numeric:
        values = pd.to_numeric(series, errors="coerce").dropna()
        content_width = int(values.map(lambda value: len(f"{float(value):,.2f}")).max()) if not values.empty else 0
    else:
        values = series.dropna().astype(str)
        content_width = int(values.str.len().max()) if not values.empty else 0
    return min(max(max(len(str(column_name)), content_width) + 2, 12), 48)


def autofit_excel_columns(
    worksheet: Any,
    df: pd.DataFrame,
    numeric_columns: set[int],
    max_width: int = 48,
) -> None:
    for column_index, column_name in enumerate(df.columns, start=1):
        series = df.iloc[:, column_index - 1]
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(
            excel_column_display_width(series, str(column_name), column_index in numeric_columns),
            max_width,
        )


def write_table_sheet(writer: Any, df: pd.DataFrame, sheet_name: str, table_name: str) -> None:
    safe_df = make_excel_safe_dataframe(df)
    safe_df.to_excel(writer, index=False, sheet_name=sheet_name)
    worksheet = writer.sheets[sheet_name]
    worksheet.freeze_panes = "A2"
    numeric_columns = numeric_excel_column_indexes(safe_df)
    autofit_excel_columns(worksheet, safe_df, numeric_columns)
    add_excel_table(worksheet, table_name)
    apply_teal_excel_table_format(worksheet, numeric_columns)


def to_excel_bytes(clean_df: pd.DataFrame, pivot_analysis_df: pd.DataFrame | None = None) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        write_table_sheet(writer, clean_df, "Clean Dataset", "AtlasFlowCleanDataset")

        if pivot_analysis_df is not None and not pivot_analysis_df.empty:
            write_table_sheet(writer, pivot_analysis_df, "Summary Analysis", "AtlasFlowSummaryAnalysis")

    return output.getvalue()


def to_displayed_table_excel_bytes(display_df: pd.DataFrame, sheet_name: str = "Displayed Table") -> bytes:
    """Export only the table currently visible to the user.

    This avoids preparing multiple hidden sheets during a normal tab export and keeps
    memory usage lower on Streamlit Cloud.
    """
    safe_df = make_excel_safe_dataframe(display_df)
    output = BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter", datetime_format="dd/mm/yyyy hh:mm") as writer:
        safe_df.to_excel(writer, index=False, sheet_name=sheet_name[:31])
        workbook = writer.book
        worksheet = writer.sheets[sheet_name[:31]]
        header_format = workbook.add_format({
            "bg_color": "#006B68", "font_color": "#FFFFFF", "bold": True,
            "align": "center", "valign": "vcenter", "text_wrap": True,
            "border": 1, "border_color": "#B7DCD8",
        })
        stripe_format = workbook.add_format({"bg_color": "#EAF7F5"})
        numeric_format = workbook.add_format({"num_format": "#,##0.00"})
        numeric_columns = numeric_excel_column_indexes(safe_df)

        for column_index, column_name in enumerate(safe_df.columns):
            is_numeric = (column_index + 1) in numeric_columns
            worksheet.set_column(
                column_index,
                column_index,
                excel_column_display_width(safe_df.iloc[:, column_index], str(column_name), is_numeric),
                numeric_format if is_numeric else None,
            )

        if safe_df.empty:
            worksheet.set_row(0, None, header_format)
        else:
            worksheet.add_table(
                0,
                0,
                len(safe_df),
                len(safe_df.columns) - 1,
                {
                    "name": "AtlasFlowDisplayedTable",
                    "style": None,
                    "columns": [
                        {"header": str(column), "header_format": header_format}
                        for column in safe_df.columns
                    ],
                },
            )
            worksheet.conditional_format(
                1,
                0,
                len(safe_df),
                len(safe_df.columns) - 1,
                {"type": "formula", "criteria": "=MOD(ROW(),2)=0", "format": stripe_format},
            )
        worksheet.freeze_panes(1, 0)
    return output.getvalue()


def flatten_pivot_columns(df: pd.DataFrame) -> pd.DataFrame:
    flat_df = df.copy()
    if isinstance(flat_df.columns, pd.MultiIndex):
        flat_df.columns = [
            " | ".join(str(part) for part in column if str(part) not in {"", "nan", "NaT"})
            for column in flat_df.columns.to_flat_index()
        ]
    else:
        flat_df.columns = [str(column) for column in flat_df.columns]
    return flat_df


def build_summary_analysis(
    clean_df: pd.DataFrame,
    group_fields: list[str],
    value_fields: list[str],
    aggregation: str,
) -> pd.DataFrame:
    if clean_df.empty or not group_fields or not value_fields:
        return pd.DataFrame()

    valid_group_fields = [column for column in group_fields if column in clean_df.columns]
    valid_value_fields = [column for column in value_fields if column in clean_df.columns]
    if not valid_group_fields or not valid_value_fields:
        return pd.DataFrame()

    source_df = clean_df.copy()
    for column in valid_value_fields:
        source_df[column] = pd.to_numeric(source_df[column], errors="coerce")

    aggregation_map = {
        "Average": "mean",
        "Sum": "sum",
        "Count": "count",
        "Minimum": "min",
        "Maximum": "max",
        "Median": "median",
    }
    aggfunc = aggregation_map.get(aggregation, "mean")

    summary_df = (
        source_df
        .groupby(valid_group_fields, dropna=False, as_index=False)[valid_value_fields]
        .agg(aggfunc)
    )

    if aggregation != "Count":
        for column in valid_value_fields:
            summary_df[column] = pd.to_numeric(summary_df[column], errors="coerce").round(3)

    rename_map = {column: f"{aggregation} {column}" for column in valid_value_fields}
    return summary_df.rename(columns=rename_map)


def numeric_column_options(df: pd.DataFrame) -> list[str]:
    options: list[str] = []
    for column in df.columns:
        values = pd.to_numeric(df[column], errors="coerce")
        if values.notna().any():
            options.append(column)
    return options


def filter_digest(column: str) -> str:
    return sha256(column.encode("utf-8")).hexdigest()[:10]


def parse_optional_float(value: str) -> tuple[float | None, bool]:
    text = str(value or "").strip()
    if not text:
        return None, True
    normalized = text.replace(" ", "").replace(",", "")
    try:
        return float(normalized), True
    except ValueError:
        return None, False


def parse_optional_date(value: str) -> tuple[pd.Timestamp | None, bool]:
    text = str(value or "").strip()
    if not text:
        return None, True
    parsed = pd.to_datetime(text, dayfirst=True, errors="coerce", utc=True)
    if pd.isna(parsed):
        return None, False
    return parsed, True


def unique_display_values(series: pd.Series, limit: int = 500) -> list[str]:
    values = series.astype("string").fillna("(Blank)").drop_duplicates().tolist()
    values = sorted(values, key=lambda value: str(value).casefold())
    return values[:limit]


def is_numeric_like(series: pd.Series) -> bool:
    values = pd.to_numeric(series, errors="coerce")
    return values.notna().any()


def render_column_filters(df: pd.DataFrame, filter_columns: list[str]) -> list[dict[str, Any]]:
    specs: list[dict[str, Any]] = []
    for column in filter_columns:
        if column not in df.columns:
            continue

        st.caption(f"Filter: {column}")
        series = df[column]
        digest = filter_digest(column)

        if pd.api.types.is_datetime64_any_dtype(series):
            left, right = st.columns(2)
            from_text = left.text_input("From", key=f"atlas_filter_{digest}_from", placeholder="dd/mm/yyyy")
            to_text = right.text_input("To", key=f"atlas_filter_{digest}_to", placeholder="dd/mm/yyyy")
            from_value, from_ok = parse_optional_date(from_text)
            to_value, to_ok = parse_optional_date(to_text)
            if not from_ok or not to_ok:
                st.warning(f"{column}: enter dates as dd/mm/yyyy or yyyy-mm-dd.")
            specs.append({"column": column, "kind": "datetime", "from": from_value, "to": to_value})
            continue

        if is_numeric_like(series):
            values = pd.to_numeric(series, errors="coerce").dropna()
            if not values.empty:
                st.caption(f"Loaded range: {values.min():,.3f} to {values.max():,.3f}")
            left, right = st.columns(2)
            min_text = left.text_input("Min", key=f"atlas_filter_{digest}_min", placeholder="no minimum")
            max_text = right.text_input("Max", key=f"atlas_filter_{digest}_max", placeholder="no maximum")
            minimum, min_ok = parse_optional_float(min_text)
            maximum, max_ok = parse_optional_float(max_text)
            if not min_ok or not max_ok:
                st.warning(f"{column}: enter numeric Min/Max values only.")
            if minimum is not None and maximum is not None and minimum > maximum:
                minimum, maximum = maximum, minimum
            specs.append({"column": column, "kind": "numeric", "min": minimum, "max": maximum})
            continue

        values_key = f"atlas_filter_{digest}_values"
        selected_values = st.multiselect(
            "Values",
            options=unique_display_values(series),
            key=values_key,
            help="Leave blank to include all values for this column.",
        )
        specs.append({"column": column, "kind": "categorical", "values": selected_values})

    return specs


def apply_column_filters(df: pd.DataFrame, specs: list[dict[str, Any]]) -> pd.DataFrame:
    filtered = df.copy()
    for spec in specs:
        column = spec.get("column")
        if column not in filtered.columns:
            continue

        kind = spec.get("kind")
        if kind == "numeric":
            values = pd.to_numeric(filtered[column], errors="coerce")
            minimum = spec.get("min")
            maximum = spec.get("max")
            if minimum is not None:
                filtered = filtered[values >= minimum]
                values = pd.to_numeric(filtered[column], errors="coerce")
            if maximum is not None:
                filtered = filtered[values <= maximum]

        elif kind == "datetime":
            values = pd.to_datetime(filtered[column], errors="coerce", utc=True)
            from_value = spec.get("from")
            to_value = spec.get("to")
            if from_value is not None:
                filtered = filtered[values >= from_value]
                values = pd.to_datetime(filtered[column], errors="coerce", utc=True)
            if to_value is not None:
                filtered = filtered[values < (to_value + pd.Timedelta(days=1))]

        elif kind == "categorical":
            selected_values = spec.get("values") or []
            if selected_values:
                values = filtered[column].astype("string").fillna("(Blank)")
                filtered = filtered[values.isin(selected_values)]

    return filtered



# =============================================================================
# Descriptive statistics helpers
# =============================================================================


def dataframe_numeric_options(df: pd.DataFrame) -> list[str]:
    return [column for column in df.columns if pd.to_numeric(df[column], errors="coerce").notna().any()]


def dataframe_categorical_options(df: pd.DataFrame) -> list[str]:
    options: list[str] = []
    for column in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[column]):
            continue
        numeric_values = pd.to_numeric(df[column], errors="coerce")
        if numeric_values.notna().any():
            continue
        if df[column].astype("string").dropna().nunique() <= 250:
            options.append(column)
    return options


def detect_analysis_datetime_column(df: pd.DataFrame) -> str | None:
    preferred_columns = ["StartDateTimeGMT", "EndDateTimeGMT", "DateTime", "ReportDateTime", "Timestamp"]
    for column in preferred_columns:
        if column in df.columns and pd.to_datetime(df[column], errors="coerce", utc=True).notna().any():
            return column
    for column in df.columns:
        lower = str(column).lower()
        if ("date" in lower or "time" in lower) and pd.to_datetime(df[column], errors="coerce", utc=True).notna().any():
            return column
    return None


def build_descriptive_statistics(df: pd.DataFrame, metric_column: str) -> pd.DataFrame:
    values = pd.to_numeric(df[metric_column], errors="coerce")
    clean_values = values.dropna()
    if clean_values.empty:
        return pd.DataFrame()
    rows = [
        ("Rows", len(df)),
        ("Numeric values", int(clean_values.count())),
        ("Missing values", int(values.isna().sum())),
        ("Sum", clean_values.sum()),
        ("Mean", clean_values.mean()),
        ("Median", clean_values.median()),
        ("Std dev", clean_values.std()),
        ("Minimum", clean_values.min()),
        ("P10", clean_values.quantile(0.10)),
        ("P25", clean_values.quantile(0.25)),
        ("P75", clean_values.quantile(0.75)),
        ("P90", clean_values.quantile(0.90)),
        ("Maximum", clean_values.max()),
    ]
    return pd.DataFrame(rows, columns=["Statistic", "Value"])


def build_grouped_descriptive_statistics(df: pd.DataFrame, metric_column: str, group_column: str) -> pd.DataFrame:
    if group_column not in df.columns or metric_column not in df.columns:
        return pd.DataFrame()
    source = df[[group_column, metric_column]].copy()
    source[metric_column] = pd.to_numeric(source[metric_column], errors="coerce")
    source = source[source[metric_column].notna()].copy()
    if source.empty:
        return pd.DataFrame()
    grouped = (
        source
        .groupby(group_column, dropna=False)[metric_column]
        .agg(Count="count", Sum="sum", Mean="mean", Median="median", Minimum="min", Maximum="max")
        .reset_index()
        .sort_values("Sum", ascending=False)
    )
    for column in ["Sum", "Mean", "Median", "Minimum", "Maximum"]:
        grouped[column] = pd.to_numeric(grouped[column], errors="coerce").round(3)
    return grouped


def build_monthly_trend(df: pd.DataFrame, metric_column: str, datetime_column: str) -> pd.DataFrame:
    source = df[[datetime_column, metric_column]].copy()
    source[datetime_column] = pd.to_datetime(source[datetime_column], errors="coerce", utc=True)
    source[metric_column] = pd.to_numeric(source[metric_column], errors="coerce")
    source = source[source[datetime_column].notna() & source[metric_column].notna()].copy()
    if source.empty:
        return pd.DataFrame()
    source["Month"] = source[datetime_column].dt.to_period("M").astype(str)
    trend = (
        source
        .groupby("Month", as_index=False)[metric_column]
        .agg(Count="count", Sum="sum", Mean="mean", Median="median")
        .sort_values("Month")
    )
    for column in ["Sum", "Mean", "Median"]:
        trend[column] = pd.to_numeric(trend[column], errors="coerce").round(3)
    return trend


def render_descriptive_statistics_tab(
    custom_df: pd.DataFrame,
    reportpivots_df: pd.DataFrame,
    shippivots_df: pd.DataFrame,
) -> None:
    st.markdown('<div class="section-title">Descriptive Statistics</div>', unsafe_allow_html=True)
    st.caption("Analyze exactly the same filtered/export-ready tables from each source. No extra KPI logic is applied here.")

    sources = {
        "Custom Analytics": custom_df,
        "Noon & Manual Reports": reportpivots_df,
        "High-Frequency": shippivots_df,
    }
    available_sources = [label for label, df in sources.items() if isinstance(df, pd.DataFrame) and not df.empty]
    if not available_sources:
        st.info("No filtered source table is available for descriptive statistics yet.")
        return

    selected_source = st.selectbox("Source table", options=available_sources, key="atlas_descriptive_source")
    analysis_df = sources[selected_source].copy()
    numeric_options = dataframe_numeric_options(analysis_df)
    if not numeric_options:
        st.info("The selected source table has no numeric columns to analyze.")
        return

    metric_column = st.selectbox("Metric to analyze", options=numeric_options, key="atlas_descriptive_metric")
    group_options = ["None"] + dataframe_categorical_options(analysis_df)
    default_group_index = group_options.index("ShipName") if "ShipName" in group_options else 0
    group_column = st.selectbox("Optional group by", options=group_options, index=default_group_index, key="atlas_descriptive_group")

    stats_df = build_descriptive_statistics(analysis_df, metric_column)
    if stats_df.empty:
        st.info("No numeric values were found for the selected metric.")
        return

    values = pd.to_numeric(analysis_df[metric_column], errors="coerce")
    render_metric_cards(
        [
            ("Numeric Values", f"{values.notna().sum():,}", "numeric"),
            ("Total", f"{values.sum(skipna=True):,.3f}", "total"),
            ("Average", f"{values.mean(skipna=True):,.3f}", "average"),
            ("Missing", f"{values.isna().sum():,}", "missing"),
        ]
    )

    st.markdown('<div class="section-title">Overall statistics</div>', unsafe_allow_html=True)
    st.dataframe(format_display_dataframe(stats_df), use_container_width=True, hide_index=True)

    if group_column != "None":
        grouped_df = build_grouped_descriptive_statistics(analysis_df, metric_column, group_column)
        if not grouped_df.empty:
            st.markdown('<div class="section-title">Grouped statistics</div>', unsafe_allow_html=True)
            st.dataframe(format_display_dataframe(grouped_df.head(100)), use_container_width=True, hide_index=True)

    datetime_column = detect_analysis_datetime_column(analysis_df)
    if datetime_column:
        trend_df = build_monthly_trend(analysis_df, metric_column, datetime_column)
        if not trend_df.empty:
            st.markdown('<div class="section-title">Monthly trend</div>', unsafe_allow_html=True)
            st.dataframe(format_display_dataframe(trend_df), use_container_width=True, hide_index=True)
            chart_df = trend_df.set_index("Month")[["Sum", "Mean"]]
            st.line_chart(chart_df)

    outlier_values = values.dropna()
    if len(outlier_values) >= 4:
        q1 = outlier_values.quantile(0.25)
        q3 = outlier_values.quantile(0.75)
        iqr = q3 - q1
        lower = q1 - 1.5 * iqr
        upper = q3 + 1.5 * iqr
        outliers = analysis_df[(values < lower) | (values > upper)].copy()
        if not outliers.empty:
            st.markdown('<div class="section-title">Potential outlier rows</div>', unsafe_allow_html=True)
            display_cols = [column for column in ["ShipName", "ReportType", "StartDateTimeGMT", "DateTime", metric_column] if column in outliers.columns]
            if not display_cols:
                display_cols = list(outliers.columns[: min(8, len(outliers.columns))])
            st.dataframe(format_display_dataframe(outliers[display_cols].head(100)), use_container_width=True, hide_index=True)


# =============================================================================
# Sidebar/session helpers
# =============================================================================


def request_signature(username: str, auth_method: str, start_date: date) -> dict[str, Any]:
    return {
        "endpoint": ODATA_ENDPOINT,
        "username_hash": sha256(username.encode("utf-8")).hexdigest()[:12],
        "auth_method": auth_method.lower(),
        "start_date": start_date.isoformat(),
    }


def raw_data_covers_request(
    loaded_signature: dict[str, Any] | None,
    metadata: dict[str, Any] | None,
    requested_signature: dict[str, Any],
    requested_start_date: date,
) -> bool:
    if not loaded_signature or not metadata:
        return False
    for key in ["endpoint", "username_hash", "auth_method"]:
        if loaded_signature.get(key) != requested_signature.get(key):
            return False
    loaded_start_text = metadata.get("loaded_start_date") or loaded_signature.get("start_date")
    try:
        loaded_start_date = date.fromisoformat(str(loaded_start_text))
    except ValueError:
        return False
    return loaded_start_date <= requested_start_date


def get_loaded_state() -> tuple[pd.DataFrame | None, pd.DataFrame | None, dict[str, Any] | None]:
    return (
        st.session_state.get("loaded_raw_df"),
        st.session_state.get("loaded_long_df"),
        st.session_state.get("loaded_metadata"),
    )


def set_loaded_raw_state(raw_df: pd.DataFrame, metadata: dict[str, Any], signature: dict[str, Any]) -> None:
    metadata = metadata.copy()
    metadata.setdefault("loaded_at_utc", datetime.now(timezone.utc).strftime("%d-%m-%Y %H:%M:%S UTC"))
    metadata.setdefault("loaded_at_local", local_time_label())
    metadata["loaded_start_date"] = signature["start_date"]
    st.session_state["loaded_raw_df"] = raw_df
    st.session_state["loaded_metadata"] = metadata
    st.session_state["loaded_request_signature"] = signature
    st.session_state.pop("loaded_long_df", None)
    st.session_state.pop("loaded_prepare_signature", None)


def save_raw_snapshot(raw_df: pd.DataFrame, metadata: dict[str, Any], signature: dict[str, Any]) -> None:
    """Persist the latest successful raw API pull as a local fallback after app restarts."""
    try:
        SNAPSHOT_DIR.mkdir(parents=True, exist_ok=True)
        raw_df.to_parquet(RAW_SNAPSHOT_FILE, index=False)
        snapshot_payload = {
            "metadata": metadata,
            "signature": signature,
            "saved_at_utc": datetime.now(timezone.utc).strftime("%d-%m-%Y %H:%M:%S UTC"),
        }
        METADATA_SNAPSHOT_FILE.write_text(json.dumps(snapshot_payload, indent=2, default=str), encoding="utf-8")
    except Exception:
        # Snapshot persistence is a speed fallback only; never break the app if it fails.
        return


def load_raw_snapshot(
    requested_signature: dict[str, Any],
    requested_start_date: date,
) -> tuple[pd.DataFrame, dict[str, Any], dict[str, Any]] | None:
    """Load the latest local raw-data snapshot if it covers the current API request."""
    try:
        if not RAW_SNAPSHOT_FILE.is_file() or not METADATA_SNAPSHOT_FILE.is_file():
            return None
        snapshot_payload = json.loads(METADATA_SNAPSHOT_FILE.read_text(encoding="utf-8"))
        metadata = snapshot_payload.get("metadata") or {}
        signature = snapshot_payload.get("signature") or {}
        if not raw_data_covers_request(signature, metadata, requested_signature, requested_start_date):
            return None
        raw_df = pd.read_parquet(RAW_SNAPSHOT_FILE)
        if not isinstance(raw_df, pd.DataFrame) or raw_df.empty:
            return None
        metadata = metadata.copy()
        metadata["loaded_from_snapshot"] = True
        metadata.setdefault("snapshot_saved_at_utc", snapshot_payload.get("saved_at_utc", "-"))
        return raw_df, metadata, signature
    except Exception:
        return None


def set_loaded_long_state(long_df: pd.DataFrame, signature: dict[str, Any]) -> None:
    st.session_state["loaded_long_df"] = long_df
    st.session_state["loaded_prepare_signature"] = signature


def activate_reportdata_snapshot(
    username: str,
    auth_method: str,
    start_date: date,
) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, Any]]:
    """Load the latest ReportData snapshot and seed the shared transform cache once."""
    signature = request_signature(username, auth_method, start_date)
    snapshot = load_raw_snapshot(signature, start_date)
    if snapshot is None:
        raise FileNotFoundError("The refreshed ReportData snapshot could not be loaded.")

    raw_df, metadata, snapshot_signature = snapshot
    cached_prepare_long_data.clear()
    build_pivot_table.clear()
    long_df = cached_prepare_long_data(raw_df)

    set_loaded_raw_state(raw_df, metadata, snapshot_signature)
    prepare_signature = {
        **signature,
        "prepare_version": "atlasflow_dynamic_pivot_v3_oil_stats",
    }
    set_loaded_long_state(long_df, prepare_signature)

    active_metadata = st.session_state.get("loaded_metadata")
    if isinstance(active_metadata, dict):
        active_metadata["long_rows"] = int(len(long_df))
        active_metadata["available_variables"] = (
            int(long_df["ValueDescription"].nunique())
            if "ValueDescription" in long_df.columns
            else 0
        )
        st.session_state["loaded_metadata"] = active_metadata
        metadata = active_metadata

    return raw_df, long_df, metadata


def refresh_all_atlasflow_snapshots(
    username: str,
    password: str,
    token: str,
    auth_method: str,
    start_date: date,
) -> dict[str, dict[str, Any]]:
    """Refresh each API once into an atomic snapshot, then activate the new data."""
    results: dict[str, dict[str, Any]] = {}

    results["reportdata"] = fetch_report_data_to_snapshot(
        username, password, token, auth_method, start_date
    )
    results["reportpivots"] = fetch_wide_source_to_snapshot(
        "reportpivots", username, password, token, auth_method, start_date
    )
    results["shippivots"] = fetch_wide_source_to_snapshot(
        "shippivots", username, password, token, auth_method, start_date
    )

    cached_fetch_report_data.clear()
    cached_fetch_wide_odata_source.clear()
    clear_wide_source_state("reportpivots")
    clear_wide_source_state("shippivots")
    activate_reportdata_snapshot(username, auth_method, start_date)
    gc.collect()
    return results


def selected_vessel_controls() -> tuple[str, list[str]]:
    group_options = ["Single vessel", "All fleets"] + list(VESSEL_GROUPS.keys())
    selected_group = st.sidebar.selectbox("Fleet group", options=group_options, key="atlas_fleet_group")

    if selected_group == "Single vessel":
        vessel = st.sidebar.selectbox("Vessel to include", options=VESSEL_OPTIONS, key="atlas_single_vessel")
        st.session_state["atlas_last_fleet_group"] = selected_group
        return selected_group, [vessel]

    if selected_group == "All fleets":
        group_vessels = VESSEL_OPTIONS
    else:
        group_vessels = VESSEL_GROUPS[selected_group]

    vessel_key = "atlas_selected_vessels"
    last_group_key = "atlas_last_fleet_group"
    previous_group = st.session_state.get(last_group_key)

    # Streamlit keeps multiselect state after the first render, so changing from
    # Fleet 1 to All fleets could otherwise keep only the old Fleet 1 vessels.
    # Reset to the full new group whenever the fleet-group selector changes.
    if previous_group != selected_group:
        st.session_state[vessel_key] = list(group_vessels)
        st.session_state[last_group_key] = selected_group

    previous_vessels = st.session_state.get(vessel_key, group_vessels)
    if not isinstance(previous_vessels, list):
        previous_vessels = group_vessels
    valid_default_vessels = [vessel for vessel in previous_vessels if vessel in group_vessels]
    if not valid_default_vessels:
        valid_default_vessels = list(group_vessels)
    if valid_default_vessels != previous_vessels:
        st.session_state[vessel_key] = valid_default_vessels

    vessels = st.sidebar.multiselect(
        "Vessels to include",
        options=group_vessels,
        default=valid_default_vessels,
        key=vessel_key,
        help="This controls the displayed datasets only. The API data is loaded broadly.",
    )

    if not vessels:
        st.sidebar.caption("No vessels selected manually, so all vessels in this fleet group are included.")
        vessels = group_vessels

    return selected_group, list(vessels)

def sidebar_refresh_control() -> bool:
    refresh_requested = st.sidebar.button("Refresh all APIs", use_container_width=False)
    if refresh_requested:
        st.session_state["confirm_api_refresh"] = True

    refresh = False
    if st.session_state.get("confirm_api_refresh"):
        metadata = st.session_state.get("loaded_metadata") or {}
        last_load = metadata.get("loaded_at_local") or metadata.get("loaded_at_utc") or "-"
        last_load_display = str(last_load).replace(" EEST", "").replace(" EET", "")
        st.sidebar.warning(
            "Refresh will call all Atlas Flow APIs and may take a while.\n\n"
            f"Last updated data was on: {last_load_display} LT"
        )
        col1, col2 = st.sidebar.columns(2)
        if col1.button("Confirm"):
            refresh = True
            st.session_state["confirm_api_refresh"] = False
        if col2.button("Cancel"):
            st.session_state["confirm_api_refresh"] = False
            st.rerun()
    return refresh


def add_months(base_date: date, months: int) -> date:
    month_index = base_date.month - 1 + months
    year = base_date.year + month_index // 12
    month = month_index % 12 + 1
    days_in_month = [31, 29 if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0) else 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    day = min(base_date.day, days_in_month[month - 1])
    return date(year, month, day)


def first_day_of_month(value: date) -> date:
    return date(value.year, value.month, 1)


def last_day_previous_month(value: date) -> date:
    return first_day_of_month(value) - timedelta(days=1)


def clamp_period(start_value: date, end_value: date, min_date: date, max_date: date) -> tuple[date, date]:
    start_value = max(start_value, min_date)
    end_value = min(end_value, max_date)
    if start_value > end_value:
        return min_date, max_date
    return start_value, end_value


def dynamic_period_dates(period_label: str, min_date: date, max_date: date) -> tuple[date, date]:
    anchor_date = min(max_date, date.today())

    if period_label == "YTD":
        start_value = date(anchor_date.year, 1, 1)
        end_value = anchor_date
    elif period_label == "Previous month":
        end_value = last_day_previous_month(anchor_date)
        start_value = first_day_of_month(end_value)
    elif period_label == "Year to previous month":
        end_value = last_day_previous_month(anchor_date)
        start_value = date(end_value.year, 1, 1)
    elif period_label == "Previous 3 months":
        end_value = anchor_date
        start_value = add_months(anchor_date, -3) + timedelta(days=1)
    elif period_label == "Previous 6 months":
        end_value = anchor_date
        start_value = add_months(anchor_date, -6) + timedelta(days=1)
    elif period_label == "Previous 12 months":
        end_value = anchor_date
        start_value = add_months(anchor_date, -12) + timedelta(days=1)
    else:
        start_value = min_date
        end_value = max_date

    return clamp_period(start_value, end_value, min_date, max_date)


def render_date_slicer(df: pd.DataFrame) -> tuple[date, date]:
    min_date, max_date = dataframe_date_window(df)
    st.sidebar.markdown("### Period")
    if min_date >= max_date:
        st.sidebar.caption(f"Available data period: {min_date.strftime('%d/%m/%Y')}")
        return min_date, max_date

    period_mode = st.sidebar.selectbox(
        "Period preset",
        options=[
            "Custom range",
            "YTD",
            "Previous month",
            "Year to previous month",
            "Previous 3 months",
            "Previous 6 months",
            "Previous 12 months",
            "Full available period",
        ],
        index=0,
        key="atlas_period_preset",
        help="Use a dynamic preset or choose Custom range to control the period manually with the slider.",
    )

    if period_mode == "Custom range":
        selected_start, selected_end = st.sidebar.slider(
            "Report period",
            min_value=min_date,
            max_value=max_date,
            value=(min_date, max_date),
            format="DD/MM/YYYY",
            key="atlas_period_slicer",
        )
    else:
        selected_start, selected_end = dynamic_period_dates(period_mode, min_date, max_date)
        st.sidebar.caption(
            f"Selected period: {selected_start.strftime('%d/%m/%Y')} to {selected_end.strftime('%d/%m/%Y')}"
        )

    return selected_start, selected_end




# =============================================================================
# Streamlit Cloud-safe snapshot refresh helpers
# =============================================================================


def normalize_snapshot_values(df: pd.DataFrame) -> pd.DataFrame:
    """Store raw API values as nullable strings to keep Parquet schemas stable page by page."""
    if df.empty:
        return df.copy()
    safe_df = df.copy()
    for column in safe_df.columns:
        safe_df[column] = safe_df[column].astype("string")
    return safe_df


def write_parquet_pages(page_frames: list[pd.DataFrame], output_file: Path) -> int:
    """Write already-normalized page frames into one Parquet file with a stable schema."""
    output_file.parent.mkdir(parents=True, exist_ok=True)
    writer = None
    row_count = 0
    try:
        for frame in page_frames:
            if frame.empty:
                continue
            normalized = normalize_snapshot_values(frame)
            table = pa.Table.from_pandas(normalized, preserve_index=False)
            if writer is None:
                writer = pq.ParquetWriter(output_file, table.schema, compression="zstd")
            writer.write_table(table)
            row_count += len(normalized)
    finally:
        if writer is not None:
            writer.close()
    return row_count


def fetch_report_data_to_snapshot(
    username: str,
    password: str,
    token: str,
    auth_method: str,
    start_date: date,
) -> dict[str, Any]:
    """Fetch ReportData page-by-page and write to Parquet without keeping the full dataset in memory."""
    started_at = time.perf_counter()
    next_url = build_odata_url(start_date)
    first_url = next_url
    seen_urls: set[str] = set()
    pages = 0
    total_bytes = 0
    scanned_rows = 0
    kept_rows_total = 0
    consecutive_empty_pages = 0
    paging_stop_reason = "max_page_limit"
    # Use a unique temporary file per warmup request. Streamlit Cloud can run
    # multiple warmup/browser sessions at the same time; a fixed .tmp.parquet
    # name can be deleted by another session before this request reaches replace().
    tmp_file = RAW_SNAPSHOT_FILE.with_name(
        f"{RAW_SNAPSHOT_FILE.stem}.{os.getpid()}.{int(time.time() * 1000)}.tmp.parquet"
    )
    tmp_file.parent.mkdir(parents=True, exist_ok=True)
    if tmp_file.exists():
        tmp_file.unlink()

    auth = request_auth(username, password, auth_method)
    headers = request_headers(token, auth_method)
    writer = None

    try:
        with requests.Session() as session:
            session.headers.update(headers)
            for _ in range(MAX_ODATA_PAGES):
                if next_url in seen_urls:
                    paging_stop_reason = "repeated_current_url"
                    break
                seen_urls.add(next_url)

                response = request_with_retry(session, next_url, auth=auth, timeout=90)
                total_bytes += len(response.content)
                response.raise_for_status()
                pages += 1

                page_rows, next_link = extract_odata_page(response.json())
                scanned_rows += len(page_rows)
                compact_rows = compact_odata_rows(page_rows)
                if compact_rows:
                    page_df = normalize_snapshot_values(rows_to_dataframe(compact_rows))
                    table = pa.Table.from_pandas(page_df, preserve_index=False)
                    if writer is None:
                        writer = pq.ParquetWriter(tmp_file, table.schema, compression="zstd")
                    writer.write_table(table)
                    kept_rows_total += len(page_df)
                    del page_df, table

                consecutive_empty_pages = consecutive_empty_pages + 1 if len(page_rows) == 0 else 0
                del page_rows, compact_rows
                gc.collect()

                should_continue, resolved_next_url, stop_reason = should_continue_odata_paging(
                    current_url=next_url,
                    next_link=next_link,
                    seen_urls=seen_urls,
                    consecutive_empty_pages=consecutive_empty_pages,
                )
                if not should_continue:
                    paging_stop_reason = stop_reason
                    break
                next_url = resolved_next_url or next_url
    finally:
        if writer is not None:
            writer.close()

    RAW_SNAPSHOT_FILE.parent.mkdir(parents=True, exist_ok=True)

    if kept_rows_total == 0:
        empty_df = normalize_snapshot_values(pd.DataFrame(columns=SOURCE_COLUMNS))
        empty_table = pa.Table.from_pandas(empty_df, preserve_index=False)
        pq.write_table(empty_table, tmp_file, compression="zstd")
        del empty_df, empty_table

    if not tmp_file.exists():
        raise FileNotFoundError(f"ReportData snapshot temporary file was not created: {tmp_file}")

    tmp_file.replace(RAW_SNAPSHOT_FILE)
    loaded_at_utc = datetime.now(timezone.utc)
    metadata = {
        "loaded_at_utc": loaded_at_utc.strftime("%d-%m-%Y %H:%M:%S UTC"),
        "loaded_at_local": local_time_label(loaded_at_utc),
        "rows": kept_rows_total,
        "kept_rows": kept_rows_total,
        "scanned_rows": scanned_rows,
        "discarded_rows": max(scanned_rows - kept_rows_total, 0),
        "pages": pages,
        "downloaded_mb": round(total_bytes / 1024 / 1024, 2),
        "fetch_seconds": round(time.perf_counter() - started_at, 2),
        "first_url": first_url,
        "hit_page_limit": pages >= MAX_ODATA_PAGES and paging_stop_reason == "max_page_limit",
        "paging_stop_reason": paging_stop_reason,
        "max_pages": MAX_ODATA_PAGES,
        "loaded_start_date": start_date.isoformat(),
        "snapshot_format": "parquet",
        "reportdata_mode": "atlasflow_performance_oil_cargo_whitelist",
        "value_description_whitelist_count": len(REPORTDATA_VALUE_WHITELIST),
    }
    signature = request_signature(username, auth_method, start_date)
    snapshot_payload = {
        "metadata": metadata,
        "signature": signature,
        "saved_at_utc": datetime.now(timezone.utc).strftime("%d-%m-%Y %H:%M:%S UTC"),
    }
    SNAPSHOT_DIR.mkdir(parents=True, exist_ok=True)
    METADATA_SNAPSHOT_FILE.write_text(json.dumps(snapshot_payload, indent=2, default=str), encoding="utf-8")
    return metadata


def fetch_wide_source_to_snapshot(
    source_key: str,
    username: str,
    password: str,
    token: str,
    auth_method: str,
    start_date: date,
) -> dict[str, Any]:
    """Fetch a wide OData source page-by-page and write to Parquet safely."""
    config = SOURCE_CONFIGS[source_key]
    endpoint = str(config["endpoint"])
    datetime_column = str(config.get("datetime_candidates", ["DateTime"])[0])
    next_url = build_wide_odata_url(endpoint, start_date, datetime_column)
    first_url = next_url
    seen_urls: set[str] = set()
    pages = 0
    total_bytes = 0
    row_count = 0
    consecutive_empty_pages = 0
    paging_stop_reason = "max_page_limit"
    all_columns: list[str] = []
    target_file = Path(config["snapshot_file"])
    # Use a unique temporary file per warmup request to avoid cross-session
    # collisions when several source warmups or browser tabs run concurrently.
    tmp_file = target_file.with_name(
        f"{target_file.stem}.{os.getpid()}.{int(time.time() * 1000)}.tmp.parquet"
    )
    tmp_file.parent.mkdir(parents=True, exist_ok=True)
    if tmp_file.exists():
        tmp_file.unlink()

    auth = request_auth(username, password, auth_method)
    headers = request_headers(token, auth_method)
    writer = None
    started_at = time.perf_counter()

    try:
        with requests.Session() as session:
            session.headers.update(headers)
            for _ in range(MAX_ODATA_PAGES):
                if next_url in seen_urls:
                    paging_stop_reason = "repeated_current_url"
                    break
                seen_urls.add(next_url)
                response = request_with_retry(session, next_url, auth=auth, timeout=90)
                total_bytes += len(response.content)
                response.raise_for_status()
                pages += 1

                page_rows, next_link = extract_odata_page(response.json())
                page_df = pd.DataFrame(page_rows)
                if "__metadata" in page_df.columns:
                    page_df = page_df.drop(columns=["__metadata"])

                if not page_df.empty:
                    if not all_columns:
                        all_columns = list(page_df.columns)
                    else:
                        for column in page_df.columns:
                            if column not in all_columns:
                                all_columns.append(column)
                        for column in all_columns:
                            if column not in page_df.columns:
                                page_df[column] = pd.NA
                        page_df = page_df[all_columns]

                    page_df = normalize_snapshot_values(page_df)
                    table = pa.Table.from_pandas(page_df, preserve_index=False)
                    if writer is None:
                        writer = pq.ParquetWriter(tmp_file, table.schema, compression="zstd")
                    else:
                        # New columns after the first page are rare; keep only first-page schema to avoid schema errors.
                        schema_names = writer.schema.names
                        page_df = page_df[[column for column in schema_names if column in page_df.columns]]
                        for column in schema_names:
                            if column not in page_df.columns:
                                page_df[column] = pd.NA
                        page_df = page_df[schema_names]
                        table = pa.Table.from_pandas(page_df, schema=writer.schema, preserve_index=False)
                    writer.write_table(table)
                    row_count += len(page_df)
                    del page_df, table

                consecutive_empty_pages = consecutive_empty_pages + 1 if len(page_rows) == 0 else 0
                del page_rows
                gc.collect()

                should_continue, resolved_next_url, stop_reason = should_continue_odata_paging(
                    current_url=next_url,
                    next_link=next_link,
                    seen_urls=seen_urls,
                    consecutive_empty_pages=consecutive_empty_pages,
                )
                if not should_continue:
                    paging_stop_reason = stop_reason
                    break
                next_url = resolved_next_url or next_url
    finally:
        if writer is not None:
            writer.close()

    target_file.parent.mkdir(parents=True, exist_ok=True)

    if row_count == 0:
        # Some wide endpoints can legally return no rows for the selected API window.
        # In that case, create an empty snapshot with discovered columns if possible.
        empty_columns = all_columns if all_columns else ["NoData"]
        empty_df = normalize_snapshot_values(pd.DataFrame(columns=empty_columns))
        empty_df.to_parquet(tmp_file, index=False, compression="zstd")
        del empty_df

    if not tmp_file.exists():
        raise FileNotFoundError(f"{config['label']} snapshot temporary file was not created: {tmp_file}")

    # Validate before replacing the previous good snapshot. Do not accept a placeholder
    # NoData file when API rows were written.
    try:
        parquet_columns = pq.ParquetFile(tmp_file).schema.names
    except Exception as exc:
        raise RuntimeError(f"{config['label']} snapshot validation failed before save: {exc}") from exc
    if row_count > 0 and parquet_columns == ["NoData"]:
        raise RuntimeError(f"{config['label']} snapshot validation failed: placeholder NoData file for {row_count:,} rows.")

    if target_file.exists():
        target_file.unlink()
    os.replace(str(tmp_file), str(target_file))
    loaded_at_utc = datetime.now(timezone.utc)
    metadata = {
        "source": config["label"],
        "endpoint": endpoint,
        "loaded_at_utc": loaded_at_utc.strftime("%d-%m-%Y %H:%M:%S UTC"),
        "loaded_at_local": local_time_label(loaded_at_utc),
        "loaded_start_date": start_date.isoformat(),
        "rows": int(row_count),
        "columns": int(len(all_columns)),
        "pages": pages,
        "downloaded_mb": round(total_bytes / 1024 / 1024, 2),
        "fetch_seconds": round(time.perf_counter() - started_at, 2),
        "first_url": first_url,
        "hit_page_limit": pages >= MAX_ODATA_PAGES and paging_stop_reason == "max_page_limit",
        "paging_stop_reason": paging_stop_reason,
        "max_pages": MAX_ODATA_PAGES,
        "snapshot_format": "parquet",
    }
    signature = source_signature(source_key, username, auth_method, start_date)
    payload = {
        "metadata": metadata,
        "signature": signature,
        "saved_at_utc": datetime.now(timezone.utc).strftime("%d-%m-%Y %H:%M:%S UTC"),
    }
    SNAPSHOT_DIR.mkdir(parents=True, exist_ok=True)
    Path(config["metadata_file"]).write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")
    return metadata




# =============================================================================
# Persistent prepared multi-source snapshots and incremental refresh
# =============================================================================


class AtlasRefreshAlreadyRunning(RuntimeError):
    pass


def read_int_secret(
    name: str,
    default: int,
    *,
    minimum: int = 0,
    maximum: int = 10000,
) -> int:
    try:
        value = int(read_secret(name, str(default)))
    except (TypeError, ValueError):
        value = default
    return min(max(value, minimum), maximum)


def source_secret_name(source_key: str, suffix: str) -> str:
    return f"ATLASFLOW_{source_key.upper()}_{suffix}"


def source_primary_datetime_column(source_key: str) -> str:
    if source_key == "reportdata":
        return "StartDateTimeGMT"
    candidates = list(SOURCE_CONFIGS[source_key].get("datetime_candidates", []))
    return str(candidates[0] if candidates else "DateTime")


def source_manifest_path(source_key: str) -> Path:
    return SNAPSHOT_DIR / f"{source_key}_prepared_manifest.json"


def source_lock_path(source_key: str) -> Path:
    return SNAPSHOT_DIR / f"{source_key}_refresh.lock"


def source_status_path(source_key: str) -> Path:
    return SNAPSHOT_DIR / f"{source_key}_refresh_status.json"


def source_snapshot_path(source_key: str, generation: str) -> Path:
    return SNAPSHOT_DIR / f"{source_key}_prepared_{generation}.parquet"


def source_data_signature(source_key: str) -> str:
    if source_key == "reportdata":
        signature_text = "|".join(
            [
                ATLAS_SNAPSHOT_SCHEMA_VERSION,
                ATLAS_PREPARE_VERSION,
                *REPORTDATA_VALUE_WHITELIST,
                *SOURCE_COLUMNS,
            ]
        )
    else:
        config = SOURCE_CONFIGS[source_key]
        signature_text = "|".join(
            [
                ATLAS_SNAPSHOT_SCHEMA_VERSION,
                source_key,
                str(config["endpoint"]),
                *map(str, config.get("datetime_candidates", [])),
            ]
        )
    return sha256(signature_text.encode("utf-8")).hexdigest()[:16]


def atlas_source_signature(
    source_key: str,
    username: str,
    auth_method: str,
    start_date: date,
) -> dict[str, Any]:
    config = SOURCE_CONFIGS[source_key]
    return {
        "source": source_key,
        "endpoint": str(config["endpoint"]),
        "username_hash": sha256(username.encode("utf-8")).hexdigest()[:12],
        "auth_method": auth_method.lower(),
        "start_date": start_date.isoformat(),
        "data_signature": source_data_signature(source_key),
    }


def source_signature_covers_request(
    stored_signature: dict[str, Any] | None,
    requested_signature: dict[str, Any] | None,
    metadata: dict[str, Any] | None,
    requested_start_date: date,
) -> bool:
    if not stored_signature or not requested_signature or not metadata:
        return False
    for key in ["endpoint", "username_hash", "auth_method"]:
        if stored_signature.get(key) != requested_signature.get(key):
            return False
    source_key = str(requested_signature.get("source") or stored_signature.get("source") or "reportdata")
    expected_data_signature = source_data_signature(source_key)
    if stored_signature.get("data_signature") != expected_data_signature:
        return False
    loaded_start_text = metadata.get("loaded_start_date") or stored_signature.get("start_date")
    try:
        loaded_start_date = date.fromisoformat(str(loaded_start_text))
    except ValueError:
        return False
    return loaded_start_date <= requested_start_date


def _atomic_write_text(path: Path, text_value: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = path.with_name(
        f"{path.name}.{os.getpid()}.{int(time.time() * 1000)}.tmp"
    )
    try:
        temp_path.write_text(text_value, encoding="utf-8")
        os.replace(str(temp_path), str(path))
    finally:
        if temp_path.exists():
            try:
                temp_path.unlink()
            except OSError:
                pass


def read_source_refresh_status(source_key: str) -> dict[str, Any] | None:
    try:
        path = source_status_path(source_key)
        if not path.is_file():
            return None
        payload = json.loads(path.read_text(encoding="utf-8"))
        return payload if isinstance(payload, dict) else None
    except Exception:
        return None


def update_source_refresh_status(source_key: str, **updates: Any) -> None:
    payload = read_source_refresh_status(source_key) or {}
    payload.update(updates)
    payload["source"] = source_key
    payload["updated_at_utc"] = datetime.now(timezone.utc).isoformat()
    payload.setdefault("pid", os.getpid())
    try:
        _atomic_write_text(
            source_status_path(source_key),
            json.dumps(payload, indent=2, default=str),
        )
    except Exception:
        return


def source_refresh_status_summary(source_key: str) -> str:
    status = read_source_refresh_status(source_key) or {}
    stage = str(status.get("stage", "refreshing"))
    refresh_mode = str(status.get("refresh_mode", "refresh"))
    chunk_index = int(status.get("chunk_index", 0) or 0)
    chunks_total = int(status.get("chunks_total", 0) or 0)
    chunk_start = status.get("chunk_start_date")
    chunk_end = status.get("chunk_end_date_exclusive")
    parts = [f"{source_key}: {refresh_mode} {stage}"]
    if chunk_index and chunks_total:
        parts.append(f"window {chunk_index} of {chunks_total}")
    if chunk_start and chunk_end:
        parts.append(f"{chunk_start} to {chunk_end}")
    return "; ".join(parts)


@contextmanager
def source_refresh_lock(source_key: str) -> Any:
    """Prevent duplicate refreshes of the same API source."""
    SNAPSHOT_DIR.mkdir(parents=True, exist_ok=True)
    lock_path = source_lock_path(source_key)

    if fcntl is not None:
        handle = lock_path.open("a+", encoding="utf-8")
        try:
            try:
                fcntl.flock(handle.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
            except BlockingIOError:
                yield False
                return
            handle.seek(0)
            handle.truncate()
            handle.write(
                json.dumps(
                    {
                        "source": source_key,
                        "pid": os.getpid(),
                        "started_at_utc": datetime.now(timezone.utc).isoformat(),
                    }
                )
            )
            handle.flush()
            yield True
        finally:
            try:
                fcntl.flock(handle.fileno(), fcntl.LOCK_UN)
            except Exception:
                pass
            handle.close()
        return

    lock_fd: int | None = None
    try:
        try:
            lock_fd = os.open(str(lock_path), os.O_CREAT | os.O_EXCL | os.O_WRONLY)
        except FileExistsError:
            yield False
            return
        os.write(
            lock_fd,
            json.dumps(
                {
                    "source": source_key,
                    "pid": os.getpid(),
                    "started_at_utc": datetime.now(timezone.utc).isoformat(),
                }
            ).encode("utf-8"),
        )
        yield True
    finally:
        if lock_fd is not None:
            os.close(lock_fd)
            try:
                lock_path.unlink()
            except FileNotFoundError:
                pass


def read_source_manifest(source_key: str) -> dict[str, Any] | None:
    try:
        path = source_manifest_path(source_key)
        if not path.is_file():
            return None
        payload = json.loads(path.read_text(encoding="utf-8"))
        return payload if isinstance(payload, dict) else None
    except Exception:
        return None


def source_manifest_is_valid(
    source_key: str,
    manifest: dict[str, Any] | None,
    requested_signature: dict[str, Any],
    requested_start_date: date,
) -> bool:
    if not manifest:
        return False
    if manifest.get("snapshot_schema_version") != ATLAS_SNAPSHOT_SCHEMA_VERSION:
        return False
    if manifest.get("source") != source_key:
        return False
    metadata = manifest.get("metadata") or {}
    stored_signature = manifest.get("signature") or {}
    if not source_signature_covers_request(
        stored_signature,
        requested_signature,
        metadata,
        requested_start_date,
    ):
        return False
    snapshot_file = SNAPSHOT_DIR / str(manifest.get("prepared_file", ""))
    return bool(manifest.get("generation")) and snapshot_file.is_file()


@st.cache_data(show_spinner=False)
def cached_read_prepared_source_snapshot(
    source_key: str,
    generation: str,
    snapshot_file: str,
) -> pd.DataFrame:
    del source_key, generation  # Deliberate cache keys.
    return pd.read_parquet(snapshot_file)


def load_source_snapshot(
    source_key: str,
    requested_signature: dict[str, Any],
    requested_start_date: date,
) -> tuple[pd.DataFrame, dict[str, Any], dict[str, Any]] | None:
    # Older call sites do not include source/data_signature in the requested signature.
    requested_signature = dict(requested_signature or {})
    requested_signature.setdefault("source", source_key)
    requested_signature.setdefault("data_signature", source_data_signature(source_key))
    manifest = read_source_manifest(source_key)
    if not source_manifest_is_valid(
        source_key,
        manifest,
        requested_signature,
        requested_start_date,
    ):
        return None
    assert manifest is not None
    generation = str(manifest["generation"])
    snapshot_path = SNAPSHOT_DIR / str(manifest["prepared_file"])
    try:
        df = cached_read_prepared_source_snapshot(
            source_key,
            generation,
            str(snapshot_path),
        )
    except Exception:
        return None
    if not isinstance(df, pd.DataFrame):
        return None
    metadata = dict(manifest.get("metadata") or {})
    metadata["loaded_from_snapshot"] = True
    metadata["snapshot_generation"] = generation
    metadata.setdefault("snapshot_saved_at_utc", manifest.get("saved_at_utc", "-"))
    metadata.setdefault("snapshot_schema_version", ATLAS_SNAPSHOT_SCHEMA_VERSION)
    return df, metadata, dict(manifest.get("signature") or {})


def source_snapshot_info(
    source_key: str,
    username: str,
    auth_method: str,
    start_date: date,
) -> tuple[dict[str, Any], dict[str, Any]] | None:
    requested_signature = atlas_source_signature(
        source_key,
        username,
        auth_method,
        start_date,
    )
    manifest = read_source_manifest(source_key)
    if not source_manifest_is_valid(
        source_key,
        manifest,
        requested_signature,
        start_date,
    ):
        return None
    assert manifest is not None
    return dict(manifest.get("metadata") or {}), manifest


def build_refresh_windows(
    start_date: date,
    end_date_exclusive: date,
    chunk_days: int,
) -> list[tuple[date, date]]:
    if start_date >= end_date_exclusive:
        return []
    windows: list[tuple[date, date]] = []
    cursor = start_date
    while cursor < end_date_exclusive:
        next_cursor = min(cursor + timedelta(days=chunk_days), end_date_exclusive)
        windows.append((cursor, next_cursor))
        cursor = next_cursor
    return windows


def build_source_window_url(
    source_key: str,
    window_start: date,
    window_end_exclusive: date,
) -> str:
    config = SOURCE_CONFIGS[source_key]
    datetime_column = source_primary_datetime_column(source_key)
    # Query one day earlier because Marorka's OData V1 filter uses strict gt.
    query_start = window_start - timedelta(days=1)
    filter_text = (
        f"{datetime_column} gt DateTime'{query_start.isoformat()}' and "
        f"{datetime_column} lt DateTime'{window_end_exclusive.isoformat()}'"
    )
    params: dict[str, str] = {"$filter": filter_text}
    if source_key == "reportdata":
        params["$select"] = ",".join(SOURCE_COLUMNS)
    return f"{config['endpoint']}?{urlencode(params)}"


def trim_frame_to_window(
    df: pd.DataFrame,
    datetime_column: str,
    window_start: date,
    window_end_exclusive: date,
) -> pd.DataFrame:
    if df.empty or datetime_column not in df.columns:
        return df.iloc[0:0].copy()
    values = parse_datetime_series(df[datetime_column])
    start_ts = pd.Timestamp(window_start, tz="UTC")
    end_ts = pd.Timestamp(window_end_exclusive, tz="UTC")
    return df.loc[values.ge(start_ts) & values.lt(end_ts)].copy()


def prepare_reportdata_snapshot_frame(df: pd.DataFrame) -> pd.DataFrame:
    for column in SOURCE_COLUMNS:
        if column not in df.columns:
            df[column] = pd.NA
    prepared = df[SOURCE_COLUMNS].copy()
    prepared["ReportId"] = pd.to_numeric(prepared["ReportId"], errors="coerce").astype("Int64")
    prepared["StartDateTimeGMT"] = parse_datetime_series(prepared["StartDateTimeGMT"])
    prepared["EndDateTimeGMT"] = parse_datetime_series(prepared["EndDateTimeGMT"])
    prepared["LapTime"] = parse_numeric_series(prepared["LapTime"])
    prepared["ParsedValue"] = parse_numeric_series(prepared["ReportedValue"])
    prepared = prepared[
        prepared["ValueDescription"].notna()
        & ~prepared["ReportType"].isin(EXCLUDED_REPORT_TYPES)
    ].copy()
    for column in ["ShipName", "ReportType", "StateName", "ValueDescription", "ReportedValue"]:
        prepared[column] = prepared[column].astype("string")
    return prepared[[*SOURCE_COLUMNS, "ParsedValue"]]


def normalize_wide_snapshot_frame(source_key: str, df: pd.DataFrame) -> pd.DataFrame:
    prepared = df.copy()
    prepared.columns = [str(column) for column in prepared.columns]
    if "__metadata" in prepared.columns:
        prepared = prepared.drop(columns=["__metadata"])
    datetime_column = source_primary_datetime_column(source_key)
    for column in prepared.columns:
        if column == datetime_column:
            prepared[column] = parse_datetime_series(prepared[column])
        else:
            prepared[column] = prepared[column].astype("string")
    return prepared


def normalize_source_snapshot_frame(source_key: str, df: pd.DataFrame) -> pd.DataFrame:
    if source_key == "reportdata":
        return prepare_reportdata_snapshot_frame(df)
    return normalize_wide_snapshot_frame(source_key, df)


def deduplicate_source_window(source_key: str, df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    if source_key == "reportdata":
        report_ids = df["ReportId"].astype("string").fillna("")
        value_keys = df["ValueDescription"].map(normalize_text)
        with_id = df.loc[report_ids.str.len().gt(0)].copy()
        if not with_id.empty:
            with_id["_rid"] = report_ids.loc[with_id.index]
            with_id["_value_key"] = value_keys.loc[with_id.index]
            with_id = with_id.drop_duplicates(["_rid", "_value_key"], keep="last").drop(columns=["_rid", "_value_key"])
        without_id = df.loc[report_ids.str.len().eq(0)].drop_duplicates(keep="last")
        return pd.concat([with_id, without_id], ignore_index=True)
    if "ReportId" in df.columns and df["ReportId"].notna().any():
        return df.drop_duplicates(["ReportId"], keep="last").reset_index(drop=True)
    datetime_column = source_primary_datetime_column(source_key)
    keys = [column for column in ["ShipName", datetime_column] if column in df.columns]
    if len(keys) == 2:
        return df.drop_duplicates(keys, keep="last").reset_index(drop=True)
    return df.drop_duplicates(keep="last").reset_index(drop=True)


def fetch_source_window(
    source_key: str,
    username: str,
    password: str,
    token: str,
    auth_method: str,
    window_start: date,
    window_end_exclusive: date,
    *,
    deadline_monotonic: float,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    started_at = time.perf_counter()
    next_url = build_source_window_url(source_key, window_start, window_end_exclusive)
    first_url = next_url
    seen_urls: set[str] = set()
    frames: list[pd.DataFrame] = []
    pages = 0
    total_bytes = 0
    scanned_rows = 0
    consecutive_empty_pages = 0
    paging_stop_reason = "max_page_limit"
    auth = request_auth(username, password, auth_method)
    headers = request_headers(token, auth_method)
    datetime_column = source_primary_datetime_column(source_key)

    with requests.Session() as session:
        session.headers.update(headers)
        for _ in range(MAX_ODATA_PAGES):
            if time.perf_counter() >= deadline_monotonic:
                raise TimeoutError(f"{SOURCE_CONFIGS[source_key]['label']} refresh exceeded its safety time limit.")
            if next_url in seen_urls:
                paging_stop_reason = "repeated_current_url"
                break
            seen_urls.add(next_url)
            response = request_with_retry(
                session,
                next_url,
                auth=auth,
                timeout=API_REQUEST_TIMEOUT_SECONDS,
                max_attempts=API_REQUEST_MAX_ATTEMPTS,
            )
            total_bytes += len(response.content)
            response.raise_for_status()
            pages += 1
            page_rows, next_link = extract_odata_page(response.json())
            scanned_rows += len(page_rows)
            consecutive_empty_pages = consecutive_empty_pages + 1 if len(page_rows) == 0 else 0

            if source_key == "reportdata":
                page_rows = compact_odata_rows(page_rows)
            page_df = pd.DataFrame(page_rows)
            if not page_df.empty:
                if source_key == "reportdata":
                    for column in SOURCE_COLUMNS:
                        if column not in page_df.columns:
                            page_df[column] = pd.NA
                    page_df = page_df[SOURCE_COLUMNS]
                page_df = trim_frame_to_window(
                    page_df,
                    datetime_column,
                    window_start,
                    window_end_exclusive,
                )
                if not page_df.empty:
                    frames.append(page_df)

            should_continue, resolved_next_url, stop_reason = should_continue_odata_paging(
                current_url=next_url,
                next_link=next_link,
                seen_urls=seen_urls,
                consecutive_empty_pages=consecutive_empty_pages,
            )
            if not should_continue:
                paging_stop_reason = stop_reason or "end_of_feed"
                break
            next_url = resolved_next_url or next_url

    hit_page_limit = pages >= MAX_ODATA_PAGES and paging_stop_reason == "max_page_limit"
    if hit_page_limit:
        raise RuntimeError(
            f"{SOURCE_CONFIGS[source_key]['label']} reached {MAX_ODATA_PAGES:,} pages inside "
            f"the bounded window {window_start} to {window_end_exclusive}. Reduce the configured chunk size."
        )

    if frames:
        window_df = pd.concat(frames, ignore_index=True, sort=False)
    elif source_key == "reportdata":
        window_df = pd.DataFrame(columns=SOURCE_COLUMNS)
    else:
        window_df = pd.DataFrame()
    window_df = normalize_source_snapshot_frame(source_key, window_df)
    window_df = deduplicate_source_window(source_key, window_df)

    date_values = (
        pd.to_datetime(window_df.get(datetime_column), errors="coerce", utc=True)
        if datetime_column in window_df.columns
        else pd.Series(dtype="datetime64[ns, UTC]")
    )
    metadata = {
        "window_start_date": window_start.isoformat(),
        "window_end_date_exclusive": window_end_exclusive.isoformat(),
        "rows": int(len(window_df)),
        "scanned_rows": int(scanned_rows),
        "pages": int(pages),
        "downloaded_mb": round(total_bytes / 1024 / 1024, 2),
        "fetch_seconds": round(time.perf_counter() - started_at, 2),
        "first_url": first_url,
        "paging_stop_reason": paging_stop_reason,
        "hit_page_limit": False,
        "latest_source_date": date_values.max().date().isoformat() if not date_values.empty and date_values.notna().any() else None,
    }
    return window_df, metadata


def write_temp_chunk(df: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_parquet(path, index=False, compression="zstd")
    if not path.is_file() or path.stat().st_size <= 0:
        raise RuntimeError(f"Temporary snapshot chunk was not created: {path}")


def collect_source_chunks(
    source_key: str,
    username: str,
    password: str,
    token: str,
    auth_method: str,
    refresh_start_date: date,
    end_date_exclusive: date,
    chunk_days: int,
    max_duration_seconds: int,
    refresh_mode: str,
) -> tuple[Path, list[Path], list[str], dict[str, Any]]:
    work_dir = SNAPSHOT_DIR / f".refresh_{source_key}_{os.getpid()}_{int(time.time() * 1000)}"
    work_dir.mkdir(parents=True, exist_ok=False)
    windows = build_refresh_windows(refresh_start_date, end_date_exclusive, chunk_days)
    deadline = time.perf_counter() + max_duration_seconds
    chunk_files: list[Path] = []
    union_columns: list[str] = []
    total_rows = total_scanned = total_pages = 0
    total_downloaded_mb = total_fetch_seconds = 0.0
    first_url = "-"
    latest_source_date: str | None = None

    for index, (window_start, window_end) in enumerate(windows, start=1):
        update_source_refresh_status(
            source_key,
            state="running",
            stage="fetching",
            refresh_mode=refresh_mode,
            chunk_index=index,
            chunks_total=len(windows),
            chunk_start_date=window_start.isoformat(),
            chunk_end_date_exclusive=window_end.isoformat(),
            pages_completed=total_pages,
            rows_kept=total_rows,
        )
        frame, window_meta = fetch_source_window(
            source_key,
            username,
            password,
            token,
            auth_method,
            window_start,
            window_end,
            deadline_monotonic=deadline,
        )
        chunk_path = work_dir / f"chunk_{index:04d}.parquet"
        write_temp_chunk(frame, chunk_path)
        chunk_files.append(chunk_path)
        for column in frame.columns:
            if column not in union_columns:
                union_columns.append(str(column))
        total_rows += len(frame)
        total_scanned += int(window_meta["scanned_rows"])
        total_pages += int(window_meta["pages"])
        total_downloaded_mb += float(window_meta["downloaded_mb"])
        total_fetch_seconds += float(window_meta["fetch_seconds"])
        if first_url == "-":
            first_url = str(window_meta.get("first_url", "-"))
        latest_value = window_meta.get("latest_source_date")
        if latest_value and (latest_source_date is None or str(latest_value) > latest_source_date):
            latest_source_date = str(latest_value)
        del frame
        gc.collect()

    metadata = {
        "refresh_mode": refresh_mode,
        "refresh_api_start_date": refresh_start_date.isoformat(),
        "refresh_end_date_exclusive": end_date_exclusive.isoformat(),
        "chunk_days": int(chunk_days),
        "chunks_total": len(windows),
        "chunks_completed": len(chunk_files),
        "refresh_rows": int(total_rows),
        "scanned_rows": int(total_scanned),
        "discarded_rows": max(int(total_scanned) - int(total_rows), 0),
        "pages": int(total_pages),
        "downloaded_mb": round(total_downloaded_mb, 2),
        "fetch_seconds": round(total_fetch_seconds, 2),
        "prepare_seconds": 0,
        "first_url": first_url,
        "paging_stop_reason": "all_windows_completed",
        "hit_page_limit": False,
        "max_pages": MAX_ODATA_PAGES,
        "max_pages_per_window": MAX_ODATA_PAGES,
        "latest_source_date": latest_source_date,
    }
    return work_dir, chunk_files, union_columns, metadata


def source_snapshot_columns(source_key: str, existing_path: Path | None, fresh_columns: list[str]) -> list[str]:
    if source_key == "reportdata":
        return [*SOURCE_COLUMNS, "ParsedValue"]
    columns: list[str] = []
    if existing_path is not None and existing_path.is_file():
        try:
            for column in pq.ParquetFile(existing_path).schema.names:
                if column not in columns:
                    columns.append(column)
        except Exception:
            pass
    for column in fresh_columns:
        if column not in columns:
            columns.append(column)
    datetime_column = source_primary_datetime_column(source_key)
    if datetime_column not in columns:
        columns.insert(0, datetime_column)
    return columns or [datetime_column]


def source_arrow_schema(source_key: str, columns: list[str]) -> pa.Schema:
    if source_key == "reportdata":
        float_columns = {"LapTime", "ParsedValue"}
        datetime_columns = {"StartDateTimeGMT", "EndDateTimeGMT"}
        fields = []
        for column in columns:
            if column == "ReportId":
                field_type = pa.int64()
            elif column in float_columns:
                field_type = pa.float64()
            elif column in datetime_columns:
                field_type = pa.timestamp("ns", tz="UTC")
            else:
                field_type = pa.string()
            fields.append(pa.field(column, field_type, nullable=True))
        return pa.schema(fields)

    datetime_column = source_primary_datetime_column(source_key)
    return pa.schema(
        [
            pa.field(
                column,
                pa.timestamp("ns", tz="UTC") if column == datetime_column else pa.string(),
                nullable=True,
            )
            for column in columns
        ]
    )


def align_source_frame(
    source_key: str,
    frame: pd.DataFrame,
    columns: list[str],
) -> pd.DataFrame:
    aligned = frame.copy()
    for column in columns:
        if column not in aligned.columns:
            aligned[column] = pd.NA
    aligned = aligned[columns]
    if source_key == "reportdata":
        aligned["ReportId"] = pd.to_numeric(aligned["ReportId"], errors="coerce").astype("Int64")
        for column in ["StartDateTimeGMT", "EndDateTimeGMT"]:
            aligned[column] = pd.to_datetime(aligned[column], errors="coerce", utc=True)
        for column in ["LapTime", "ParsedValue"]:
            aligned[column] = pd.to_numeric(aligned[column], errors="coerce")
        for column in columns:
            if column not in {"ReportId", "StartDateTimeGMT", "EndDateTimeGMT", "LapTime", "ParsedValue"}:
                aligned[column] = aligned[column].astype("string")
        return aligned

    datetime_column = source_primary_datetime_column(source_key)
    for column in columns:
        if column == datetime_column:
            aligned[column] = pd.to_datetime(aligned[column], errors="coerce", utc=True)
        else:
            aligned[column] = aligned[column].astype("string")
    return aligned


def append_frame_to_parquet_writer(
    writer: pq.ParquetWriter,
    source_key: str,
    frame: pd.DataFrame,
    columns: list[str],
    schema: pa.Schema,
) -> int:
    if frame.empty:
        return 0
    aligned = align_source_frame(source_key, frame, columns)
    table = pa.Table.from_pandas(aligned, schema=schema, preserve_index=False)
    writer.write_table(table)
    return int(len(aligned))


def stream_existing_snapshot_before_cutoff(
    writer: pq.ParquetWriter,
    source_key: str,
    existing_path: Path,
    cutoff_date: date,
    columns: list[str],
    schema: pa.Schema,
) -> int:
    row_count = 0
    datetime_column = source_primary_datetime_column(source_key)
    cutoff_ts = pd.Timestamp(cutoff_date, tz="UTC")
    parquet_file = pq.ParquetFile(existing_path)
    for batch in parquet_file.iter_batches(batch_size=25000):
        frame = batch.to_pandas()
        if datetime_column in frame.columns:
            dates = pd.to_datetime(frame[datetime_column], errors="coerce", utc=True)
            frame = frame.loc[dates.isna() | dates.lt(cutoff_ts)].copy()
        row_count += append_frame_to_parquet_writer(
            writer,
            source_key,
            frame,
            columns,
            schema,
        )
        del frame
        gc.collect()
    return row_count


def stream_chunk_file(
    writer: pq.ParquetWriter,
    source_key: str,
    chunk_path: Path,
    columns: list[str],
    schema: pa.Schema,
) -> int:
    row_count = 0
    parquet_file = pq.ParquetFile(chunk_path)
    for batch in parquet_file.iter_batches(batch_size=25000):
        frame = batch.to_pandas()
        row_count += append_frame_to_parquet_writer(
            writer,
            source_key,
            frame,
            columns,
            schema,
        )
        del frame
        gc.collect()
    return row_count


def source_snapshot_latest_date(source_key: str, snapshot_path: Path) -> date | None:
    datetime_column = source_primary_datetime_column(source_key)
    try:
        parquet_file = pq.ParquetFile(snapshot_path)
        if datetime_column not in parquet_file.schema.names:
            return None
        latest: pd.Timestamp | None = None
        for batch in parquet_file.iter_batches(columns=[datetime_column], batch_size=100000):
            values = pd.to_datetime(batch.column(0).to_pandas(), errors="coerce", utc=True)
            if values.notna().any():
                batch_max = values.max()
                if latest is None or batch_max > latest:
                    latest = batch_max
        return latest.date() if latest is not None else None
    except Exception:
        return None


def snapshot_generation() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S%fZ") + f"-{os.getpid()}"


def cleanup_old_source_generations(source_key: str) -> None:
    try:
        files = sorted(
            SNAPSHOT_DIR.glob(f"{source_key}_prepared_*.parquet"),
            key=lambda path: path.stat().st_mtime,
            reverse=True,
        )
        for path in files[ATLAS_SNAPSHOT_GENERATIONS_TO_KEEP:]:
            try:
                path.unlink()
            except OSError:
                pass
        for temp_path in SNAPSHOT_DIR.glob(f".refresh_{source_key}_*"):
            try:
                if temp_path.is_dir() and time.time() - temp_path.stat().st_mtime > 3600:
                    for child in temp_path.iterdir():
                        child.unlink(missing_ok=True)
                    temp_path.rmdir()
            except OSError:
                pass
    except Exception:
        return


def cleanup_legacy_source_files(source_key: str) -> None:
    """Remove superseded fixed-name snapshots only after a new manifest is live."""
    try:
        config = SOURCE_CONFIGS[source_key]
        for key in ["snapshot_file", "metadata_file"]:
            legacy_path = Path(config[key])
            if legacy_path.is_file():
                legacy_path.unlink()
    except OSError:
        pass


def publish_source_snapshot(
    source_key: str,
    existing_manifest: dict[str, Any] | None,
    refresh_start_date: date,
    work_dir: Path,
    chunk_files: list[Path],
    fresh_columns: list[str],
    refresh_metadata: dict[str, Any],
    signature: dict[str, Any],
) -> tuple[dict[str, Any], dict[str, Any]]:
    existing_path: Path | None = None
    if existing_manifest is not None:
        candidate = SNAPSHOT_DIR / str(existing_manifest.get("prepared_file", ""))
        if candidate.is_file():
            existing_path = candidate

    columns = source_snapshot_columns(source_key, existing_path, fresh_columns)
    schema = source_arrow_schema(source_key, columns)
    generation = snapshot_generation()
    final_path = source_snapshot_path(source_key, generation)
    temp_path = final_path.with_name(f"{final_path.name}.tmp")
    temp_path.parent.mkdir(parents=True, exist_ok=True)
    writer = pq.ParquetWriter(temp_path, schema, compression="zstd")
    total_rows = 0
    try:
        if existing_path is not None:
            total_rows += stream_existing_snapshot_before_cutoff(
                writer,
                source_key,
                existing_path,
                refresh_start_date,
                columns,
                schema,
            )
        for chunk_path in chunk_files:
            total_rows += stream_chunk_file(
                writer,
                source_key,
                chunk_path,
                columns,
                schema,
            )
    finally:
        writer.close()

    try:
        if not temp_path.is_file() or temp_path.stat().st_size <= 0:
            raise RuntimeError(f"{SOURCE_CONFIGS[source_key]['label']} prepared snapshot was not created.")
        os.replace(str(temp_path), str(final_path))
    finally:
        if temp_path.exists():
            try:
                temp_path.unlink()
            except OSError:
                pass

    latest_date = source_snapshot_latest_date(source_key, final_path)
    loaded_at_utc = datetime.now(timezone.utc)
    metadata = dict(refresh_metadata)
    metadata.update(
        {
            "source": SOURCE_CONFIGS[source_key]["label"],
            "endpoint": str(SOURCE_CONFIGS[source_key]["endpoint"]),
            "loaded_at_utc": loaded_at_utc.strftime("%d-%m-%Y %H:%M:%S UTC"),
            "loaded_at_local": local_time_label(loaded_at_utc),
            "loaded_start_date": API_FULL_START_DATE.isoformat(),
            "rows": int(total_rows),
            "kept_rows": int(total_rows),
            "columns": int(len(columns)),
            "snapshot_generation": generation,
            "snapshot_format": "prepared_parquet",
            "snapshot_schema_version": ATLAS_SNAPSHOT_SCHEMA_VERSION,
            "latest_source_date": latest_date.isoformat() if latest_date else refresh_metadata.get("latest_source_date"),
        }
    )
    saved_at_utc = loaded_at_utc.strftime("%d-%m-%Y %H:%M:%S UTC")
    manifest = {
        "snapshot_schema_version": ATLAS_SNAPSHOT_SCHEMA_VERSION,
        "source": source_key,
        "generation": generation,
        "prepared_file": final_path.name,
        "signature": signature,
        "metadata": metadata,
        "saved_at_utc": saved_at_utc,
    }
    _atomic_write_text(
        source_manifest_path(source_key),
        json.dumps(manifest, indent=2, default=str),
    )

    cached_read_prepared_source_snapshot.clear()
    if source_key == "reportdata":
        cached_prepare_long_data.clear()
        build_pivot_table.clear()
    else:
        clear_wide_source_state(source_key)
    cleanup_old_source_generations(source_key)
    cleanup_legacy_source_files(source_key)
    update_source_refresh_status(
        source_key,
        state="complete",
        stage="published",
        refresh_mode=metadata.get("refresh_mode"),
        rows_kept=total_rows,
        snapshot_generation=generation,
    )
    return metadata, manifest


def remove_refresh_work_dir(work_dir: Path | None) -> None:
    if work_dir is None:
        return
    try:
        if work_dir.is_dir():
            for child in work_dir.iterdir():
                try:
                    child.unlink()
                except OSError:
                    pass
            work_dir.rmdir()
    except OSError:
        pass


def refresh_source_snapshot(
    source_key: str,
    username: str,
    password: str,
    token: str,
    auth_method: str,
    *,
    full_refresh: bool,
    acquire_lock: bool = True,
) -> tuple[dict[str, Any], dict[str, Any]]:
    if source_key not in SOURCE_CONFIGS:
        raise ValueError(f"Unsupported AtlasFlow source: {source_key}")

    requested_signature = atlas_source_signature(
        source_key,
        username,
        auth_method,
        API_FULL_START_DATE,
    )

    def execute_refresh() -> tuple[dict[str, Any], dict[str, Any]]:
        existing_manifest = read_source_manifest(source_key)
        if not source_manifest_is_valid(
            source_key,
            existing_manifest,
            requested_signature,
            API_FULL_START_DATE,
        ):
            existing_manifest = None

        refresh_mode = "full"
        refresh_start_date = API_FULL_START_DATE
        if not full_refresh and existing_manifest is not None:
            existing_path = SNAPSHOT_DIR / str(existing_manifest["prepared_file"])
            latest_date_text = (existing_manifest.get("metadata") or {}).get("latest_source_date")
            try:
                latest_date = date.fromisoformat(str(latest_date_text)) if latest_date_text else None
            except ValueError:
                latest_date = None
            if latest_date is None:
                latest_date = source_snapshot_latest_date(source_key, existing_path)
            if latest_date is not None:
                overlap_days = read_int_secret(
                    source_secret_name(source_key, "INCREMENTAL_OVERLAP_DAYS"),
                    DEFAULT_SOURCE_OVERLAP_DAYS[source_key],
                    minimum=1,
                    maximum=90,
                )
                refresh_start_date = max(API_FULL_START_DATE, latest_date - timedelta(days=overlap_days))
                refresh_mode = "incremental"

        chunk_days = read_int_secret(
            source_secret_name(source_key, "REFRESH_CHUNK_DAYS"),
            DEFAULT_SOURCE_CHUNK_DAYS[source_key],
            minimum=1,
            maximum=62,
        )
        default_minutes = (
            DEFAULT_SOURCE_FULL_REFRESH_MAX_MINUTES[source_key]
            if refresh_mode == "full"
            else DEFAULT_SOURCE_INCREMENTAL_REFRESH_MAX_MINUTES[source_key]
        )
        max_minutes = read_int_secret(
            source_secret_name(
                source_key,
                "FULL_REFRESH_MAX_MINUTES" if refresh_mode == "full" else "INCREMENTAL_REFRESH_MAX_MINUTES",
            ),
            default_minutes,
            minimum=5,
            maximum=720,
        )
        end_date_exclusive = date.today() + timedelta(days=1)
        work_dir: Path | None = None
        try:
            update_source_refresh_status(
                source_key,
                state="running",
                stage="starting",
                refresh_mode=refresh_mode,
                refresh_start_date=refresh_start_date.isoformat(),
                end_date_exclusive=end_date_exclusive.isoformat(),
                chunk_days=chunk_days,
                max_minutes=max_minutes,
            )
            work_dir, chunk_files, fresh_columns, refresh_metadata = collect_source_chunks(
                source_key,
                username,
                password,
                token,
                auth_method,
                refresh_start_date,
                end_date_exclusive,
                chunk_days,
                max_minutes * 60,
                refresh_mode,
            )
            if (
                int(refresh_metadata.get("scanned_rows", 0) or 0) == 0
                or int(refresh_metadata.get("refresh_rows", 0) or 0) == 0
            ):
                if existing_manifest is not None:
                    metadata = dict(existing_manifest.get("metadata") or {})
                    metadata.update(
                        {
                            "refresh_mode": "no_changes",
                            "refresh_api_start_date": refresh_start_date.isoformat(),
                            "refresh_checked_at_local": local_time_label(),
                        }
                    )
                    update_source_refresh_status(source_key, state="complete", stage="no_changes", refresh_mode="no_changes")
                    return metadata, existing_manifest
                raise RuntimeError(f"{SOURCE_CONFIGS[source_key]['label']} returned zero usable rows during initial bootstrap.")

            update_source_refresh_status(
                source_key,
                state="running",
                stage="publishing",
                refresh_mode=refresh_mode,
                pages_completed=int(refresh_metadata.get("pages", 0) or 0),
                rows_kept=int(refresh_metadata.get("refresh_rows", 0) or 0),
            )
            metadata, manifest = publish_source_snapshot(
                source_key,
                existing_manifest if refresh_mode == "incremental" else None,
                refresh_start_date,
                work_dir,
                chunk_files,
                fresh_columns,
                refresh_metadata,
                requested_signature,
            )
            return metadata, manifest
        except Exception as exc:
            update_source_refresh_status(
                source_key,
                state="failed",
                stage="failed",
                refresh_mode=refresh_mode,
                error=str(exc),
            )
            raise
        finally:
            remove_refresh_work_dir(work_dir)

    if not acquire_lock:
        return execute_refresh()

    with source_refresh_lock(source_key) as lock_acquired:
        if not lock_acquired:
            existing = source_snapshot_info(source_key, username, auth_method, API_FULL_START_DATE)
            if existing is not None:
                metadata, manifest = existing
                metadata = dict(metadata)
                metadata["refresh_skipped_due_to_lock"] = True
                metadata["refresh_status"] = source_refresh_status_summary(source_key)
                return metadata, manifest
            raise AtlasRefreshAlreadyRunning(source_refresh_status_summary(source_key))
        return execute_refresh()


def migrate_legacy_source_snapshot(
    source_key: str,
    username: str,
    auth_method: str,
) -> tuple[dict[str, Any], dict[str, Any]] | None:
    config = SOURCE_CONFIGS[source_key]
    legacy_file = Path(config["snapshot_file"])
    legacy_metadata_file = Path(config["metadata_file"])
    if not legacy_file.is_file() or not legacy_metadata_file.is_file():
        return None
    try:
        legacy_df = pd.read_parquet(legacy_file)
        if source_key == "reportdata":
            prepared = prepare_reportdata_snapshot_frame(legacy_df)
        else:
            prepared = normalize_wide_snapshot_frame(source_key, legacy_df)
        if prepared.empty:
            return None
        work_dir = SNAPSHOT_DIR / f".refresh_{source_key}_migration_{os.getpid()}_{int(time.time() * 1000)}"
        work_dir.mkdir(parents=True, exist_ok=False)
        chunk_path = work_dir / "chunk_0001.parquet"
        write_temp_chunk(prepared, chunk_path)
        signature = atlas_source_signature(source_key, username, auth_method, API_FULL_START_DATE)
        metadata = {
            "refresh_mode": "legacy_migration",
            "refresh_api_start_date": API_FULL_START_DATE.isoformat(),
            "refresh_end_date_exclusive": (date.today() + timedelta(days=1)).isoformat(),
            "chunk_days": 0,
            "chunks_total": 1,
            "chunks_completed": 1,
            "refresh_rows": int(len(prepared)),
            "scanned_rows": int(len(prepared)),
            "pages": 0,
            "downloaded_mb": 0,
            "fetch_seconds": 0,
            "first_url": "legacy_snapshot",
            "paging_stop_reason": "legacy_migration",
            "hit_page_limit": False,
            "latest_source_date": None,
        }
        published = publish_source_snapshot(
            source_key,
            None,
            API_FULL_START_DATE,
            work_dir,
            [chunk_path],
            list(prepared.columns),
            metadata,
            signature,
        )
        remove_refresh_work_dir(work_dir)
        return published
    except Exception:
        return None


def ensure_source_snapshot(
    source_key: str,
    username: str,
    auth_method: str,
) -> tuple[dict[str, Any], dict[str, Any]] | None:
    existing = source_snapshot_info(source_key, username, auth_method, API_FULL_START_DATE)
    if existing is not None:
        return existing
    with source_refresh_lock(source_key) as lock_acquired:
        if not lock_acquired:
            return source_snapshot_info(source_key, username, auth_method, API_FULL_START_DATE)
        existing = source_snapshot_info(source_key, username, auth_method, API_FULL_START_DATE)
        if existing is not None:
            return existing
        return migrate_legacy_source_snapshot(source_key, username, auth_method)


def parse_wide_source_datetimes(df: pd.DataFrame) -> pd.DataFrame:
    """Prepared snapshots already type the primary filter datetime column."""
    if df.empty:
        return df
    parsed_df = df
    for column in parsed_df.columns:
        if pd.api.types.is_datetime64_any_dtype(parsed_df[column]):
            continue
        lower = str(column).lower()
        if "datetime" in lower or lower in {"date", "timestamp"}:
            parsed = parse_datetime_series(parsed_df[column])
            if parsed.notna().any():
                if parsed_df is df:
                    parsed_df = df.copy()
                parsed_df[column] = parsed
    return parsed_df


def get_loaded_state() -> tuple[pd.DataFrame | None, pd.DataFrame | None, dict[str, Any] | None]:
    return (
        None,
        st.session_state.get("loaded_long_df"),
        st.session_state.get("loaded_metadata"),
    )


def activate_reportdata_snapshot(
    username: str,
    auth_method: str,
    start_date: date,
) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, Any]]:
    signature = atlas_source_signature("reportdata", username, auth_method, start_date)
    snapshot = load_source_snapshot("reportdata", signature, start_date)
    if snapshot is None:
        raise FileNotFoundError("The prepared ReportData snapshot could not be loaded.")
    long_df, metadata, snapshot_signature = snapshot
    st.session_state.pop("loaded_raw_df", None)
    st.session_state["loaded_long_df"] = long_df
    st.session_state["loaded_metadata"] = dict(metadata)
    st.session_state["loaded_request_signature"] = snapshot_signature
    st.session_state["loaded_prepare_signature"] = source_data_signature("reportdata")
    st.session_state["loaded_reportdata_generation"] = metadata.get("snapshot_generation")
    return pd.DataFrame(), long_df, metadata


def load_or_fetch_source(
    source_key: str,
    username: str,
    password: str,
    token: str,
    auth_method: str,
    start_date: date,
    refresh: bool,
    auto_fetch: bool = False,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    del auto_fetch
    requested_signature = atlas_source_signature(source_key, username, auth_method, start_date)
    state_df_key = f"loaded_{source_key}_df"
    state_meta_key = f"loaded_{source_key}_metadata"
    state_sig_key = f"loaded_{source_key}_signature"
    state_generation_key = f"loaded_{source_key}_generation"

    if refresh:
        refresh_source_snapshot(
            source_key,
            username,
            password,
            token,
            auth_method,
            full_refresh=False,
        )

    manifest = read_source_manifest(source_key)
    current_generation = manifest.get("generation") if isinstance(manifest, dict) else None
    df = st.session_state.get(state_df_key)
    metadata = st.session_state.get(state_meta_key)
    current_signature = st.session_state.get(state_sig_key)
    session_generation = st.session_state.get(state_generation_key)
    session_ready = (
        isinstance(df, pd.DataFrame)
        and isinstance(metadata, dict)
        and source_signature_covers_request(current_signature, requested_signature, metadata, start_date)
        and session_generation == current_generation
    )
    if session_ready:
        return df, metadata

    snapshot = load_source_snapshot(source_key, requested_signature, start_date)
    if snapshot is None:
        migrated = ensure_source_snapshot(source_key, username, auth_method)
        if migrated is not None:
            snapshot = load_source_snapshot(source_key, requested_signature, start_date)
    if snapshot is None:
        config = SOURCE_CONFIGS[source_key]
        return pd.DataFrame(), {
            "source": config["label"],
            "endpoint": str(config["endpoint"]),
            "loaded_at_utc": "-",
            "loaded_at_local": "No prepared snapshot yet",
            "loaded_from_snapshot": False,
            "rows": 0,
            "columns": 0,
            "pages": 0,
            "first_url": "-",
            "needs_warmup": True,
        }

    df, metadata, snapshot_signature = snapshot
    st.session_state[state_df_key] = df
    st.session_state[state_meta_key] = metadata
    st.session_state[state_sig_key] = snapshot_signature
    st.session_state[state_generation_key] = metadata.get("snapshot_generation")
    return df, metadata


def load_raw_snapshot(
    requested_signature: dict[str, Any],
    requested_start_date: date,
) -> tuple[pd.DataFrame, dict[str, Any], dict[str, Any]] | None:
    requested_signature = dict(requested_signature or {})
    requested_signature.setdefault("source", "reportdata")
    requested_signature.setdefault("data_signature", source_data_signature("reportdata"))
    return load_source_snapshot("reportdata", requested_signature, requested_start_date)


def refresh_all_atlasflow_snapshots(
    username: str,
    password: str,
    token: str,
    auth_method: str,
    start_date: date,
) -> dict[str, dict[str, Any]]:
    del start_date
    results: dict[str, dict[str, Any]] = {}
    errors: dict[str, str] = {}
    for source_key in ["reportdata", "reportpivots", "shippivots"]:
        try:
            metadata, _ = refresh_source_snapshot(
                source_key,
                username,
                password,
                token,
                auth_method,
                full_refresh=False,
            )
            results[source_key] = metadata
        except Exception as exc:
            errors[source_key] = str(exc)
        finally:
            gc.collect()
    if errors:
        raise RuntimeError(
            "AtlasFlow source refresh failures: "
            + "; ".join(f"{key}: {value}" for key, value in errors.items())
        )
    cached_fetch_report_data.clear()
    cached_fetch_wide_odata_source.clear()
    activate_reportdata_snapshot(username, auth_method, API_FULL_START_DATE)
    return results



# =============================================================================
# Monthly partitioned storage and cross-source monthly comparison (v2)
# =============================================================================

# The original prepared-snapshot layer is retained for backwards compatibility,
# but this version publishes immutable monthly partitions. Incremental warmups
# rewrite only the months touched by the overlap window. Monthly summaries are
# generated at publish time, so all vessels and full months can be compared
# without opening millions of 15-minute rows in a browser session.
ATLAS_SNAPSHOT_SCHEMA_VERSION = "2026-07-15-monthly-partitioned-comparison-v2"
ATLAS_MONTHLY_COMPARISON_VERSION = "2026-07-15-cross-source-monthly-v1"
ATLAS_RAW_INTERACTIVE_ROW_LIMIT = 50_000
ATLAS_REPORTPIVOTS_INTERACTIVE_ROW_LIMIT = 250_000

COMPARISON_SOURCE_LABELS = {
    "reportdata": "ReportData",
    "reportpivots": "ReportPivots",
    "shippivots": "ShipPivots",
}

MONTHLY_COMPARISON_METRICS: dict[str, dict[str, Any]] = {
    "Average Speed [kn]": {
        "aggregation": "mean",
        "candidates": [
            "GPSSpeed", "GPS Speed [kn]", "GPS Speed", "Speed Over Ground [kn]", "Speed Over Ground", "SOG",
            "LogSpeed", "Log Speed [kn]", "Log Speed", "Speed Through Water [kn]", "Speed Through Water", "STW",
            "Average Speed", "Vessel Speed",
        ],
    },
    "Average Shaft Power [kW]": {
        "aggregation": "mean",
        "candidates": [
            "ShaftPower", "Shaft Power", "Power from Torque Meter [kW]",
            "Total Shaft Power [kW]", "Total Shaft Power [kW] (kW)",
            "ME Power", "Main Engine Power",
        ],
    },
    "Average ME Load [%]": {
        "aggregation": "mean",
        "percentage": True,
        "candidates": [
            "ME Load [%MCR]", "ME Load [% MCR]", "MELoad",
            "ME Load", "Main Engine Load", "MainEngineLoad",
        ],
    },
    "Total ME Fuel [MT]": {
        "aggregation": "sum",
        "candidates": [
            "Main Engine Total Consumed", "ME Total Consumed",
            "MEConsumed", "Main Engine Consumption", "ME Consumption",
        ],
        "component_candidates": ME_FUEL_COLUMNS,
    },
    "Total DG Fuel [MT]": {
        "aggregation": "sum",
        "candidates": [
            "Diesel Generator Total Consumed", "DG Total Consumed",
            "DG Totals Consumed", "DGTotalsConsumed", "DGTotalConsumed",
            "DGConsumed", "Generator Total Consumed",
        ],
        "component_candidates": DG_FUEL_COLUMNS,
    },
    "Total Auxiliary Fuel [MT]": {
        "aggregation": "sum",
        "candidates": [
            "Auxiliary Engine Total Consumed", "Aux Engine Total Consumed",
            "Aux Total Consumed", "AuxConsumed",
        ],
        "component_candidates": AUXILIARY_FUEL_COLUMNS,
    },
    "Total Boiler Fuel [MT]": {
        "aggregation": "sum",
        "candidates": [
            "Boiler Total Consumed", "BoilerConsumed", "Boiler Consumption",
        ],
        "component_candidates": BOILER_FUEL_COLUMNS,
    },
    "Total Fuel [MT]": {
        "aggregation": "sum",
        "candidates": [
            "Total Fuel Consumed", "Total Consumed", "Total Consumption",
            "Bunker Consumption", "Fuel Consumption", "FuelConsumed",
        ],
    },
    "Total Distance [nm]": {
        "aggregation": "sum",
        "candidates": [
            "Distance Over Ground [nm]", "DistanceOverGround", "Distance Over Ground",
            "Engine Distance [nm]", "EngineDistance", "Sailed Distance", "Distance",
        ],
    },
    "Total Running Hours [h]": {
        "aggregation": "sum",
        "candidates": [
            "Steaming Time Since Last Report [hh:mm]",
            "Steaming Time Since Last Report", "RunningHours", "Running Hours",
            "LapTime", "Operating Hours",
        ],
    },
    "Average SFOC [g/kWh]": {
        "aggregation": "mean",
        "candidates": ["SFOC [g/kWh]", "SFOC [gr/Kwh]", "SFOC"],
    },
}


def source_partition_root(source_key: str) -> Path:
    return SNAPSHOT_DIR / "monthly" / source_key


def source_partition_file(source_key: str, month_key: str, generation: str) -> Path:
    return source_partition_root(source_key) / f"data_{month_key.replace('-', '')}_{generation}.parquet"


def source_summary_file(source_key: str, month_key: str, generation: str) -> Path:
    return source_partition_root(source_key) / f"summary_{month_key.replace('-', '')}_{generation}.parquet"


def month_start_from_key(month_key: str) -> date:
    year_text, month_text = month_key.split("-", 1)
    return date(int(year_text), int(month_text), 1)


def next_month_start(value: date) -> date:
    return date(value.year + 1, 1, 1) if value.month == 12 else date(value.year, value.month + 1, 1)


def month_keys_for_range(start_value: date, end_exclusive: date) -> list[str]:
    if start_value >= end_exclusive:
        return []
    cursor = date(start_value.year, start_value.month, 1)
    keys: list[str] = []
    while cursor < end_exclusive:
        keys.append(f"{cursor.year:04d}-{cursor.month:02d}")
        cursor = next_month_start(cursor)
    return keys


def manifest_partitions(manifest: dict[str, Any] | None) -> dict[str, dict[str, Any]]:
    if not isinstance(manifest, dict) or not isinstance(manifest.get("partitions"), dict):
        return {}
    return {
        str(month_key): dict(entry)
        for month_key, entry in manifest["partitions"].items()
        if isinstance(entry, dict)
    }


def partition_entry_path(entry: dict[str, Any], field: str = "file") -> Path:
    return SNAPSHOT_DIR / str(entry.get(field, ""))


def source_manifest_is_valid(
    source_key: str,
    manifest: dict[str, Any] | None,
    requested_signature: dict[str, Any],
    requested_start_date: date,
) -> bool:
    if not manifest:
        return False
    if manifest.get("snapshot_schema_version") != ATLAS_SNAPSHOT_SCHEMA_VERSION:
        return False
    if manifest.get("source") != source_key:
        return False
    metadata = manifest.get("metadata") or {}
    stored_signature = manifest.get("signature") or {}
    if not source_signature_covers_request(
        stored_signature,
        requested_signature,
        metadata,
        requested_start_date,
    ):
        return False
    partitions = manifest_partitions(manifest)
    if not partitions:
        return False
    for entry in partitions.values():
        if not partition_entry_path(entry, "file").is_file():
            return False
        if entry.get("summary_file") and not partition_entry_path(entry, "summary_file").is_file():
            return False
    return True


@st.cache_data(show_spinner=False)
def cached_read_partitioned_source_snapshot(
    source_key: str,
    generation: str,
    partition_files: tuple[str, ...],
) -> pd.DataFrame:
    del source_key, generation
    frames = [pd.read_parquet(file_name) for file_name in partition_files]
    frames = [frame for frame in frames if isinstance(frame, pd.DataFrame) and not frame.empty]
    return pd.concat(frames, ignore_index=True, sort=False) if frames else pd.DataFrame()


@st.cache_data(show_spinner=False)
def cached_read_monthly_summary_files(
    generation_signature: str,
    summary_files: tuple[str, ...],
) -> pd.DataFrame:
    del generation_signature
    frames = [pd.read_parquet(file_name) for file_name in summary_files]
    frames = [frame for frame in frames if isinstance(frame, pd.DataFrame) and not frame.empty]
    return pd.concat(frames, ignore_index=True, sort=False) if frames else pd.DataFrame()


def load_source_snapshot(
    source_key: str,
    requested_signature: dict[str, Any],
    requested_start_date: date,
) -> tuple[pd.DataFrame, dict[str, Any], dict[str, Any]] | None:
    requested_signature = dict(requested_signature or {})
    requested_signature.setdefault("source", source_key)
    requested_signature.setdefault("data_signature", source_data_signature(source_key))
    manifest = read_source_manifest(source_key)
    if not source_manifest_is_valid(source_key, manifest, requested_signature, requested_start_date):
        return None
    assert manifest is not None
    partitions = manifest_partitions(manifest)
    total_rows = int((manifest.get("metadata") or {}).get("rows", 0) or 0)
    # Wide sources are intentionally read through load_wide_source_for_view(),
    # which applies predicate pushdown. Refuse a full wide-source materialization.
    if source_key != "reportdata" and total_rows > ATLAS_REPORTPIVOTS_INTERACTIVE_ROW_LIMIT:
        return None
    files = tuple(
        str(partition_entry_path(partitions[month_key], "file"))
        for month_key in sorted(partitions)
    )
    try:
        frame = cached_read_partitioned_source_snapshot(
            source_key,
            str(manifest.get("generation", "")),
            files,
        )
    except Exception:
        return None
    metadata = dict(manifest.get("metadata") or {})
    metadata["loaded_from_snapshot"] = True
    metadata["snapshot_generation"] = manifest.get("generation")
    metadata.setdefault("snapshot_saved_at_utc", manifest.get("saved_at_utc", "-"))
    metadata["partition_count"] = len(partitions)
    return frame, metadata, dict(manifest.get("signature") or {})


def source_snapshot_info(
    source_key: str,
    username: str,
    auth_method: str,
    start_date: date,
) -> tuple[dict[str, Any], dict[str, Any]] | None:
    signature = atlas_source_signature(source_key, username, auth_method, start_date)
    manifest = read_source_manifest(source_key)
    if not source_manifest_is_valid(source_key, manifest, signature, start_date):
        return None
    assert manifest is not None
    return dict(manifest.get("metadata") or {}), manifest


def _candidate_column(df: pd.DataFrame, candidates: list[str]) -> str | None:
    normalized_columns = {normalize_text(column): str(column) for column in df.columns}
    for candidate in candidates:
        match = normalized_columns.get(normalize_text(candidate))
        if match is not None:
            return match
    for candidate in candidates:
        candidate_key = normalize_text(candidate)
        if len(candidate_key) < 6:
            continue
        for normalized_column, original_column in normalized_columns.items():
            if candidate_key in normalized_column or normalized_column in candidate_key:
                return original_column
    return None


def _reportdata_metric_rows(
    frame: pd.DataFrame,
    spec: dict[str, Any],
) -> tuple[pd.DataFrame, str | None]:
    if frame.empty or "ValueDescription" not in frame.columns:
        return pd.DataFrame(), None
    description_keys = frame["ValueDescription"].map(normalize_text)
    for candidate in list(spec.get("candidates") or []):
        mask = description_keys.eq(normalize_text(candidate))
        if mask.any():
            selected = frame.loc[mask, ["ShipName", "ReportId", "ParsedValue"]].copy()
            selected["MetricValue"] = pd.to_numeric(selected["ParsedValue"], errors="coerce")
            return selected, candidate
    components = list(spec.get("component_candidates") or [])
    component_keys = {normalize_text(value) for value in components}
    if component_keys:
        mask = description_keys.isin(component_keys)
        if mask.any():
            selected = frame.loc[mask, ["ShipName", "ReportId", "ParsedValue"]].copy()
            selected["MetricValue"] = pd.to_numeric(selected["ParsedValue"], errors="coerce")
            selected = (
                selected.groupby(["ShipName", "ReportId"], dropna=False, as_index=False)["MetricValue"]
                .sum(min_count=1)
            )
            return selected, " + ".join(components)
    return pd.DataFrame(), None


def _period_day_count(month_key: str) -> int:
    month_start = month_start_from_key(month_key)
    end_exclusive = min(next_month_start(month_start), date.today() + timedelta(days=1))
    return max((end_exclusive - month_start).days, 1)


def build_monthly_source_summary(
    source_key: str,
    partition_df: pd.DataFrame,
    month_key: str,
) -> pd.DataFrame:
    if partition_df.empty:
        return pd.DataFrame()
    datetime_column = source_primary_datetime_column(source_key)
    if datetime_column not in partition_df.columns or "ShipName" not in partition_df.columns:
        return pd.DataFrame()
    frame = partition_df.copy()
    frame[datetime_column] = pd.to_datetime(frame[datetime_column], errors="coerce", utc=True)
    frame = frame[frame[datetime_column].notna() & frame["ShipName"].notna()].copy()
    if frame.empty:
        return pd.DataFrame()
    frame["ShipName"] = frame["ShipName"].astype("string")
    frame["_ObservedDate"] = frame[datetime_column].dt.date
    grouped = frame.groupby("ShipName", dropna=False)
    records = grouped["ReportId"].nunique(dropna=True) if source_key == "reportdata" and "ReportId" in frame.columns else grouped.size()
    first_timestamp = grouped[datetime_column].min().reindex(records.index)
    last_timestamp = grouped[datetime_column].max().reindex(records.index)
    observed_days = grouped["_ObservedDate"].nunique(dropna=True).reindex(records.index)
    summary = pd.DataFrame(
        {
            "ShipName": records.index.astype(str),
            "Records": records.to_numpy(),
            "First Timestamp": first_timestamp.to_numpy(),
            "Last Timestamp": last_timestamp.to_numpy(),
            "Observed Days": observed_days.to_numpy(),
        }
    )
    summary.insert(0, "Source", COMPARISON_SOURCE_LABELS[source_key])
    summary.insert(0, "Month", month_key)
    period_days = _period_day_count(month_key)
    summary["Period Days"] = period_days
    summary["Month Complete"] = next_month_start(month_start_from_key(month_key)) <= date.today()
    summary["Day Coverage [%]"] = (
        pd.to_numeric(summary["Observed Days"], errors="coerce") / period_days * 100
    ).clip(upper=100).round(2)
    if source_key == "shippivots":
        unique_timestamps = grouped[datetime_column].nunique(dropna=True).reindex(records.index)
        observation_coverage = pd.Series(
            unique_timestamps.to_numpy() / max(period_days * 24 * 4, 1) * 100,
            index=summary.index,
            dtype="float64",
        )
        summary["Observation Coverage [%]"] = observation_coverage.clip(upper=100).round(2)
    else:
        summary["Observation Coverage [%]"] = pd.NA

    for metric_name, spec in MONTHLY_COMPARISON_METRICS.items():
        mapping_column = f"Mapping: {metric_name}"
        aggregation = str(spec.get("aggregation", "mean"))
        if source_key == "reportdata":
            metric_rows, mapping = _reportdata_metric_rows(frame, spec)
            if metric_rows.empty:
                summary[metric_name] = pd.NA
                summary[mapping_column] = pd.NA
                continue
            metric_group = metric_rows.groupby("ShipName", dropna=False)["MetricValue"]
        else:
            source_column = _candidate_column(frame, list(spec.get("candidates") or []))
            if source_column is None:
                summary[metric_name] = pd.NA
                summary[mapping_column] = pd.NA
                continue
            metric_rows = pd.DataFrame(
                {
                    "ShipName": frame["ShipName"],
                    "MetricValue": pd.to_numeric(frame[source_column], errors="coerce"),
                }
            )
            metric_group = metric_rows.groupby("ShipName", dropna=False)["MetricValue"]
            mapping = source_column
        aggregated = metric_group.sum(min_count=1) if aggregation == "sum" else metric_group.mean()
        values = summary["ShipName"].map(aggregated)
        if spec.get("percentage"):
            numeric = pd.to_numeric(values, errors="coerce")
            non_null = numeric.dropna()
            if not non_null.empty and non_null.abs().median() <= 1.5:
                numeric = numeric * 100
            values = numeric
        summary[metric_name] = pd.to_numeric(values, errors="coerce").round(3)
        summary[mapping_column] = mapping

    component_total = pd.concat(
        [
            pd.to_numeric(summary.get("Total ME Fuel [MT]"), errors="coerce"),
            pd.to_numeric(summary.get("Total DG Fuel [MT]"), errors="coerce"),
            pd.to_numeric(summary.get("Total Auxiliary Fuel [MT]"), errors="coerce"),
            pd.to_numeric(summary.get("Total Boiler Fuel [MT]"), errors="coerce"),
        ],
        axis=1,
    ).sum(axis=1, min_count=1)
    direct_total = pd.to_numeric(summary.get("Total Fuel [MT]"), errors="coerce")
    summary["Total Fuel [MT]"] = direct_total.fillna(component_total).round(3)
    return summary.sort_values("ShipName").reset_index(drop=True)


def _source_partition_columns(
    source_key: str,
    existing_manifest: dict[str, Any] | None,
    fresh_columns: list[str],
) -> list[str]:
    columns: list[str] = []
    for entry in manifest_partitions(existing_manifest).values():
        path = partition_entry_path(entry, "file")
        if not path.is_file():
            continue
        try:
            for column in pq.ParquetFile(path).schema.names:
                if column not in columns:
                    columns.append(column)
        except Exception:
            continue
    for column in fresh_columns:
        if column not in columns:
            columns.append(column)
    if source_key == "reportdata":
        ordered = [*SOURCE_COLUMNS, "ParsedValue"]
        return ordered
    datetime_column = source_primary_datetime_column(source_key)
    if datetime_column not in columns:
        columns.insert(0, datetime_column)
    return columns or [datetime_column]


def _split_fresh_chunks_to_month_files(
    source_key: str,
    chunk_files: list[Path],
    work_dir: Path,
    columns: list[str],
) -> dict[str, Path]:
    schema = source_arrow_schema(source_key, columns)
    datetime_column = source_primary_datetime_column(source_key)
    writers: dict[str, pq.ParquetWriter] = {}
    month_paths: dict[str, Path] = {}
    try:
        for chunk_path in chunk_files:
            parquet_file = pq.ParquetFile(chunk_path)
            for batch in parquet_file.iter_batches(batch_size=25_000):
                frame = batch.to_pandas()
                if datetime_column not in frame.columns:
                    continue
                dates = pd.to_datetime(frame[datetime_column], errors="coerce", utc=True)
                valid = dates.notna()
                if not valid.any():
                    continue
                frame = frame.loc[valid].copy()
                month_values = dates.loc[valid].dt.strftime("%Y-%m")
                for month_key in month_values.dropna().unique().tolist():
                    subset = frame.loc[month_values.eq(month_key)].copy()
                    if subset.empty:
                        continue
                    aligned = align_source_frame(source_key, subset, columns)
                    table = pa.Table.from_pandas(aligned, schema=schema, preserve_index=False)
                    if month_key not in writers:
                        path = work_dir / f"fresh_{month_key.replace('-', '')}.parquet"
                        month_paths[month_key] = path
                        writers[month_key] = pq.ParquetWriter(path, schema, compression="zstd")
                    writers[month_key].write_table(table)
                    del subset, aligned, table
                del frame, dates, month_values
                gc.collect()
    finally:
        for writer in writers.values():
            writer.close()
    return month_paths


def _read_month_frame(
    source_key: str,
    path: Path | None,
    *,
    before_date: date | None = None,
) -> pd.DataFrame:
    if path is None or not path.is_file():
        return pd.DataFrame()
    frame = pd.read_parquet(path)
    if before_date is not None and not frame.empty:
        datetime_column = source_primary_datetime_column(source_key)
        if datetime_column in frame.columns:
            values = pd.to_datetime(frame[datetime_column], errors="coerce", utc=True)
            frame = frame.loc[
                values.isna() | values.lt(pd.Timestamp(before_date, tz="UTC"))
            ].copy()
    return frame


def _partition_file_metadata(
    source_key: str,
    month_key: str,
    data_path: Path,
    summary_path: Path,
    frame: pd.DataFrame,
    summary: pd.DataFrame,
) -> dict[str, Any]:
    datetime_column = source_primary_datetime_column(source_key)
    values = (
        pd.to_datetime(frame[datetime_column], errors="coerce", utc=True)
        if datetime_column in frame.columns
        else pd.Series(dtype="datetime64[ns, UTC]")
    )
    return {
        "month": month_key,
        "file": str(data_path.relative_to(SNAPSHOT_DIR)),
        "summary_file": str(summary_path.relative_to(SNAPSHOT_DIR)),
        "rows": int(len(frame)),
        "summary_rows": int(len(summary)),
        "columns": int(len(frame.columns)),
        "min_datetime": (
            values.min().isoformat()
            if not values.empty and values.notna().any()
            else None
        ),
        "max_datetime": (
            values.max().isoformat()
            if not values.empty and values.notna().any()
            else None
        ),
    }


def _cleanup_partition_files(source_key: str, current_manifest: dict[str, Any]) -> None:
    root = source_partition_root(source_key)
    if not root.is_dir():
        return
    referenced = {
        str(partition_entry_path(entry, field).resolve())
        for entry in manifest_partitions(current_manifest).values()
        for field in ["file", "summary_file"]
        if entry.get(field)
    }
    grouped: dict[str, list[Path]] = {}
    for path in root.glob("*.parquet"):
        parts = path.stem.split("_")
        group_key = "_".join(parts[:2]) if len(parts) >= 2 else path.stem
        grouped.setdefault(group_key, []).append(path)
    for paths in grouped.values():
        paths.sort(key=lambda item: item.stat().st_mtime, reverse=True)
        protected = {str(path.resolve()) for path in paths[:2]} | referenced
        for path in paths:
            if str(path.resolve()) not in protected:
                try:
                    path.unlink()
                except OSError:
                    pass


def publish_source_snapshot(
    source_key: str,
    existing_manifest: dict[str, Any] | None,
    refresh_start_date: date,
    work_dir: Path,
    chunk_files: list[Path],
    fresh_columns: list[str],
    refresh_metadata: dict[str, Any],
    signature: dict[str, Any],
) -> tuple[dict[str, Any], dict[str, Any]]:
    """Publish immutable monthly source partitions and monthly summaries."""
    existing_partitions = manifest_partitions(existing_manifest)
    columns = _source_partition_columns(source_key, existing_manifest, fresh_columns)
    generation = snapshot_generation()
    source_partition_root(source_key).mkdir(parents=True, exist_ok=True)
    fresh_month_files = _split_fresh_chunks_to_month_files(
        source_key,
        chunk_files,
        work_dir,
        columns,
    )
    try:
        end_exclusive = date.fromisoformat(
            str(refresh_metadata.get("refresh_end_date_exclusive"))
        )
    except ValueError:
        end_exclusive = date.today() + timedelta(days=1)
    affected_months = month_keys_for_range(refresh_start_date, end_exclusive)
    new_partitions = dict(existing_partitions)

    for month_key in affected_months:
        old_entry = existing_partitions.get(month_key)
        old_path = partition_entry_path(old_entry, "file") if old_entry else None
        month_start = month_start_from_key(month_key)
        before_date = (
            refresh_start_date
            if month_start <= refresh_start_date < next_month_start(month_start)
            else None
        )
        old_frame = _read_month_frame(source_key, old_path, before_date=before_date)
        fresh_frame = _read_month_frame(source_key, fresh_month_files.get(month_key))
        combined = pd.concat([old_frame, fresh_frame], ignore_index=True, sort=False)
        if combined.empty:
            new_partitions.pop(month_key, None)
            continue
        combined = normalize_source_snapshot_frame(source_key, combined)
        combined = deduplicate_source_window(source_key, combined)
        datetime_column = source_primary_datetime_column(source_key)
        sort_columns = [
            column for column in [datetime_column, "ShipName"]
            if column in combined.columns
        ]
        if sort_columns:
            combined = combined.sort_values(sort_columns)
        data_path = source_partition_file(source_key, month_key, generation)
        summary_path = source_summary_file(source_key, month_key, generation)
        combined.to_parquet(data_path, index=False, compression="zstd")
        summary = build_monthly_source_summary(source_key, combined, month_key)
        if summary.empty:
            summary = pd.DataFrame(
                columns=[
                    "Month", "Source", "ShipName", "Records",
                    "First Timestamp", "Last Timestamp", "Observed Days",
                    "Period Days", "Day Coverage [%]", "Observation Coverage [%]",
                ]
            )
        summary.to_parquet(summary_path, index=False, compression="zstd")
        new_partitions[month_key] = _partition_file_metadata(
            source_key,
            month_key,
            data_path,
            summary_path,
            combined,
            summary,
        )
        del old_frame, fresh_frame, combined, summary
        gc.collect()

    if not new_partitions:
        raise RuntimeError(
            f"{SOURCE_CONFIGS[source_key]['label']} produced no monthly partitions."
        )

    total_rows = sum(
        int(entry.get("rows", 0) or 0)
        for entry in new_partitions.values()
    )
    latest_values = [
        entry.get("max_datetime")
        for entry in new_partitions.values()
        if entry.get("max_datetime")
    ]
    latest_source_date = (
        max(pd.Timestamp(value) for value in latest_values).date().isoformat()
        if latest_values
        else None
    )
    loaded_at_utc = datetime.now(timezone.utc)
    metadata = dict(refresh_metadata)
    metadata.update(
        {
            "source": SOURCE_CONFIGS[source_key]["label"],
            "endpoint": str(SOURCE_CONFIGS[source_key]["endpoint"]),
            "loaded_at_utc": loaded_at_utc.strftime("%d-%m-%Y %H:%M:%S UTC"),
            "loaded_at_local": local_time_label(loaded_at_utc),
            "loaded_start_date": API_FULL_START_DATE.isoformat(),
            "rows": int(total_rows),
            "kept_rows": int(total_rows),
            "columns": int(len(columns)),
            "partition_count": int(len(new_partitions)),
            "snapshot_generation": generation,
            "snapshot_format": "monthly_partitioned_parquet",
            "snapshot_schema_version": ATLAS_SNAPSHOT_SCHEMA_VERSION,
            "comparison_schema_version": ATLAS_MONTHLY_COMPARISON_VERSION,
            "latest_source_date": latest_source_date,
        }
    )
    saved_at_utc = loaded_at_utc.strftime("%d-%m-%Y %H:%M:%S UTC")
    manifest = {
        "snapshot_schema_version": ATLAS_SNAPSHOT_SCHEMA_VERSION,
        "comparison_schema_version": ATLAS_MONTHLY_COMPARISON_VERSION,
        "source": source_key,
        "generation": generation,
        "signature": signature,
        "partitions": {
            key: new_partitions[key]
            for key in sorted(new_partitions)
        },
        "metadata": metadata,
        "saved_at_utc": saved_at_utc,
    }
    _atomic_write_text(
        source_manifest_path(source_key),
        json.dumps(manifest, indent=2, default=str),
    )

    cached_read_prepared_source_snapshot.clear()
    cached_read_partitioned_source_snapshot.clear()
    cached_read_monthly_summary_files.clear()
    if source_key == "reportdata":
        cached_prepare_long_data.clear()
        build_pivot_table.clear()
    clear_wide_source_state(source_key)
    _cleanup_partition_files(source_key, manifest)
    cleanup_legacy_source_files(source_key)
    update_source_refresh_status(
        source_key,
        state="complete",
        stage="published",
        refresh_mode=metadata.get("refresh_mode"),
        rows_kept=total_rows,
        partition_count=len(new_partitions),
        snapshot_generation=generation,
    )
    return metadata, manifest


def _legacy_source_file(source_key: str) -> Path | None:
    manifest = read_source_manifest(source_key)
    if isinstance(manifest, dict) and manifest.get("prepared_file"):
        candidate = SNAPSHOT_DIR / str(manifest["prepared_file"])
        if candidate.is_file():
            return candidate
    candidate = Path(SOURCE_CONFIGS[source_key]["snapshot_file"])
    return candidate if candidate.is_file() else None


def migrate_legacy_source_snapshot(
    source_key: str,
    username: str,
    auth_method: str,
) -> tuple[dict[str, Any], dict[str, Any]] | None:
    # ReportData is compacted through a ValueDescription whitelist. This release
    # adds comparison fields (including speed), so an old compact snapshot cannot
    # be considered historically complete. Force one full ReportData bootstrap;
    # wide ReportPivots/ShipPivots snapshots can be partition-migrated safely.
    if source_key == "reportdata":
        return None
    legacy_file = _legacy_source_file(source_key)
    if legacy_file is None:
        return None
    try:
        parquet_file = pq.ParquetFile(legacy_file)
        fresh_columns = list(parquet_file.schema.names)
        if not fresh_columns:
            return None
        work_dir = SNAPSHOT_DIR / (
            f".refresh_{source_key}_migration_{os.getpid()}_"
            f"{int(time.time() * 1000)}"
        )
        work_dir.mkdir(parents=True, exist_ok=False)
        signature = atlas_source_signature(
            source_key,
            username,
            auth_method,
            API_FULL_START_DATE,
        )
        metadata = {
            "refresh_mode": "legacy_partition_migration",
            "refresh_api_start_date": API_FULL_START_DATE.isoformat(),
            "refresh_end_date_exclusive": (
                date.today() + timedelta(days=1)
            ).isoformat(),
            "chunk_days": 0,
            "chunks_total": 1,
            "chunks_completed": 1,
            "refresh_rows": int(parquet_file.metadata.num_rows),
            "scanned_rows": int(parquet_file.metadata.num_rows),
            "pages": 0,
            "downloaded_mb": 0,
            "fetch_seconds": 0,
            "first_url": "legacy_snapshot",
            "paging_stop_reason": "legacy_partition_migration",
            "hit_page_limit": False,
            "latest_source_date": None,
        }
        published = publish_source_snapshot(
            source_key,
            None,
            API_FULL_START_DATE,
            work_dir,
            [legacy_file],
            fresh_columns,
            metadata,
            signature,
        )
        remove_refresh_work_dir(work_dir)
        return published
    except Exception:
        return None


def ensure_source_snapshot(
    source_key: str,
    username: str,
    auth_method: str,
) -> tuple[dict[str, Any], dict[str, Any]] | None:
    existing = source_snapshot_info(
        source_key,
        username,
        auth_method,
        API_FULL_START_DATE,
    )
    if existing is not None:
        return existing
    with source_refresh_lock(source_key) as lock_acquired:
        if not lock_acquired:
            return source_snapshot_info(
                source_key,
                username,
                auth_method,
                API_FULL_START_DATE,
            )
        existing = source_snapshot_info(
            source_key,
            username,
            auth_method,
            API_FULL_START_DATE,
        )
        if existing is not None:
            return existing
        return migrate_legacy_source_snapshot(
            source_key,
            username,
            auth_method,
        )


def refresh_source_snapshot(
    source_key: str,
    username: str,
    password: str,
    token: str,
    auth_method: str,
    *,
    full_refresh: bool,
    acquire_lock: bool = True,
) -> tuple[dict[str, Any], dict[str, Any]]:
    if source_key not in SOURCE_CONFIGS:
        raise ValueError(f"Unsupported AtlasFlow source: {source_key}")
    requested_signature = atlas_source_signature(
        source_key,
        username,
        auth_method,
        API_FULL_START_DATE,
    )

    def execute_refresh() -> tuple[dict[str, Any], dict[str, Any]]:
        existing_manifest = read_source_manifest(source_key)
        if not source_manifest_is_valid(
            source_key,
            existing_manifest,
            requested_signature,
            API_FULL_START_DATE,
        ):
            migrated = migrate_legacy_source_snapshot(
                source_key,
                username,
                auth_method,
            )
            existing_manifest = (
                read_source_manifest(source_key)
                if migrated is not None
                else None
            )
        if not source_manifest_is_valid(
            source_key,
            existing_manifest,
            requested_signature,
            API_FULL_START_DATE,
        ):
            existing_manifest = None

        refresh_mode = "full"
        refresh_start_date = API_FULL_START_DATE
        if not full_refresh and existing_manifest is not None:
            latest_text = (existing_manifest.get("metadata") or {}).get(
                "latest_source_date"
            )
            try:
                latest_date = (
                    date.fromisoformat(str(latest_text))
                    if latest_text
                    else None
                )
            except ValueError:
                latest_date = None
            if latest_date is not None:
                overlap_days = read_int_secret(
                    source_secret_name(
                        source_key,
                        "INCREMENTAL_OVERLAP_DAYS",
                    ),
                    DEFAULT_SOURCE_OVERLAP_DAYS[source_key],
                    minimum=1,
                    maximum=90,
                )
                refresh_start_date = max(
                    API_FULL_START_DATE,
                    latest_date - timedelta(days=overlap_days),
                )
                refresh_mode = "incremental"

        chunk_days = read_int_secret(
            source_secret_name(source_key, "REFRESH_CHUNK_DAYS"),
            DEFAULT_SOURCE_CHUNK_DAYS[source_key],
            minimum=1,
            maximum=62,
        )
        default_minutes = (
            DEFAULT_SOURCE_FULL_REFRESH_MAX_MINUTES[source_key]
            if refresh_mode == "full"
            else DEFAULT_SOURCE_INCREMENTAL_REFRESH_MAX_MINUTES[source_key]
        )
        max_minutes = read_int_secret(
            source_secret_name(
                source_key,
                (
                    "FULL_REFRESH_MAX_MINUTES"
                    if refresh_mode == "full"
                    else "INCREMENTAL_REFRESH_MAX_MINUTES"
                ),
            ),
            default_minutes,
            minimum=5,
            maximum=720,
        )
        end_exclusive = date.today() + timedelta(days=1)
        work_dir: Path | None = None
        try:
            update_source_refresh_status(
                source_key,
                state="running",
                stage="starting",
                refresh_mode=refresh_mode,
                refresh_start_date=refresh_start_date.isoformat(),
                end_date_exclusive=end_exclusive.isoformat(),
                chunk_days=chunk_days,
                max_minutes=max_minutes,
            )
            (
                work_dir,
                chunk_files,
                fresh_columns,
                refresh_metadata,
            ) = collect_source_chunks(
                source_key,
                username,
                password,
                token,
                auth_method,
                refresh_start_date,
                end_exclusive,
                chunk_days,
                max_minutes * 60,
                refresh_mode,
            )
            if int(refresh_metadata.get("scanned_rows", 0) or 0) == 0:
                if existing_manifest is not None:
                    metadata = dict(existing_manifest.get("metadata") or {})
                    metadata.update(
                        {
                            "refresh_mode": "no_changes",
                            "refresh_api_start_date": refresh_start_date.isoformat(),
                            "refresh_checked_at_local": local_time_label(),
                        }
                    )
                    update_source_refresh_status(
                        source_key,
                        state="complete",
                        stage="no_changes",
                        refresh_mode="no_changes",
                    )
                    return metadata, existing_manifest
                raise RuntimeError(
                    f"{SOURCE_CONFIGS[source_key]['label']} returned zero rows "
                    "during initial bootstrap."
                )
            update_source_refresh_status(
                source_key,
                state="running",
                stage="partitioning",
                refresh_mode=refresh_mode,
                pages_completed=int(
                    refresh_metadata.get("pages", 0) or 0
                ),
                rows_kept=int(
                    refresh_metadata.get("refresh_rows", 0) or 0
                ),
            )
            return publish_source_snapshot(
                source_key,
                (
                    existing_manifest
                    if refresh_mode == "incremental"
                    else None
                ),
                refresh_start_date,
                work_dir,
                chunk_files,
                fresh_columns,
                refresh_metadata,
                requested_signature,
            )
        except Exception as exc:
            update_source_refresh_status(
                source_key,
                state="failed",
                stage="failed",
                refresh_mode=refresh_mode,
                error=str(exc),
            )
            raise
        finally:
            remove_refresh_work_dir(work_dir)

    if not acquire_lock:
        return execute_refresh()
    with source_refresh_lock(source_key) as lock_acquired:
        if not lock_acquired:
            existing = source_snapshot_info(
                source_key,
                username,
                auth_method,
                API_FULL_START_DATE,
            )
            if existing is not None:
                metadata, manifest = existing
                metadata = dict(metadata)
                metadata["refresh_skipped_due_to_lock"] = True
                metadata["refresh_status"] = source_refresh_status_summary(
                    source_key
                )
                return metadata, manifest
            raise AtlasRefreshAlreadyRunning(
                source_refresh_status_summary(source_key)
            )
        return execute_refresh()


def _partition_entries_for_period(
    manifest: dict[str, Any],
    selected_start: date,
    selected_end: date,
) -> list[dict[str, Any]]:
    wanted_months = set(
        month_keys_for_range(
            selected_start,
            selected_end + timedelta(days=1),
        )
    )
    partitions = manifest_partitions(manifest)
    return [
        partitions[month_key]
        for month_key in sorted(partitions)
        if month_key in wanted_months
    ]


def _resolve_partition_vessel_names(
    dataset: ds.Dataset,
    entries: list[dict[str, Any]],
    selected_vessels: list[str] | None,
) -> list[str] | None:
    """Map UI vessel labels to the exact ShipName strings stored in Parquet.

    The UI fleet list is upper-case, while ShipPivots/ReportPivots may store the
    same vessel with mixed case or extra spaces. Arrow predicate pushdown is
    case-sensitive, so filtering with the UI label directly can return zero
    rows even though the vessel is present. Monthly summary files are tiny and
    provide the exact stored labels; a one-column dataset scan is used only as
    a fallback.
    """
    if not selected_vessels or "ShipName" not in dataset.schema.names:
        return selected_vessels

    selected_keys = {normalize_text(vessel) for vessel in selected_vessels}
    resolved: dict[str, str] = {}

    # First use the small monthly summaries, avoiding a scan of the large raw
    # ShipPivots partitions on normal app reruns.
    for entry in entries:
        summary_path = partition_entry_path(entry, "summary_file")
        if not summary_path.is_file():
            continue
        try:
            summary_table = pq.read_table(summary_path, columns=["ShipName"])
            for value in summary_table.column("ShipName").to_pylist():
                if value is None:
                    continue
                actual_name = str(value)
                key = normalize_text(actual_name)
                if key in selected_keys and key not in resolved:
                    resolved[key] = actual_name
            del summary_table
        except Exception:
            continue
        if selected_keys.issubset(resolved):
            break

    # Fallback for old/missing summary files: scan only the ShipName column and
    # stop as soon as all requested vessels have been resolved.
    if not selected_keys.issubset(resolved):
        try:
            scanner = dataset.scanner(columns=["ShipName"], batch_size=100_000)
            for batch in scanner.to_batches():
                for value in batch.column(0).to_pylist():
                    if value is None:
                        continue
                    actual_name = str(value)
                    key = normalize_text(actual_name)
                    if key in selected_keys and key not in resolved:
                        resolved[key] = actual_name
                if selected_keys.issubset(resolved):
                    break
        except Exception:
            pass

    # Preserve the original labels for unresolved vessels so a genuinely absent
    # vessel still produces a clean zero-row result rather than loading all data.
    return [
        resolved.get(normalize_text(vessel), vessel)
        for vessel in selected_vessels
    ]


def _dataset_filter_expression(
    source_key: str,
    schema_names: list[str],
    selected_vessels: list[str] | None,
    selected_start: date | None,
    selected_end: date | None,
) -> Any:
    expression = None
    if selected_vessels and "ShipName" in schema_names:
        expression = ds.field("ShipName").isin(list(selected_vessels))
    datetime_column = source_primary_datetime_column(source_key)
    if datetime_column in schema_names:
        if selected_start is not None:
            start_value = datetime.combine(
                selected_start,
                datetime.min.time(),
                tzinfo=timezone.utc,
            )
            start_expression = ds.field(datetime_column) >= pa.scalar(
                start_value,
                type=pa.timestamp("ns", tz="UTC"),
            )
            expression = (
                start_expression
                if expression is None
                else expression & start_expression
            )
        if selected_end is not None:
            end_value = datetime.combine(
                selected_end + timedelta(days=1),
                datetime.min.time(),
                tzinfo=timezone.utc,
            )
            end_expression = ds.field(datetime_column) < pa.scalar(
                end_value,
                type=pa.timestamp("ns", tz="UTC"),
            )
            expression = (
                end_expression
                if expression is None
                else expression & end_expression
            )
    return expression


def read_partitioned_source_slice(
    source_key: str,
    manifest: dict[str, Any],
    selected_vessels: list[str] | None,
    selected_start: date,
    selected_end: date,
    *,
    row_limit: int | None,
) -> tuple[pd.DataFrame, int, bool]:
    entries = _partition_entries_for_period(
        manifest,
        selected_start,
        selected_end,
    )
    files = [
        str(partition_entry_path(entry, "file"))
        for entry in entries
        if partition_entry_path(entry, "file").is_file()
    ]
    if not files:
        return pd.DataFrame(), 0, False
    dataset = ds.dataset(files, format="parquet")
    schema_names = list(dataset.schema.names)
    resolved_vessels = _resolve_partition_vessel_names(
        dataset,
        entries,
        selected_vessels,
    )
    expression = _dataset_filter_expression(
        source_key,
        schema_names,
        resolved_vessels,
        selected_start,
        selected_end,
    )
    try:
        matching_rows = int(dataset.count_rows(filter=expression))
    except Exception:
        matching_rows = sum(
            int(entry.get("rows", 0) or 0)
            for entry in entries
        )
    truncated = row_limit is not None and matching_rows > row_limit
    try:
        table = (
            dataset.head(row_limit, filter=expression)
            if truncated and row_limit is not None
            else dataset.to_table(filter=expression)
        )
        frame = table.to_pandas()
        del table
    except Exception:
        frames: list[pd.DataFrame] = []
        remaining = row_limit
        for file_name in files:
            parquet_file = pq.ParquetFile(file_name)
            for batch in parquet_file.iter_batches(batch_size=50_000):
                batch_frame = batch.to_pandas()
                batch_frame = filter_wide_source_data(
                    batch_frame,
                    source_key,
                    selected_vessels or [],
                    selected_start,
                    selected_end,
                )
                if batch_frame.empty:
                    continue
                if remaining is not None:
                    batch_frame = batch_frame.head(remaining)
                    remaining -= len(batch_frame)
                frames.append(batch_frame)
                if remaining is not None and remaining <= 0:
                    break
            if remaining is not None and remaining <= 0:
                break
        frame = (
            pd.concat(frames, ignore_index=True, sort=False)
            if frames
            else pd.DataFrame(columns=schema_names)
        )
    return frame, matching_rows, truncated


def load_wide_source_for_view(
    source_key: str,
    username: str,
    password: str,
    token: str,
    auth_method: str,
    start_date: date,
    refresh: bool,
    selected_vessels: list[str] | None = None,
    selected_start: date | None = None,
    selected_end: date | None = None,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    if refresh:
        refresh_source_snapshot(
            source_key,
            username,
            password,
            token,
            auth_method,
            full_refresh=False,
        )
    requested_signature = atlas_source_signature(
        source_key,
        username,
        auth_method,
        start_date,
    )
    manifest = read_source_manifest(source_key)
    if not source_manifest_is_valid(
        source_key,
        manifest,
        requested_signature,
        start_date,
    ):
        migrated = ensure_source_snapshot(
            source_key,
            username,
            auth_method,
        )
        manifest = (
            read_source_manifest(source_key)
            if migrated is not None
            else None
        )
    if not source_manifest_is_valid(
        source_key,
        manifest,
        requested_signature,
        start_date,
    ):
        config = SOURCE_CONFIGS[source_key]
        return pd.DataFrame(), {
            "source": config["label"],
            "endpoint": str(config["endpoint"]),
            "loaded_at_local": "No prepared snapshot yet",
            "rows": 0,
            "needs_warmup": True,
        }
    assert manifest is not None
    selected_start = selected_start or start_date
    selected_end = selected_end or date.today()
    row_limit = (
        ATLAS_RAW_INTERACTIVE_ROW_LIMIT
        if source_key == "shippivots"
        else ATLAS_REPORTPIVOTS_INTERACTIVE_ROW_LIMIT
    )
    view_signature = (
        manifest.get("generation"),
        tuple(sorted((str(vessel) for vessel in (selected_vessels or [])), key=str.casefold)),
        selected_start.isoformat(),
        selected_end.isoformat(),
        row_limit,
    )
    cached_frame = st.session_state.get(f"loaded_{source_key}_df")
    cached_metadata = st.session_state.get(f"loaded_{source_key}_metadata")
    if (
        not refresh
        and st.session_state.get(f"loaded_{source_key}_signature") == view_signature
        and isinstance(cached_frame, pd.DataFrame)
        and isinstance(cached_metadata, dict)
    ):
        return cached_frame, dict(cached_metadata)

    frame, matching_rows, truncated = read_partitioned_source_slice(
        source_key,
        manifest,
        selected_vessels,
        selected_start,
        selected_end,
        row_limit=row_limit,
    )
    metadata = dict(manifest.get("metadata") or {})
    metadata["loaded_from_snapshot"] = True
    metadata["snapshot_generation"] = manifest.get("generation")
    metadata["snapshot_saved_at_utc"] = manifest.get(
        "saved_at_utc",
        "-",
    )
    metadata["view_rows_matching"] = int(matching_rows)
    metadata["view_rows_loaded"] = int(len(frame))
    metadata["view_truncated"] = bool(truncated)
    metadata["interactive_row_limit"] = int(row_limit)
    st.session_state[f"loaded_{source_key}_df"] = frame
    st.session_state[f"loaded_{source_key}_metadata"] = metadata
    st.session_state[f"loaded_{source_key}_signature"] = view_signature
    return frame, metadata


def render_wide_source_tab(
    source_label: str,
    df: pd.DataFrame,
    metadata: dict[str, Any],
    source_key: str,
    selected_vessels: list[str],
    selected_start: date,
    selected_end: date,
) -> pd.DataFrame:
    """Render ReportPivots/ShipPivots with the same three-view structure as Custom Analytics.

    This is a presentation-layer change only. The API warmup, monthly partitions,
    manifests, incremental refresh, and memory safeguards remain unchanged.
    """
    st.markdown(
        f'<div class="section-title">{escape(source_label)} Preview & Export</div>',
        unsafe_allow_html=True,
    )
    st.caption(
        "Choose which table you want to preview and export. Clean Dataset uses the "
        "selected source columns, Summary Analysis aggregates that cleaned table, and "
        "Source Data shows all available API columns for the loaded vessel/date slice."
    )
    render_api_load_caption(metadata)

    # The partition reader already applies vessel/date predicate pushdown. Keep the
    # defensive in-memory filter so migrated/legacy snapshots behave identically.
    source_df = filter_wide_source_data(
        df,
        source_key,
        selected_vessels,
        selected_start,
        selected_end,
    )
    matching_rows = int(
        metadata.get("view_rows_matching", len(source_df)) or 0
    )
    interactive_limit = int(
        metadata.get("interactive_row_limit", len(source_df)) or len(source_df)
    )

    if metadata.get("view_truncated"):
        st.warning(
            f"The selected source contains {matching_rows:,} rows. AtlasFlow loaded "
            f"the first {interactive_limit:,} rows to protect Streamlit Cloud memory. "
            "Clean Dataset, Source Data, and Summary Analysis below use the loaded "
            "rows. Narrow the vessel/date selection for a complete raw-detail summary, "
            "or use Monthly Comparison for complete all-vessel monthly analysis."
        )

    default_columns = [
        column
        for column in [
            "ShipName",
            "DateTime",
            "State",
            "StateName",
            "GPSSpeed",
            "LogSpeed",
            "MEConsumed",
            "ShaftPower",
        ]
        if column in source_df.columns
    ]
    if not default_columns:
        default_columns = list(
            source_df.columns[: min(12, len(source_df.columns))]
        )

    state_columns_key = f"{source_key}_preview_columns"
    previous_columns = st.session_state.get(state_columns_key, default_columns)
    if not isinstance(previous_columns, list):
        previous_columns = default_columns
    valid_previous_columns = [
        column for column in previous_columns if column in source_df.columns
    ]
    if not valid_previous_columns:
        valid_previous_columns = default_columns
    if valid_previous_columns != previous_columns:
        st.session_state[state_columns_key] = valid_previous_columns

    selected_columns = st.multiselect(
        f"{source_label} columns to include",
        options=list(source_df.columns),
        default=valid_previous_columns,
        key=state_columns_key,
        help=(
            "These columns define Clean Dataset and the available fields in Summary "
            "Analysis. Source Data always retains all available API columns."
        ),
    )
    if not selected_columns:
        selected_columns = default_columns

    clean_df = (
        source_df[selected_columns].copy()
        if selected_columns
        else source_df.copy()
    )

    preview_options = ["Clean Dataset", "Summary Analysis", "Source Data"]
    preview_state_key = f"atlas_{source_key}_preview_mode"
    preview_mode = get_tab_selection(
        f"{source_key}_preview",
        preview_options,
        st.session_state.get(preview_state_key, "Clean Dataset"),
    )
    preview_mode = render_text_tab_bar(
        preview_options,
        preview_mode,
        param_name=f"{source_key}_preview",
        css_class="compact",
    )
    st.session_state[preview_state_key] = preview_mode

    summary_group_fields: list[str] = []
    summary_value_fields: list[str] = []
    aggregation_options = [
        "Average",
        "Sum",
        "Count",
        "Minimum",
        "Maximum",
        "Median",
    ]
    aggregation_key = f"atlas_{source_key}_summary_aggregation"
    summary_aggregation = st.session_state.get(aggregation_key, "Average")
    if summary_aggregation not in aggregation_options:
        summary_aggregation = "Average"

    group_key = f"atlas_{source_key}_summary_groups"
    legacy_value_key = f"atlas_{source_key}_summary_values"

    if preview_mode == "Summary Analysis":
        st.markdown(
            '<div class="section-title">Summary Builder</div>',
            unsafe_allow_html=True,
        )

        previous_groups = st.session_state.get(group_key, [])
        if not isinstance(previous_groups, list):
            previous_groups = []
        valid_groups = [
            column for column in previous_groups if column in clean_df.columns
        ]
        if not valid_groups and group_key not in st.session_state:
            valid_groups = [
                column
                for column in ["ShipName", "ReportType", "State", "StateName"]
                if column in clean_df.columns
            ][:2]
        if valid_groups != previous_groups:
            st.session_state[group_key] = valid_groups

        builder_cols = st.columns(2)
        with builder_cols[0]:
            summary_group_fields = st.multiselect(
                "Group by fields",
                options=list(clean_df.columns),
                default=valid_groups,
                key=group_key,
                help="Choose the fields that define each summary row.",
            )
        with builder_cols[1]:
            summary_aggregation = st.selectbox(
                "Aggregation",
                options=aggregation_options,
                index=aggregation_options.index(summary_aggregation),
                key=aggregation_key,
            )
    else:
        stored_groups = st.session_state.get(group_key, [])
        summary_group_fields = stored_groups if isinstance(stored_groups, list) else []

    # Summary metrics now follow the Clean Dataset selection automatically. The
    # user selects source columns once above, then chooses only grouping fields
    # and the aggregation method here. Exclude grouping and obvious identifier
    # columns so IDs are not accidentally averaged or summed.
    non_metric_keys = {
        "reportid",
        "shipid",
        "vesselid",
        "voyageid",
        "imo",
        "imonumber",
        "mmsi",
    }
    summary_value_fields = [
        column
        for column in numeric_column_options(clean_df)
        if column not in summary_group_fields
        and normalize_text(column) not in non_metric_keys
    ]

    # Remove the obsolete widget state left by earlier AtlasFlow versions. It is
    # no longer used because value fields are inherited from Clean Dataset.
    st.session_state.pop(legacy_value_key, None)

    if preview_mode == "Summary Analysis":
        if summary_value_fields:
            metric_preview = ", ".join(summary_value_fields[:8])
            if len(summary_value_fields) > 8:
                metric_preview += f", +{len(summary_value_fields) - 8} more"
            st.caption(
                f"{len(summary_value_fields):,} numeric field(s) inherited automatically "
                f"from Clean Dataset: {metric_preview}. Change the columns above to "
                "change the summary metrics."
            )
        else:
            st.info(
                "No numeric metric columns are available after excluding the selected "
                "Group by fields. Add a numeric column to Clean Dataset or change the grouping."
            )

    export_sheet_name = "Clean Dataset"
    if preview_mode == "Summary Analysis":
        if summary_group_fields and summary_value_fields:
            displayed_table_df = build_summary_analysis(
                clean_df,
                group_fields=summary_group_fields,
                value_fields=summary_value_fields,
                aggregation=summary_aggregation,
            )
        else:
            displayed_table_df = pd.DataFrame()
            st.info(
                "Select at least one Group by field and include at least one numeric metric in Clean Dataset to preview Summary Analysis."
            )
        export_sheet_name = "Summary Analysis"
    elif preview_mode == "Source Data":
        displayed_table_df = source_df.copy()
        export_sheet_name = "Source Data"
    else:
        displayed_table_df = clean_df.copy()

    render_metric_cards(
        [
            ("Displayed Rows", f"{len(displayed_table_df):,}", "table_eye"),
            ("Selected Columns", f"{len(selected_columns):,}", "checked_columns"),
            ("Rows in Selection", f"{matching_rows:,}", "database_rows"),
            ("Available Columns", f"{len(source_df.columns):,}", "columns_plus"),
        ]
    )

    render_preview_table(displayed_table_df)
    if len(displayed_table_df) > TABLE_PREVIEW_ROW_LIMIT:
        st.caption(
            f"Showing first {TABLE_PREVIEW_ROW_LIMIT:,} of "
            f"{len(displayed_table_df):,} loaded rows. Excel export includes the "
            "full displayed table currently held by the app."
        )

    export_signature_payload = "|".join(
        [
            source_key,
            preview_mode,
            ",".join(selected_vessels),
            selected_start.isoformat(),
            selected_end.isoformat(),
            ",".join(selected_columns),
            str(len(source_df)),
            str(matching_rows),
            str(len(displayed_table_df)),
            ",".join(summary_group_fields),
            ",".join(summary_value_fields),
            summary_aggregation,
            ",".join(displayed_table_df.columns.astype(str).tolist())
            if not displayed_table_df.empty
            else "empty",
        ]
    )
    export_signature = sha256(
        export_signature_payload.encode("utf-8")
    ).hexdigest()
    signature_key = f"atlas_{source_key}_display_export_signature"
    bytes_key = f"atlas_{source_key}_display_export_bytes"
    if st.session_state.get(signature_key) != export_signature:
        st.session_state.pop(bytes_key, None)

    export_ready = (
        st.session_state.get(signature_key) == export_signature
        and bytes_key in st.session_state
    )
    if st.button(
        f"Prepare {source_label} displayed table Excel",
        type="primary",
        disabled=displayed_table_df.empty,
        key=f"atlas_{source_key}_prepare_displayed_excel",
    ):
        with st.spinner("Preparing Excel file..."):
            st.session_state[bytes_key] = to_displayed_table_excel_bytes(
                displayed_table_df,
                sheet_name=export_sheet_name,
            )
            st.session_state[signature_key] = export_signature
            gc.collect()
        export_ready = True

    if export_ready:
        st.download_button(
            f"Download {source_label} displayed table Excel",
            data=st.session_state[bytes_key],
            file_name=f"atlasflow_{source_key}_{slugify_tab_label(preview_mode)}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"atlas_{source_key}_download_displayed_excel",
        )
    else:
        st.caption(
            "Excel generation is prepared on demand. The download contains only "
            "the table shown by the selected sub-tab."
        )

    return displayed_table_df

def load_monthly_comparison_data(
    username: str,
    auth_method: str,
    selected_vessels: list[str],
    selected_start: date,
    selected_end: date,
) -> tuple[
    pd.DataFrame,
    dict[str, dict[str, Any]],
    list[str],
]:
    summary_files: list[str] = []
    generation_parts: list[str] = []
    source_metadata: dict[str, dict[str, Any]] = {}
    missing_sources: list[str] = []
    wanted_months = set(
        month_keys_for_range(
            selected_start,
            selected_end + timedelta(days=1),
        )
    )
    for source_key in ["reportdata", "reportpivots", "shippivots"]:
        requested_signature = atlas_source_signature(
            source_key,
            username,
            auth_method,
            API_FULL_START_DATE,
        )
        manifest = read_source_manifest(source_key)
        if not source_manifest_is_valid(
            source_key,
            manifest,
            requested_signature,
            API_FULL_START_DATE,
        ):
            missing_sources.append(COMPARISON_SOURCE_LABELS[source_key])
            continue
        assert manifest is not None
        source_metadata[source_key] = dict(
            manifest.get("metadata") or {}
        )
        generation_parts.append(
            f"{source_key}:{manifest.get('generation')}"
        )
        for month_key, entry in manifest_partitions(manifest).items():
            if month_key not in wanted_months:
                continue
            summary_path = partition_entry_path(entry, "summary_file")
            if summary_path.is_file():
                summary_files.append(str(summary_path))
    comparison = cached_read_monthly_summary_files(
        "|".join(generation_parts),
        tuple(sorted(summary_files)),
    )
    if (
        not comparison.empty
        and selected_vessels
        and "ShipName" in comparison.columns
    ):
        comparison = comparison[
            match_selected_vessels(
                comparison["ShipName"],
                selected_vessels,
            )
        ].copy()
    return comparison, source_metadata, missing_sources


def to_simplified_monthly_comparison_excel_bytes(
    availability_df: pd.DataFrame,
    detail_df: pd.DataFrame,
    mapping_df: pd.DataFrame,
) -> bytes:
    """Export the simplified monthly comparison as one clear workbook."""
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        write_table_sheet(
            writer,
            availability_df,
            "Data Availability",
            "AtlasMonthlyAvailability",
        )
        write_table_sheet(
            writer,
            detail_df,
            "Monthly Values",
            "AtlasMonthlyValues",
        )
        if mapping_df is not None and not mapping_df.empty:
            write_table_sheet(
                writer,
                mapping_df,
                "Field Mapping",
                "AtlasMonthlyMapping",
            )
    return output.getvalue()




_DYNAMIC_COMPARISON_EXCLUDED_FIELDS = {
    "id",
    "reportid",
    "shipname",
    "datetime",
    "startdatetimegmt",
    "enddatetimegmt",
    "reportdatetime",
    "timestamp",
    "imo",
    "imono",
    "mmsi",
    "companyname",
    "latitude",
    "longitude",
    "voyageid",
    "voyageidinternal",
}


def _candidate_name_from_list(
    available_names: list[str],
    candidates: list[str],
) -> str | None:
    normalized = {normalize_text(name): str(name) for name in available_names}
    for candidate in candidates:
        match = normalized.get(normalize_text(candidate))
        if match is not None:
            return match
    for candidate in candidates:
        candidate_key = normalize_text(candidate)
        if len(candidate_key) < 6:
            continue
        for available_key, original_name in normalized.items():
            if candidate_key in available_key or available_key in candidate_key:
                return original_name
    return None


def _dynamic_aggregation(field_name: str) -> str:
    """Choose a conservative monthly aggregation for a raw source field."""
    key = normalize_text(field_name)
    latest_tokens = (
        "rob",
        "remaining",
        "onboard",
        "counter",
        "totalizer",
        "cumulative",
        "opening",
        "closing",
    )
    sum_tokens = (
        "consum",
        "fuel",
        "distance",
        "runninghours",
        "operatinghours",
        "steamingtime",
        "energy",
        "emission",
        "quantity",
        "received",
        "bunkered",
        "production",
    )
    if any(token in key for token in latest_tokens):
        return "latest"
    if any(token in key for token in sum_tokens):
        return "sum"
    return "mean"


def _aggregation_label(aggregation: str) -> str:
    return {
        "sum": "Sum",
        "mean": "Average",
        "latest": "Latest value",
        "min": "Minimum",
        "max": "Maximum",
    }.get(str(aggregation), str(aggregation).title())


def _source_files_for_comparison(
    manifest: dict[str, Any],
    selected_start: date,
    selected_end: date,
) -> tuple[list[dict[str, Any]], tuple[str, ...]]:
    entries = _partition_entries_for_period(
        manifest,
        selected_start,
        selected_end,
    )
    files = tuple(
        str(partition_entry_path(entry, "file"))
        for entry in entries
        if partition_entry_path(entry, "file").is_file()
    )
    return entries, files


@st.cache_data(show_spinner=False)
def cached_reportdata_field_names(
    generation_signature: str,
    partition_files: tuple[str, ...],
    resolved_vessels: tuple[str, ...],
    selected_start: date,
    selected_end: date,
) -> list[str]:
    del generation_signature
    if not partition_files:
        return []
    dataset = ds.dataset(list(partition_files), format="parquet")
    schema_names = list(dataset.schema.names)
    required = [
        column
        for column in ["ShipName", "StartDateTimeGMT", "ValueDescription"]
        if column in schema_names
    ]
    if "ValueDescription" not in required:
        return []
    expression = _dataset_filter_expression(
        "reportdata",
        schema_names,
        list(resolved_vessels),
        selected_start,
        selected_end,
    )
    values: set[str] = set()
    scanner = dataset.scanner(
        columns=required,
        filter=expression,
        batch_size=100_000,
    )
    for batch in scanner.to_batches():
        frame = batch.to_pandas()
        if "ValueDescription" not in frame.columns:
            continue
        descriptions = (
            frame["ValueDescription"]
            .dropna()
            .astype(str)
            .str.strip()
        )
        values.update(descriptions[descriptions.ne("")].tolist())
    return sorted(values, key=str.casefold)


def build_dynamic_comparison_catalog(
    source_fields: dict[str, list[str]],
) -> list[dict[str, Any]]:
    """Build standardized plus raw fields from the three current API schemas."""
    catalog: list[dict[str, Any]] = []
    used_labels: set[str] = set()

    # Keep the approved standardized fields first because these have explicit,
    # trusted source aliases and aggregation rules.
    for label, spec in MONTHLY_COMPARISON_METRICS.items():
        mappings: dict[str, list[str]] = {}
        for source_key in ["reportdata", "reportpivots", "shippivots"]:
            available = source_fields.get(source_key, [])
            matched = _candidate_name_from_list(
                available,
                list(spec.get("candidates") or []),
            )
            if matched is not None:
                mappings[source_key] = [matched]
                continue
            if source_key == "reportdata":
                components = [
                    value
                    for value in list(spec.get("component_candidates") or [])
                    if normalize_text(value)
                    in {normalize_text(name) for name in available}
                ]
                if components:
                    mappings[source_key] = components
        if mappings:
            catalog.append(
                {
                    "label": label,
                    "aggregation": str(spec.get("aggregation", "mean")),
                    "mappings": mappings,
                    "standardized": True,
                }
            )
            used_labels.add(label.casefold())

    # Group raw fields by normalized name. Exact normalized matches are treated
    # as the same field across APIs; source-only fields remain selectable too.
    grouped: dict[str, dict[str, str]] = {}
    for source_key in ["reportdata", "reportpivots", "shippivots"]:
        for raw_name in source_fields.get(source_key, []):
            if source_key != "reportdata" and normalize_text(raw_name) in _DYNAMIC_COMPARISON_EXCLUDED_FIELDS:
                continue
            key = normalize_text(raw_name)
            if not key:
                continue
            grouped.setdefault(key, {})[source_key] = str(raw_name)

    source_preference = ["reportdata", "reportpivots", "shippivots"]
    raw_entries: list[dict[str, Any]] = []
    for mappings_by_source in grouped.values():
        display_name = next(
            mappings_by_source[source]
            for source in source_preference
            if source in mappings_by_source
        )
        label = display_name
        if label.casefold() in used_labels:
            label = f"{label} (source field)"
        suffix = 2
        base_label = label
        while label.casefold() in used_labels:
            label = f"{base_label} {suffix}"
            suffix += 1
        used_labels.add(label.casefold())
        raw_entries.append(
            {
                "label": label,
                "aggregation": _dynamic_aggregation(display_name),
                "mappings": {
                    source: [raw_name]
                    for source, raw_name in mappings_by_source.items()
                },
                "standardized": False,
            }
        )

    raw_entries.sort(key=lambda entry: str(entry["label"]).casefold())
    catalog.extend(raw_entries)
    return catalog


def _update_dynamic_group_state(
    states: dict[tuple[str, str, str], dict[str, Any]],
    frame: pd.DataFrame,
    field_label: str,
    raw_values: pd.Series,
    numeric_values: pd.Series,
    datetime_column: str,
) -> None:
    if frame.empty:
        return
    work = pd.DataFrame(
        {
            "Month": pd.to_datetime(
                frame[datetime_column], errors="coerce", utc=True
            ).dt.to_period("M").astype("string"),
            "ShipName": frame["ShipName"].astype("string"),
            "Timestamp": pd.to_datetime(
                frame[datetime_column], errors="coerce", utc=True
            ),
            "RawValue": raw_values,
            "NumericValue": numeric_values,
        },
        index=frame.index,
    )
    work = work[
        work["Month"].notna()
        & work["ShipName"].notna()
        & work["Timestamp"].notna()
    ].copy()
    if work.empty:
        return
    raw_text = work["RawValue"].astype("string").str.strip()
    work["Present"] = work["RawValue"].notna() & raw_text.ne("")

    grouped = work.groupby(["Month", "ShipName"], dropna=False)
    present_counts = grouped["Present"].sum()
    numeric_counts = grouped["NumericValue"].count()
    numeric_sums = grouped["NumericValue"].sum(min_count=1)
    numeric_mins = grouped["NumericValue"].min()
    numeric_maxs = grouped["NumericValue"].max()

    latest_rows = (
        work[work["NumericValue"].notna()]
        .sort_values("Timestamp")
        .groupby(["Month", "ShipName"], dropna=False)
        .tail(1)
        .set_index(["Month", "ShipName"])
    )

    for pair, present_count in present_counts.items():
        month_value, ship_value = str(pair[0]), str(pair[1])
        state = states.setdefault(
            (month_value, ship_value, field_label),
            {
                "PresentValues": 0,
                "NumericValues": 0,
                "Sum": 0.0,
                "Minimum": None,
                "Maximum": None,
                "LatestTimestamp": None,
                "LatestValue": None,
            },
        )
        state["PresentValues"] += int(present_count or 0)
        count = int(numeric_counts.get(pair, 0) or 0)
        state["NumericValues"] += count
        if count:
            batch_sum = numeric_sums.get(pair)
            if pd.notna(batch_sum):
                state["Sum"] += float(batch_sum)
            batch_min = numeric_mins.get(pair)
            batch_max = numeric_maxs.get(pair)
            if pd.notna(batch_min):
                state["Minimum"] = (
                    float(batch_min)
                    if state["Minimum"] is None
                    else min(float(state["Minimum"]), float(batch_min))
                )
            if pd.notna(batch_max):
                state["Maximum"] = (
                    float(batch_max)
                    if state["Maximum"] is None
                    else max(float(state["Maximum"]), float(batch_max))
                )
        if pair in latest_rows.index:
            latest_row = latest_rows.loc[pair]
            if isinstance(latest_row, pd.DataFrame):
                latest_row = latest_row.iloc[-1]
            latest_timestamp = latest_row["Timestamp"]
            if (
                state["LatestTimestamp"] is None
                or latest_timestamp > state["LatestTimestamp"]
            ):
                state["LatestTimestamp"] = latest_timestamp
                state["LatestValue"] = float(latest_row["NumericValue"])


@st.cache_data(show_spinner=False)
def cached_dynamic_source_aggregates(
    source_key: str,
    generation_signature: str,
    partition_files: tuple[str, ...],
    resolved_vessels: tuple[str, ...],
    selected_start: date,
    selected_end: date,
    field_specs: tuple[tuple[str, str, tuple[str, ...]], ...],
) -> pd.DataFrame:
    """Aggregate only selected fields directly from monthly Parquet partitions."""
    del generation_signature
    if not partition_files or not field_specs:
        return pd.DataFrame()
    dataset = ds.dataset(list(partition_files), format="parquet")
    schema_names = list(dataset.schema.names)
    datetime_column = source_primary_datetime_column(source_key)
    if datetime_column not in schema_names or "ShipName" not in schema_names:
        return pd.DataFrame()

    base_expression = _dataset_filter_expression(
        source_key,
        schema_names,
        list(resolved_vessels),
        selected_start,
        selected_end,
    )
    states: dict[tuple[str, str, str], dict[str, Any]] = {}

    if source_key == "reportdata":
        required = [
            column
            for column in [
                "ShipName",
                datetime_column,
                "ReportId",
                "ValueDescription",
                "ParsedValue",
                "ReportedValue",
            ]
            if column in schema_names
        ]
        if "ValueDescription" not in required:
            return pd.DataFrame()
        raw_to_specs: dict[str, list[tuple[str, str]]] = {}
        for field_label, aggregation, mappings in field_specs:
            for raw_name in mappings:
                raw_to_specs.setdefault(str(raw_name), []).append(
                    (field_label, aggregation)
                )
        raw_names = list(raw_to_specs)
        expression = base_expression
        if raw_names:
            value_expression = ds.field("ValueDescription").isin(raw_names)
            expression = (
                value_expression
                if expression is None
                else expression & value_expression
            )
        seen_report_values: set[tuple[str, str, str]] = set()
        scanner = dataset.scanner(
            columns=required,
            filter=expression,
            batch_size=100_000,
        )
        for batch in scanner.to_batches():
            frame = batch.to_pandas()
            for raw_name, spec_pairs in raw_to_specs.items():
                subset = frame[
                    frame["ValueDescription"].astype("string").eq(raw_name)
                ].copy()
                if subset.empty:
                    continue
                if "ReportId" in subset.columns:
                    keep_mask = []
                    for report_id in subset["ReportId"].astype("string"):
                        if pd.isna(report_id) or str(report_id) in {"", "<NA>"}:
                            keep_mask.append(True)
                            continue
                        dedup_key = (raw_name, str(report_id), str(len(keep_mask)))
                        simple_key = (raw_name, str(report_id), "value")
                        if simple_key in seen_report_values:
                            keep_mask.append(False)
                        else:
                            seen_report_values.add(simple_key)
                            keep_mask.append(True)
                    subset = subset.loc[keep_mask].copy()
                raw_series = (
                    subset["ReportedValue"]
                    if "ReportedValue" in subset.columns
                    else subset.get("ParsedValue", pd.Series(index=subset.index, dtype="object"))
                )
                numeric_series = (
                    pd.to_numeric(subset["ParsedValue"], errors="coerce")
                    if "ParsedValue" in subset.columns
                    else parse_numeric_series(raw_series)
                )
                for field_label, _aggregation in spec_pairs:
                    _update_dynamic_group_state(
                        states,
                        subset,
                        field_label,
                        raw_series,
                        numeric_series,
                        datetime_column,
                    )
    else:
        source_columns = sorted(
            {
                raw_name
                for _, _, mappings in field_specs
                for raw_name in mappings
                if raw_name in schema_names
            }
        )
        columns = ["ShipName", datetime_column, *source_columns]
        scanner = dataset.scanner(
            columns=columns,
            filter=base_expression,
            batch_size=100_000,
        )
        for batch in scanner.to_batches():
            frame = batch.to_pandas()
            for field_label, _aggregation, mappings in field_specs:
                for raw_name in mappings:
                    if raw_name not in frame.columns:
                        continue
                    raw_series = frame[raw_name]
                    numeric_series = parse_numeric_series(raw_series)
                    _update_dynamic_group_state(
                        states,
                        frame,
                        field_label,
                        raw_series,
                        numeric_series,
                        datetime_column,
                    )

    aggregation_by_field = {
        field_label: aggregation
        for field_label, aggregation, _ in field_specs
    }
    rows: list[dict[str, Any]] = []
    for (month_value, ship_value, field_label), state in states.items():
        aggregation = aggregation_by_field.get(field_label, "mean")
        count = int(state["NumericValues"] or 0)
        value: float | None = None
        if count > 0:
            if aggregation == "sum":
                value = float(state["Sum"])
            elif aggregation == "latest":
                value = state["LatestValue"]
            elif aggregation == "min":
                value = state["Minimum"]
            elif aggregation == "max":
                value = state["Maximum"]
            else:
                value = float(state["Sum"]) / count
        rows.append(
            {
                "Month": month_value,
                "ShipName": ship_value,
                "Field": field_label,
                "Value": round(value, 3) if value is not None else pd.NA,
                "NumericValues": count,
                "PresentValues": int(state["PresentValues"] or 0),
            }
        )
    return pd.DataFrame(rows)


def to_dynamic_monthly_comparison_excel_bytes(
    availability_df: pd.DataFrame,
    detail_df: pd.DataFrame,
    mapping_df: pd.DataFrame,
) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        write_table_sheet(
            writer,
            availability_df,
            "Data Availability",
            "AtlasDynamicAvailability",
        )
        if detail_df is not None and not detail_df.empty:
            write_table_sheet(
                writer,
                detail_df,
                "Monthly Values",
                "AtlasDynamicMonthlyValues",
            )
        if mapping_df is not None and not mapping_df.empty:
            write_table_sheet(
                writer,
                mapping_df,
                "Field Mapping",
                "AtlasDynamicMapping",
            )
    return output.getvalue()


def render_monthly_comparison_workspace(
    username: str,
    auth_method: str,
    selected_vessels: list[str],
    selected_start: date,
    selected_end: date,
) -> None:
    """Compare every available API field, not only a fixed metric dictionary."""
    st.markdown(
        '<div class="section-title">Monthly Comparison</div>',
        unsafe_allow_html=True,
    )
    st.caption(
        "Select any field found in ReportData, ReportPivots, or ShipPivots. "
        "AtlasFlow shows the original source mapping, whether data exists in "
        "the selected period, and monthly numeric values side by side."
    )
    st.markdown(
        f'<div class="atlas-pill"><span>Period used:</span> '
        f'{selected_start.strftime("%d/%m/%Y")} to '
        f'{selected_end.strftime("%d/%m/%Y")}</div>',
        unsafe_allow_html=True,
    )

    source_order_keys = ["reportdata", "reportpivots", "shippivots"]
    source_order = [COMPARISON_SOURCE_LABELS[key] for key in source_order_keys]
    contexts: dict[str, dict[str, Any]] = {}
    source_fields: dict[str, list[str]] = {}
    source_metadata: dict[str, dict[str, Any]] = {}
    missing_sources: list[str] = []

    for source_key in source_order_keys:
        requested_signature = atlas_source_signature(
            source_key,
            username,
            auth_method,
            API_FULL_START_DATE,
        )
        manifest = read_source_manifest(source_key)
        if not source_manifest_is_valid(
            source_key,
            manifest,
            requested_signature,
            API_FULL_START_DATE,
        ):
            missing_sources.append(COMPARISON_SOURCE_LABELS[source_key])
            source_fields[source_key] = []
            continue
        assert manifest is not None
        entries, files = _source_files_for_comparison(
            manifest,
            selected_start,
            selected_end,
        )
        if not files:
            source_fields[source_key] = []
            continue
        dataset = ds.dataset(list(files), format="parquet")
        resolved_vessels = _resolve_partition_vessel_names(
            dataset,
            entries,
            selected_vessels,
        ) or selected_vessels
        contexts[source_key] = {
            "manifest": manifest,
            "entries": entries,
            "files": files,
            "resolved_vessels": tuple(resolved_vessels or []),
            "generation": str(manifest.get("generation", "")),
        }
        metadata = dict(manifest.get("metadata") or {})
        metadata["partition_count"] = len(manifest_partitions(manifest))
        source_metadata[source_key] = metadata
        if source_key == "reportdata":
            source_fields[source_key] = cached_reportdata_field_names(
                str(manifest.get("generation", "")),
                files,
                tuple(resolved_vessels or []),
                selected_start,
                selected_end,
            )
        else:
            source_fields[source_key] = [
                str(column) for column in dataset.schema.names
            ]

    if missing_sources:
        st.warning(
            "Prepared monthly partitions are not available yet for: "
            + ", ".join(missing_sources)
            + ". Those sources remain visible as No data."
        )

    catalog = build_dynamic_comparison_catalog(source_fields)
    if not catalog:
        st.info("No API fields were found for the selected period and sources.")
        return

    catalog_by_label = {str(entry["label"]): entry for entry in catalog}
    field_options = list(catalog_by_label)
    standardized_defaults = [
        str(entry["label"])
        for entry in catalog
        if entry.get("standardized")
    ][:8]
    if not standardized_defaults:
        standardized_defaults = field_options[: min(8, len(field_options))]

    state_key = "atlas_monthly_dynamic_fields"
    previous = st.session_state.get(state_key, standardized_defaults)
    if not isinstance(previous, list):
        previous = standardized_defaults
    valid_previous = [field for field in previous if field in catalog_by_label]
    if not valid_previous:
        valid_previous = standardized_defaults
    if valid_previous != previous:
        st.session_state[state_key] = valid_previous

    selected_fields = st.multiselect(
        "Fields to compare",
        options=field_options,
        default=valid_previous,
        key=state_key,
        help=(
            "This list is generated dynamically from the current ReportData "
            "ValueDescriptions and all columns stored by ReportPivots and ShipPivots."
        ),
    )
    if not selected_fields:
        st.info("Select at least one field to display the comparison.")
        return
    if len(selected_fields) > 25:
        st.warning(
            "For a clear and memory-safe comparison, select up to 25 fields at a time."
        )
        selected_fields = selected_fields[:25]

    selected_entries = [catalog_by_label[field] for field in selected_fields]
    month_values = month_keys_for_range(
        selected_start,
        selected_end + timedelta(days=1),
    )
    expected_pairs = max(len(month_values) * max(len(selected_vessels), 1), 1)

    render_metric_cards(
        [
            ("Available Fields", f"{len(field_options):,}", "columns_plus"),
            ("Selected Fields", f"{len(selected_fields):,}", "checked_columns"),
            ("Vessels", f"{len(selected_vessels):,}", "vessel"),
            ("API Sources", "3", "api_sources"),
        ]
    )

    source_results: dict[str, pd.DataFrame] = {}
    selected_ship_labels = {
        normalize_text(vessel): vessel for vessel in selected_vessels
    }
    with st.spinner("Checking selected fields across the three API sources..."):
        for source_key in source_order_keys:
            context = contexts.get(source_key)
            if context is None:
                source_results[source_key] = pd.DataFrame()
                continue
            specs: list[tuple[str, str, tuple[str, ...]]] = []
            for entry in selected_entries:
                mappings = tuple(
                    str(value)
                    for value in entry.get("mappings", {}).get(source_key, [])
                )
                if mappings:
                    specs.append(
                        (
                            str(entry["label"]),
                            str(entry["aggregation"]),
                            mappings,
                        )
                    )
            result = cached_dynamic_source_aggregates(
                source_key,
                str(context["generation"]),
                tuple(context["files"]),
                tuple(context["resolved_vessels"]),
                selected_start,
                selected_end,
                tuple(specs),
            )
            if not result.empty:
                result["ShipName"] = result["ShipName"].map(
                    lambda value: selected_ship_labels.get(
                        normalize_text(value), str(value)
                    )
                )
            source_results[source_key] = result

    availability_rows: list[dict[str, Any]] = []
    mapping_rows: list[dict[str, Any]] = []
    numeric_fields: list[str] = []

    for entry in selected_entries:
        field_label = str(entry["label"])
        aggregation = str(entry["aggregation"])
        availability_row: dict[str, Any] = {
            "Field": field_label,
            "Monthly rule": _aggregation_label(aggregation),
        }
        mapping_row: dict[str, Any] = {
            "Field": field_label,
            "Monthly rule": _aggregation_label(aggregation),
        }
        sources_present = 0
        field_has_numeric = False

        for source_key, source_label in zip(source_order_keys, source_order):
            mappings = [
                str(value)
                for value in entry.get("mappings", {}).get(source_key, [])
            ]
            mapping_text = " + ".join(mappings) if mappings else "-"
            mapping_row[source_label] = mapping_text
            result = source_results.get(source_key, pd.DataFrame())
            if result.empty or not mappings:
                availability_row[source_label] = (
                    f"{mapping_text} — No data"
                    if mappings
                    else "Not mapped"
                )
                continue
            field_rows = result[result["Field"].eq(field_label)].copy()
            numeric_pairs = int(
                field_rows.loc[
                    pd.to_numeric(field_rows["Value"], errors="coerce").notna(),
                    ["Month", "ShipName"],
                ].drop_duplicates().shape[0]
            )
            present_pairs = int(
                field_rows.loc[
                    pd.to_numeric(
                        field_rows["PresentValues"], errors="coerce"
                    ).fillna(0).gt(0),
                    ["Month", "ShipName"],
                ].drop_duplicates().shape[0]
            )
            if numeric_pairs > 0:
                sources_present += 1
                field_has_numeric = True
                status = f"Numeric {numeric_pairs:,}/{expected_pairs:,}"
            elif present_pairs > 0:
                sources_present += 1
                status = f"Present {present_pairs:,}/{expected_pairs:,} (non-numeric)"
            else:
                status = "No data"
            availability_row[source_label] = f"{mapping_text} — {status}"

        availability_row["Availability"] = (
            "All 3 sources"
            if sources_present == 3
            else f"{sources_present} of 3 sources"
            if sources_present
            else "No source data"
        )
        availability_rows.append(availability_row)
        mapping_rows.append(mapping_row)
        if field_has_numeric:
            numeric_fields.append(field_label)

    availability_df = pd.DataFrame(availability_rows)[
        ["Field", "Monthly rule", *source_order, "Availability"]
    ]
    mapping_df = pd.DataFrame(mapping_rows)[
        ["Field", "Monthly rule", *source_order]
    ]

    detail_rows: list[dict[str, Any]] = []
    value_lookup: dict[tuple[str, str, str, str], Any] = {}
    for source_key, source_label in zip(source_order_keys, source_order):
        result = source_results.get(source_key, pd.DataFrame())
        if result.empty:
            continue
        for row in result.itertuples(index=False):
            value_lookup[
                (str(row.Month), str(row.ShipName), str(row.Field), source_label)
            ] = row.Value

    for month_value in month_values:
        for vessel in selected_vessels:
            for field_label in numeric_fields:
                row = {
                    "Month": month_value,
                    "ShipName": vessel,
                    "Field": field_label,
                    "Monthly rule": _aggregation_label(
                        str(catalog_by_label[field_label]["aggregation"])
                    ),
                }
                available_count = 0
                for source_label in source_order:
                    value = value_lookup.get(
                        (month_value, vessel, field_label, source_label),
                        pd.NA,
                    )
                    row[source_label] = value
                    if pd.notna(value):
                        available_count += 1
                row["Data availability"] = {
                    3: "All 3 sources",
                    2: "2 of 3 sources",
                    1: "1 of 3 sources",
                    0: "No source data",
                }[available_count]
                detail_rows.append(row)

    detail_df = pd.DataFrame(detail_rows)
    if not detail_df.empty:
        detail_df = detail_df[
            [
                "Month",
                "ShipName",
                "Field",
                "Monthly rule",
                *source_order,
                "Data availability",
            ]
        ].sort_values(
            ["Month", "ShipName", "Field"],
            ascending=[False, True, True],
        )

    st.markdown(
        '<div class="section-title">Field availability and source mapping</div>',
        unsafe_allow_html=True,
    )
    st.caption(
        "Each API column shows the original source field followed by its data "
        "availability. Numeric x/y counts vessel-month combinations with a "
        "monthly value; Present identifies non-numeric fields."
    )
    st.dataframe(availability_df, use_container_width=True, hide_index=True)

    st.markdown(
        '<div class="section-title">Monthly numeric values side by side</div>',
        unsafe_allow_html=True,
    )
    if detail_df.empty:
        st.info(
            "The selected fields exist only as text/status data, so availability "
            "is shown above but no numeric monthly comparison is produced."
        )
    else:
        st.caption(
            "Only fields containing numeric values are shown here. The Monthly "
            "rule column explains whether AtlasFlow used Average, Sum, or Latest value."
        )
        st.dataframe(
            format_display_dataframe(detail_df),
            use_container_width=True,
            hide_index=True,
        )

    with st.expander("Source freshness and storage", expanded=False):
        freshness_rows = []
        for source_key in source_order_keys:
            source_meta = source_metadata.get(source_key, {})
            freshness_rows.append(
                {
                    "Source": COMPARISON_SOURCE_LABELS[source_key],
                    "Last API load": source_meta.get("loaded_at_local", "-"),
                    "Latest source date": source_meta.get("latest_source_date", "-"),
                    "Stored rows": int(source_meta.get("rows", 0) or 0),
                    "Monthly partitions": int(source_meta.get("partition_count", 0) or 0),
                    "Refresh mode": source_meta.get("refresh_mode", "-"),
                }
            )
        st.dataframe(
            pd.DataFrame(freshness_rows),
            use_container_width=True,
            hide_index=True,
        )

    export_signature = sha256(
        "|".join(
            [
                selected_start.isoformat(),
                selected_end.isoformat(),
                ",".join(selected_vessels),
                ",".join(selected_fields),
                str(len(availability_df)),
                str(len(detail_df)),
            ]
        ).encode("utf-8")
    ).hexdigest()
    if st.session_state.get("atlas_monthly_export_signature") != export_signature:
        st.session_state.pop("atlas_monthly_export_bytes", None)

    if st.button(
        "Prepare monthly comparison Excel",
        type="primary",
        disabled=availability_df.empty,
    ):
        with st.spinner("Preparing monthly comparison workbook..."):
            st.session_state["atlas_monthly_export_bytes"] = (
                to_dynamic_monthly_comparison_excel_bytes(
                    availability_df,
                    detail_df,
                    mapping_df,
                )
            )
            st.session_state["atlas_monthly_export_signature"] = export_signature

    if (
        st.session_state.get("atlas_monthly_export_signature") == export_signature
        and "atlas_monthly_export_bytes" in st.session_state
    ):
        st.download_button(
            "Download monthly comparison Excel",
            data=st.session_state["atlas_monthly_export_bytes"],
            file_name="atlasflow_dynamic_monthly_comparison.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        st.caption(
            "The workbook is prepared on demand and includes availability, "
            "monthly numeric values, and exact source-field mappings."
        )

def run_warmup_if_requested() -> None:
    """Build or incrementally refresh one or all prepared AtlasFlow snapshots."""
    if not is_warmup_request():
        return

    apply_custom_css()
    if not warmup_token_is_valid():
        st.error("Invalid or missing warmup token.")
        st.stop()

    username = read_secret("MARORKA_USERNAME")
    password = read_secret("MARORKA_PASSWORD")
    token = read_secret("MARORKA_TOKEN")
    auth_method = read_secret("MARORKA_AUTH_METHOD", "basic")
    requested_source = get_query_param("source", "reportdata").strip().lower()
    force_refresh = get_query_param("force", "0") == "1"
    full_refresh = get_query_param("full", "0") == "1"

    if auth_method.lower() in {"basic", "digest"} and (not username or not password):
        st.error("Warmup failed: MARORKA_USERNAME and MARORKA_PASSWORD are required.")
        st.stop()

    valid_sources = ["reportdata", "reportpivots", "shippivots"]
    if requested_source == "all":
        requested_sources = valid_sources
    elif requested_source in valid_sources:
        requested_sources = [requested_source]
    else:
        st.error("Invalid warmup source. Use reportdata, reportpivots, shippivots, or all.")
        st.stop()

    warmup_started_at = time.perf_counter()
    results: dict[str, dict[str, Any]] = {}
    failures: dict[str, str] = {}

    for source_key in requested_sources:
        try:
            if force_refresh:
                with st.spinner(f"Refreshing and preparing {SOURCE_CONFIGS[source_key]['label']}..."):
                    metadata, manifest = refresh_source_snapshot(
                        source_key,
                        username,
                        password,
                        token,
                        auth_method,
                        full_refresh=full_refresh,
                    )
            else:
                existing = ensure_source_snapshot(source_key, username, auth_method)
                if existing is None:
                    with st.spinner(f"Creating the first prepared {SOURCE_CONFIGS[source_key]['label']} snapshot..."):
                        metadata, manifest = refresh_source_snapshot(
                            source_key,
                            username,
                            password,
                            token,
                            auth_method,
                            full_refresh=True,
                        )
                else:
                    metadata, manifest = existing

            # ReportData is used by the default workspace, so seed its shared read cache.
            if source_key == "reportdata":
                signature = atlas_source_signature(source_key, username, auth_method, API_FULL_START_DATE)
                load_source_snapshot(source_key, signature, API_FULL_START_DATE)

            results[source_key] = {
                "refresh_mode": metadata.get("refresh_mode", "snapshot_only"),
                "snapshot_generation": manifest.get("generation"),
                "last_api_load_local": metadata.get("loaded_at_local", "-"),
                "rows": int(metadata.get("rows", 0) or 0),
                "columns": int(metadata.get("columns", 0) or 0),
                "monthly_partitions": int(metadata.get("partition_count", 0) or 0),
                "api_pages_last_refresh": int(metadata.get("pages", 0) or 0),
                "refresh_api_start_date": metadata.get("refresh_api_start_date", "-"),
                "refresh_skipped_due_to_lock": bool(metadata.get("refresh_skipped_due_to_lock", False)),
            }
        except AtlasRefreshAlreadyRunning as exc:
            failures[source_key] = f"already running: {exc}"
        except requests.HTTPError as exc:
            status = exc.response.status_code if exc.response is not None else "unknown"
            failures[source_key] = f"Marorka HTTP status {status}"
        except (
            MarorkaConfigError,
            ValueError,
            FileNotFoundError,
            RuntimeError,
            TimeoutError,
            OSError,
            requests.RequestException,
        ) as exc:
            failures[source_key] = str(exc)
        finally:
            gc.collect()

    if results:
        st.success("AtlasFlow prepared snapshot warmup completed.")
        st.write(
            {
                "sources": results,
                "force_refresh": force_refresh,
                "full_refresh": full_refresh,
                "warmup_seconds": round(time.perf_counter() - warmup_started_at, 2),
            }
        )
    if failures:
        st.error(
            "Some AtlasFlow sources were not refreshed: "
            + "; ".join(f"{source}: {error}" for source, error in failures.items())
        )
        for source_key in failures:
            st.caption(source_refresh_status_summary(source_key))
    st.stop()


# =============================================================================
# Voyage Analysis workspace
# =============================================================================

CARGO_REPORT_IDENTITY_COLUMNS = [
    "ReportId", "ShipName", "ReportType", "StartDateTimeGMT", "EndDateTimeGMT", "StateName"
]

CARGO_SUMMARY_FIELDS = [
    "Cargo Weight [tons]",
    "Total Number Full Units (20 and 40ft)",
    "Total Number Empty Units (20 and 40ft)",
    "Total Number Reefer Units (20 and 40ft)",
    "Total Number DG Units (20 and 40ft)",
    "Total Units Weight (All Categories)",
]

CARGO_CONTAINER_FIELDS = [
    "20ft Full Units", "40ft Full Units", "20ft Empty Units", "40ft Empty Units",
    "20ft Reefer Units", "40ft Reefer Units", "20ft DG Units", "40ft DG Units",
    "20ft Full Units Weight [tons]", "40ft Full Units Weight [tons]",
    "20ft Empty Units Weight [tons]", "40ft Empty Units Weight [tons]",
    "Reefer Units Weight [tons]", "DG Units Weight [tons]",
    "Total Number Full Units (20 and 40ft)", "Total Number Empty Units (20 and 40ft)",
    "Total Number Reefer Units (20 and 40ft)", "Total Number DG Units (20 and 40ft)",
    "Total Number of 20ft Units (Full and Empty)", "Total Number of 40ft Units (Full and Empty)",
    "Total Full Units Weight (20 and 40ft) [tons]", "Total Empty Units Weight (20 and 40ft) [tons]",
    "Total Units Weight (All Categories)",
]

CARGO_OPERATION_FIELDS = [
    "Cargo Weight Added [MT]", "Cargo Weight Removed [MT]",
    "TEU Loaded Units", "TEU Loaded Weight [tons]", "TEU Discharged Units", "TEU Discharged Weight [tons]",
    "FEU Loaded Units", "FEU Loaded Weight [tons]", "FEU Discharged Units", "FEU Discharged Weight [tons]",
    "Reefers Loaded Units", "Reefers Loaded Weight [tons]", "Reefers Discharged Units", "Reefers Discharged Weight [tons]",
    "Commenced Cargo Operation Time [dd:mm:yyyy hh:mm]", "Completed Cargo Operation Time [dd:mm:yyyy hh:mm]",
    "Cargo Operations Completed During Port Stay", "Cargo Checked: Bridges", "Cargo Checked: Lashings",
]

CARGO_DRAFT_FIELDS = [
    "Draft Forward [m] (m)", "Draft Aft [m] (m)",
    "Observed Draft Forward [m]", "Observed Draft Aft [m]", "Observed Mean Draft [m]",
    "Calculated Draft Forward [m]", "Calculated Draft Aft [m]", "Calculated Mean Draft [m]",
    "Ballast Amount [tons]", "Dead Load [tons]", "Air Draft [m]",
]

# Voyage fuel totals use the existing compact ReportData fuel-grade fields.
# The categories intentionally match the operational overview requested by the user:
# MGO, HFO-family grades, LFO-family grades, and the grand total for the voyage.
VOYAGE_FUEL_GRADE_COLUMNS = {
    "MGO": [
        "Main Engine - MGO", "Diesel Generator - MGO",
        "Auxiliary Engine - MGO", "Boiler - MGO",
    ],
    "HFO": [
        "Main Engine - HSHFO", "Main Engine - ULSHFO", "Main Engine - VLSHFO",
        "Diesel Generator - HSHFO", "Diesel Generator - ULSHFO", "Diesel Generator - VLSHFO",
        "Auxiliary Engine - HSHFO", "Auxiliary Engine - ULSHFO", "Auxiliary Engine - VLSHFO",
        "Boiler - HSHFO", "Boiler - ULSHFO", "Boiler - VLSHFO",
    ],
    "LFO": [
        "Main Engine - HSLFO", "Main Engine - ULSLFO", "Main Engine - VLSLFO",
        "Diesel Generator - HSLFO", "Diesel Generator - ULSLFO", "Diesel Generator - VLSLFO",
        "Auxiliary Engine - HSLFO", "Auxiliary Engine - ULSLFO", "Auxiliary Engine - VLSLFO",
        "Boiler - HSLFO", "Boiler - ULSLFO", "Boiler - VLSLFO",
    ],
}
VOYAGE_FUEL_CANONICAL_COLUMNS = sorted(
    {column for columns in VOYAGE_FUEL_GRADE_COLUMNS.values() for column in columns},
    key=str.casefold,
)
VOYAGE_FUEL_ALIAS_TO_CANONICAL = {
    normalize_text(alias): canonical
    for canonical in VOYAGE_FUEL_CANONICAL_COLUMNS
    for alias in PERFORMANCE_KPI_VALUE_ALIASES.get(canonical, [canonical])
}
VOYAGE_FUEL_VALUE_KEYS = set(VOYAGE_FUEL_ALIAS_TO_CANONICAL) | {
    normalize_text(alias)
    for alias in PERFORMANCE_KPI_VALUE_ALIASES.get("Total Fuel Consumed", ["Total Fuel Consumed"])
}

VOYAGE_OVERVIEW_NUMERIC_COLUMNS = [
    "Duration [days]",
    "Cargo [MT]",
    "Cargo TEU",
    "Total Fuel Consumption [MT]",
    "Total MGO Consumption [MT]",
    "Total HFO Consumption [MT]",
    "Total LFO Consumption [MT]",
]


def cargo_value_keys() -> set[str]:
    return {
        normalize_text(alias)
        for aliases in CARGO_VALUE_ALIASES.values()
        for alias in aliases
    }


def cargo_first_column(df: pd.DataFrame, candidates: list[str]) -> str | None:
    for candidate in candidates:
        if candidate in df.columns:
            return candidate
    normalized = {normalize_text(column): column for column in df.columns}
    for candidate in candidates:
        found = normalized.get(normalize_text(candidate))
        if found:
            return found
    return None


def cargo_numeric_series(df: pd.DataFrame, column: str | None) -> pd.Series:
    if not column or column not in df.columns:
        return pd.Series(pd.NA, index=df.index, dtype="Float64")
    return pd.to_numeric(df[column], errors="coerce")


def cargo_latest_value(df: pd.DataFrame, column: str | None) -> Any:
    if not column or column not in df.columns or df.empty:
        return pd.NA
    values = df[column]
    nonempty = values[values.notna() & values.astype("string").str.strip().ne("")]
    return nonempty.iloc[-1] if not nonempty.empty else pd.NA


def cargo_first_value(df: pd.DataFrame, column: str | None) -> Any:
    if not column or column not in df.columns or df.empty:
        return pd.NA
    values = df[column]
    nonempty = values[values.notna() & values.astype("string").str.strip().ne("")]
    return nonempty.iloc[0] if not nonempty.empty else pd.NA


def cargo_departure_value(
    df: pd.DataFrame,
    column: str | None,
    voyage_start: pd.Timestamp,
    voyage_end: pd.Timestamp,
    *,
    datetime_column: str = "DateTime",
    pre_departure_hours: int = 6,
) -> Any:
    """Return the cargo condition carried at voyage departure.

    Prefer the latest valid value shortly before the ShipPivots VoyageId starts.
    This captures departure reports that can be stamped a few minutes before the
    first 15-minute VoyageId point. If none exists, use the first valid value
    inside the voyage. Never use a record after the voyage end.
    """
    if not column or column not in df.columns or datetime_column not in df.columns or df.empty:
        return pd.NA
    work = df[[datetime_column, column]].copy()
    work[datetime_column] = pd.to_datetime(work[datetime_column], errors="coerce", utc=True)
    work = work[work[datetime_column].notna() & work[datetime_column].le(voyage_end)].sort_values(datetime_column)
    valid = work[column].notna() & work[column].astype("string").str.strip().ne("")
    work = work.loc[valid]
    if work.empty:
        return pd.NA
    pre_start = voyage_start - pd.Timedelta(hours=pre_departure_hours)
    before = work[work[datetime_column].ge(pre_start) & work[datetime_column].le(voyage_start)]
    if not before.empty:
        return before.iloc[-1][column]
    inside = work[work[datetime_column].gt(voyage_start) & work[datetime_column].le(voyage_end)]
    return inside.iloc[0][column] if not inside.empty else pd.NA


def report_event_time(df: pd.DataFrame) -> pd.Series:
    end_values = pd.to_datetime(df.get("EndDateTimeGMT"), errors="coerce", utc=True)
    start_values = pd.to_datetime(df.get("StartDateTimeGMT"), errors="coerce", utc=True)
    return end_values.fillna(start_values)


def fuel_reportdata_selected(long_df: pd.DataFrame, selected_vessels: list[str]) -> pd.DataFrame:
    if long_df.empty or "ValueDescription" not in long_df.columns or "ShipName" not in long_df.columns:
        return pd.DataFrame(columns=CARGO_REPORT_IDENTITY_COLUMNS)
    keys = long_df["ValueDescription"].map(normalize_text)
    mask = match_selected_vessels(long_df["ShipName"], selected_vessels) & keys.isin(VOYAGE_FUEL_VALUE_KEYS)
    return long_df.loc[mask].copy()


def pivot_fuel_reports(fuel_long: pd.DataFrame) -> pd.DataFrame:
    if fuel_long.empty:
        return pd.DataFrame(columns=[*CARGO_REPORT_IDENTITY_COLUMNS, "MGO Consumption [MT]", "HFO Consumption [MT]", "LFO Consumption [MT]", "Total Fuel Consumption [MT]"])
    work = fuel_long.copy()
    work["_key"] = work["ValueDescription"].map(normalize_text)
    total_fuel_keys = {normalize_text(alias) for alias in PERFORMANCE_KPI_VALUE_ALIASES.get("Total Fuel Consumed", ["Total Fuel Consumed"])}
    work["_canonical"] = work["_key"].map(VOYAGE_FUEL_ALIAS_TO_CANONICAL)
    work.loc[work["_key"].isin(total_fuel_keys), "_canonical"] = "Total Fuel Consumed"
    work = work[work["_canonical"].notna()].copy()
    if work.empty:
        return pd.DataFrame(columns=[*CARGO_REPORT_IDENTITY_COLUMNS, "MGO Consumption [MT]", "HFO Consumption [MT]", "LFO Consumption [MT]", "Total Fuel Consumption [MT]"])
    work["_value"] = pd.to_numeric(work["ParsedValue"], errors="coerce")
    work["_source_order"] = range(len(work))
    work = work.sort_values("_source_order").drop_duplicates(
        [*CARGO_REPORT_IDENTITY_COLUMNS, "_canonical"], keep="last"
    )
    pivot = work.pivot(index=CARGO_REPORT_IDENTITY_COLUMNS, columns="_canonical", values="_value").reset_index()
    pivot.columns.name = None
    for grade, columns in VOYAGE_FUEL_GRADE_COLUMNS.items():
        present = [column for column in columns if column in pivot.columns]
        if present:
            pivot[f"{grade} Consumption [MT]"] = pivot[present].apply(pd.to_numeric, errors="coerce").sum(axis=1, min_count=1)
        else:
            pivot[f"{grade} Consumption [MT]"] = pd.NA
    grade_columns = [f"{grade} Consumption [MT]" for grade in ["MGO", "HFO", "LFO"]]
    grade_total = pivot[grade_columns].apply(pd.to_numeric, errors="coerce").sum(axis=1, min_count=1)
    official_total = pd.to_numeric(pivot.get("Total Fuel Consumed"), errors="coerce") if "Total Fuel Consumed" in pivot.columns else pd.Series(pd.NA, index=pivot.index, dtype="Float64")
    pivot["Total Fuel Consumption [MT]"] = grade_total.fillna(official_total)
    pivot["StartDateTimeGMT"] = pd.to_datetime(pivot["StartDateTimeGMT"], errors="coerce", utc=True)
    pivot["EndDateTimeGMT"] = pd.to_datetime(pivot["EndDateTimeGMT"], errors="coerce", utc=True)
    return pivot.sort_values(["EndDateTimeGMT", "StartDateTimeGMT", "ReportId"], na_position="last").reset_index(drop=True)


def fuel_reports_for_voyage(
    fuel_reports: pd.DataFrame,
    vessel: str,
    voyage_start: pd.Timestamp,
    voyage_end: pd.Timestamp,
) -> pd.DataFrame:
    if fuel_reports.empty:
        return fuel_reports.copy()
    vessel_df = _cargo_subset_vessel(fuel_reports, vessel)
    event_time = report_event_time(vessel_df)
    # Consumption reported at a report belongs to the interval ending at that
    # report. Use only reports completing after voyage start and no later than
    # voyage end, so the grand total is restricted to the whole VoyageId span.
    return vessel_df.loc[event_time.gt(voyage_start) & event_time.le(voyage_end)].copy()


def voyage_fuel_totals(fuel_reports_voyage: pd.DataFrame) -> dict[str, Any]:
    totals: dict[str, Any] = {}
    for column in ["MGO Consumption [MT]", "HFO Consumption [MT]", "LFO Consumption [MT]"]:
        values = pd.to_numeric(fuel_reports_voyage.get(column), errors="coerce") if column in fuel_reports_voyage.columns else pd.Series(dtype="float64")
        totals[column] = values.sum(min_count=1) if not values.empty else pd.NA
    grade_values = pd.Series([totals[column] for column in ["MGO Consumption [MT]", "HFO Consumption [MT]", "LFO Consumption [MT]"]], dtype="Float64")
    total = grade_values.sum(min_count=1)
    if pd.isna(total) and "Total Fuel Consumption [MT]" in fuel_reports_voyage.columns:
        total = pd.to_numeric(fuel_reports_voyage["Total Fuel Consumption [MT]"], errors="coerce").sum(min_count=1)
    totals["Total Fuel Consumption [MT]"] = total
    return totals


def cargo_format_number(value: Any, decimals: int = 1) -> str:
    numeric = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
    if pd.isna(numeric):
        return "-"
    return f"{float(numeric):,.{decimals}f}"


def build_voyage_catalog(shippivots_df: pd.DataFrame) -> pd.DataFrame:
    if shippivots_df.empty or "VoyageId" not in shippivots_df.columns or "DateTime" not in shippivots_df.columns:
        return pd.DataFrame(columns=["VoyageId", "VoyageIdInternal", "VoyageStart", "VoyageEnd", "Points"])
    work = shippivots_df.copy()
    work["DateTime"] = pd.to_datetime(work["DateTime"], errors="coerce", utc=True)
    work["VoyageId"] = work["VoyageId"].astype("string").str.strip()
    work = work[work["DateTime"].notna() & work["VoyageId"].notna() & work["VoyageId"].ne("")].copy()
    if work.empty:
        return pd.DataFrame(columns=["VoyageId", "VoyageIdInternal", "VoyageStart", "VoyageEnd", "Points"])
    agg_map: dict[str, Any] = {"VoyageStart": ("DateTime", "min"), "VoyageEnd": ("DateTime", "max"), "Points": ("DateTime", "size")}
    if "VoyageIdInternal" in work.columns:
        catalog = work.groupby("VoyageId", dropna=False).agg(
            VoyageStart=("DateTime", "min"),
            VoyageEnd=("DateTime", "max"),
            Points=("DateTime", "size"),
            VoyageIdInternal=("VoyageIdInternal", lambda s: cargo_latest_value(pd.DataFrame({"x": s}), "x")),
        ).reset_index()
    else:
        catalog = work.groupby("VoyageId", dropna=False).agg(
            VoyageStart=("DateTime", "min"), VoyageEnd=("DateTime", "max"), Points=("DateTime", "size")
        ).reset_index()
        catalog["VoyageIdInternal"] = pd.NA
    return catalog.sort_values("VoyageEnd", ascending=False).reset_index(drop=True)


def cargo_voyage_label(row: pd.Series) -> str:
    start = pd.to_datetime(row.get("VoyageStart"), errors="coerce", utc=True)
    end = pd.to_datetime(row.get("VoyageEnd"), errors="coerce", utc=True)
    start_text = start.strftime("%d/%m/%Y %H:%M") if not pd.isna(start) else "-"
    end_text = end.strftime("%d/%m/%Y %H:%M") if not pd.isna(end) else "-"
    internal_value = row.get("VoyageIdInternal")
    internal = "" if pd.isna(internal_value) else str(internal_value).strip()
    internal_text = f" | {internal}" if internal and internal.lower() not in {"<na>", "nan", "none"} else ""
    return f"{row.get('VoyageId', '-')}{internal_text} | {start_text} → {end_text}"


def cargo_reportdata_for_voyage(
    long_df: pd.DataFrame,
    vessel: str,
    voyage_start: pd.Timestamp,
    voyage_end: pd.Timestamp,
) -> pd.DataFrame:
    if long_df.empty:
        return long_df.copy()
    vessel_mask = match_selected_vessels(long_df["ShipName"], [vessel])
    start_values = pd.to_datetime(long_df["StartDateTimeGMT"], errors="coerce", utc=True)
    end_values = pd.to_datetime(long_df["EndDateTimeGMT"], errors="coerce", utc=True)
    # A small pre-departure allowance captures departure cargo reports that can
    # be stamped shortly before ShipPivots switches to the new VoyageId.
    # Deliberately do not include anything after voyage_end.
    margin = pd.Timedelta(hours=6)
    time_mask = (
        (start_values.ge(voyage_start - margin) & start_values.le(voyage_end))
        | (end_values.ge(voyage_start - margin) & end_values.le(voyage_end))
    )
    keys = long_df["ValueDescription"].map(normalize_text)
    return long_df.loc[vessel_mask & time_mask & keys.isin(cargo_value_keys())].copy()


def pivot_cargo_reports(cargo_long: pd.DataFrame) -> pd.DataFrame:
    if cargo_long.empty:
        return pd.DataFrame(columns=CARGO_REPORT_IDENTITY_COLUMNS)
    work = cargo_long.copy()
    work["_source_order"] = range(len(work))
    work = work.sort_values("_source_order").drop_duplicates(
        [*CARGO_REPORT_IDENTITY_COLUMNS, "ValueDescription"], keep="last"
    )
    values = work["ParsedValue"].where(work["ParsedValue"].notna(), work["ReportedValue"])
    work["_cargo_value"] = values
    pivot = work.pivot(index=CARGO_REPORT_IDENTITY_COLUMNS, columns="ValueDescription", values="_cargo_value").reset_index()
    pivot.columns.name = None
    pivot["StartDateTimeGMT"] = pd.to_datetime(pivot["StartDateTimeGMT"], errors="coerce", utc=True)
    pivot["EndDateTimeGMT"] = pd.to_datetime(pivot["EndDateTimeGMT"], errors="coerce", utc=True)
    return pivot.sort_values(["StartDateTimeGMT", "ReportId"], ascending=[True, True]).reset_index(drop=True)


def filter_reportpivots_for_voyage(
    reportpivots_df: pd.DataFrame,
    voyage_start: pd.Timestamp,
    voyage_end: pd.Timestamp,
) -> pd.DataFrame:
    if reportpivots_df.empty or "DateTime" not in reportpivots_df.columns:
        return reportpivots_df.copy()
    result = reportpivots_df.copy()
    result["DateTime"] = pd.to_datetime(result["DateTime"], errors="coerce", utc=True)
    return result[result["DateTime"].ge(voyage_start) & result["DateTime"].le(voyage_end)].copy().sort_values("DateTime")


def build_cargo_evolution_series(
    rp_voyage: pd.DataFrame,
    voyage_start: pd.Timestamp,
    voyage_end: pd.Timestamp,
    cargo_weight_col: str | None,
    cargo_teu_col: str | None,
    initial_cargo_weight: Any = pd.NA,
    initial_cargo_teu: Any = pd.NA,
) -> pd.DataFrame:
    """Build continuous onboard-cargo state across the whole selected voyage.

    ReportPivots cargo values are sparse: most sea/COSP/EOSP rows do not repeat
    CargoWeight/CargoTEU. For voyage analysis those values represent onboard
    state, so carry the latest observed value forward until a new cargo value is
    reported. Seed the series from the voyage-overview departure cargo value and
    always include voyage start/end so a single known cargo state still renders
    as a visible line across the voyage.
    """
    if pd.isna(voyage_start) or pd.isna(voyage_end):
        return pd.DataFrame(columns=["DateTime", "Cargo Weight [MT]", "Cargo TEU"])

    base_times = pd.Series([voyage_start, voyage_end], dtype="datetime64[ns, UTC]")
    work = rp_voyage.copy() if isinstance(rp_voyage, pd.DataFrame) else pd.DataFrame()
    if not work.empty and "DateTime" in work.columns:
        rp_times = pd.to_datetime(work["DateTime"], errors="coerce", utc=True).dropna()
        base_times = pd.concat([base_times, rp_times], ignore_index=True)

    evolution = pd.DataFrame({"DateTime": base_times.dropna().drop_duplicates().sort_values()})
    evolution["Cargo Weight [MT]"] = pd.NA
    evolution["Cargo TEU"] = pd.NA

    # Seed the voyage with the cargo condition selected for the overview row.
    initial_weight = pd.to_numeric(pd.Series([initial_cargo_weight]), errors="coerce").iloc[0]
    initial_teu = pd.to_numeric(pd.Series([initial_cargo_teu]), errors="coerce").iloc[0]
    if pd.notna(initial_weight):
        evolution.loc[evolution["DateTime"].eq(voyage_start), "Cargo Weight [MT]"] = float(initial_weight)
    if pd.notna(initial_teu):
        evolution.loc[evolution["DateTime"].eq(voyage_start), "Cargo TEU"] = float(initial_teu)

    if not work.empty and "DateTime" in work.columns:
        work["DateTime"] = pd.to_datetime(work["DateTime"], errors="coerce", utc=True)
        work = work[work["DateTime"].notna()].copy()
        observation_frames: list[pd.DataFrame] = []
        if cargo_weight_col and cargo_weight_col in work.columns:
            weight_obs = work[["DateTime", cargo_weight_col]].copy()
            weight_obs["Cargo Weight [MT]"] = pd.to_numeric(weight_obs[cargo_weight_col], errors="coerce")
            observation_frames.append(weight_obs[["DateTime", "Cargo Weight [MT]"]])
        if cargo_teu_col and cargo_teu_col in work.columns:
            teu_obs = work[["DateTime", cargo_teu_col]].copy()
            teu_obs["Cargo TEU"] = pd.to_numeric(teu_obs[cargo_teu_col], errors="coerce")
            observation_frames.append(teu_obs[["DateTime", "Cargo TEU"]])

        for observations in observation_frames:
            value_column = next(column for column in observations.columns if column != "DateTime")
            observations = observations.dropna(subset=[value_column]).sort_values("DateTime")
            if observations.empty:
                continue
            observations = observations.drop_duplicates("DateTime", keep="last")
            mapping = observations.set_index("DateTime")[value_column]
            evolution[value_column] = evolution["DateTime"].map(mapping).combine_first(evolution[value_column])

    for value_column in ["Cargo Weight [MT]", "Cargo TEU"]:
        evolution[value_column] = pd.to_numeric(evolution[value_column], errors="coerce").ffill()

    return evolution.sort_values("DateTime").reset_index(drop=True)


def cargo_report_timeline(cargo_by_report: pd.DataFrame, rp_voyage: pd.DataFrame) -> pd.DataFrame:
    frames: list[pd.DataFrame] = []
    if not cargo_by_report.empty:
        rd = cargo_by_report.copy()
        rd["DateTime"] = pd.to_datetime(rd["EndDateTimeGMT"], errors="coerce", utc=True).fillna(
            pd.to_datetime(rd["StartDateTimeGMT"], errors="coerce", utc=True)
        )
        rd["Source"] = "ReportData"
        keep = [column for column in ["DateTime", "Source", "ReportId", "ReportType", "StateName", "Cargo Weight [tons]", "TEU Loaded Units", "TEU Discharged Units", "Reefers Loaded Units", "Reefers Discharged Units"] if column in rd.columns]
        frames.append(rd[keep])
    if not rp_voyage.empty:
        rp = rp_voyage.copy()
        rp["Source"] = "ReportPivots"
        if "ReportType" not in rp.columns:
            rp["ReportType"] = "ReportPivots"
        if "StateName" not in rp.columns:
            rp["StateName"] = pd.NA
        if "ReportId" not in rp.columns:
            rp["ReportId"] = pd.NA
        keep = [column for column in ["DateTime", "Source", "ReportId", "ReportType", "StateName", "CargoWeight", "CargoTEU", "DeparturePort", "ArrivalPort", "DraftFore", "DraftAft"] if column in rp.columns]
        frames.append(rp[keep])
    if not frames:
        return pd.DataFrame()
    timeline = pd.concat(frames, ignore_index=True, sort=False)
    timeline["DateTime"] = pd.to_datetime(timeline["DateTime"], errors="coerce", utc=True)
    return timeline.sort_values(["DateTime", "Source"], na_position="last").reset_index(drop=True)


def cargo_excel_bytes(
    overview: pd.DataFrame,
    timeline: pd.DataFrame,
    cargo_by_report: pd.DataFrame,
    reportpivots: pd.DataFrame,
    shippivots: pd.DataFrame,
    fuel_by_report: pd.DataFrame | None = None,
) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        write_table_sheet(writer, overview, "Voyage Overview", "CargoVoyageOverview")
        if not timeline.empty:
            write_table_sheet(writer, timeline, "Report Timeline", "CargoReportTimeline")
        if not cargo_by_report.empty:
            write_table_sheet(writer, cargo_by_report, "Cargo By Report", "CargoByReport")
        if fuel_by_report is not None and not fuel_by_report.empty:
            write_table_sheet(writer, fuel_by_report, "Fuel By Report", "CargoFuelByReport")
        if not reportpivots.empty:
            write_table_sheet(writer, reportpivots, "ReportPivots", "CargoReportPivots")
        if not shippivots.empty:
            write_table_sheet(writer, shippivots, "ShipPivots", "CargoShipPivots")
    return output.getvalue()


def _cargo_subset_vessel(df: pd.DataFrame, vessel: str) -> pd.DataFrame:
    if df.empty or "ShipName" not in df.columns:
        return df.copy()
    return df.loc[match_selected_vessels(df["ShipName"], [vessel])].copy()


def build_cargo_voyage_overview(
    ship_df: pd.DataFrame,
    rp_df: pd.DataFrame,
    long_df: pd.DataFrame,
    selected_vessels: list[str],
) -> pd.DataFrame:
    """Build one compact row per whole VoyageId for mass review.

    ShipPivots defines the voyage boundaries. ReportPivots route values are
    taken only from inside those boundaries. Cargo condition may use a small
    pre-departure allowance because the departure cargo report can be stamped
    minutes before the first VoyageId point. Fuel is summed only from reports
    whose reporting interval completes inside the voyage.
    """
    if not selected_vessels or ship_df.empty:
        return pd.DataFrame()

    ship_work = ship_df.copy()
    if "DateTime" in ship_work.columns:
        ship_work["DateTime"] = pd.to_datetime(ship_work["DateTime"], errors="coerce", utc=True)
    rp_work = rp_df.copy()
    if "DateTime" in rp_work.columns:
        rp_work["DateTime"] = pd.to_datetime(rp_work["DateTime"], errors="coerce", utc=True)

    cargo_reports_all = pd.DataFrame(columns=CARGO_REPORT_IDENTITY_COLUMNS)
    if isinstance(long_df, pd.DataFrame) and not long_df.empty and "ShipName" in long_df.columns and "ValueDescription" in long_df.columns:
        vessel_mask = match_selected_vessels(long_df["ShipName"], selected_vessels)
        cargo_key_mask = long_df["ValueDescription"].map(normalize_text).isin(cargo_value_keys())
        cargo_selected = long_df.loc[vessel_mask & cargo_key_mask].copy()
        if not cargo_selected.empty:
            cargo_reports_all = pivot_cargo_reports(cargo_selected)
        del cargo_selected

    fuel_reports_all = pivot_fuel_reports(fuel_reportdata_selected(long_df, selected_vessels))

    rows: list[dict[str, Any]] = []
    pre_departure_margin = pd.Timedelta(hours=6)
    for vessel in selected_vessels:
        vessel_ship = _cargo_subset_vessel(ship_work, vessel)
        vessel_rp = _cargo_subset_vessel(rp_work, vessel)
        vessel_cargo_reports = _cargo_subset_vessel(cargo_reports_all, vessel) if not cargo_reports_all.empty else cargo_reports_all.copy()

        catalog = build_voyage_catalog(vessel_ship)
        if catalog.empty:
            continue

        rp_dates = pd.to_datetime(vessel_rp.get("DateTime"), errors="coerce", utc=True) if "DateTime" in vessel_rp.columns else pd.Series(pd.NaT, index=vessel_rp.index, dtype="datetime64[ns, UTC]")
        cargo_events = report_event_time(vessel_cargo_reports) if not vessel_cargo_reports.empty else pd.Series(pd.NaT, index=vessel_cargo_reports.index, dtype="datetime64[ns, UTC]")

        for _, voyage in catalog.iterrows():
            voyage_id = str(voyage.get("VoyageId", ""))
            voyage_start = pd.to_datetime(voyage.get("VoyageStart"), errors="coerce", utc=True)
            voyage_end = pd.to_datetime(voyage.get("VoyageEnd"), errors="coerce", utc=True)
            if pd.isna(voyage_start) or pd.isna(voyage_end):
                continue

            # Strict route window: never allow the next port call/voyage to
            # overwrite Departure/Arrival for this VoyageId.
            rp_route = vessel_rp.loc[rp_dates.ge(voyage_start) & rp_dates.le(voyage_end)].copy().sort_values("DateTime")
            # Cargo condition can legitimately be a few minutes before the first
            # ShipPivots point, but never after voyage_end.
            rp_cargo = vessel_rp.loc[rp_dates.ge(voyage_start - pre_departure_margin) & rp_dates.le(voyage_end)].copy().sort_values("DateTime")
            cargo_by_report = vessel_cargo_reports.loc[cargo_events.ge(voyage_start - pre_departure_margin) & cargo_events.le(voyage_end)].copy()

            cargo_weight_col = cargo_first_column(rp_cargo, ["CargoWeight", "Cargo Weight", "CargoMT"])
            cargo_teu_col = cargo_first_column(rp_cargo, ["CargoTEU", "Cargo TEU", "TEU"])
            dep_col = cargo_first_column(rp_route, ["DeparturePort", "Departure Port", "PortFrom"])
            arr_col = cargo_first_column(rp_route, ["ArrivalPort", "Arrival Port", "PortTo"])

            cargo_weight = cargo_departure_value(rp_cargo, cargo_weight_col, voyage_start, voyage_end)
            cargo_teu = cargo_departure_value(rp_cargo, cargo_teu_col, voyage_start, voyage_end)

            # ReportData carries the full port names. Prefer the route stamped on
            # the departure cargo/report just before the VoyageId starts, then
            # fall back to the ReportPivots UN/LOCODE. This gives users e.g.
            # "Xiamen → Kaohsiung" instead of "CNXMN → TWKHH".
            departure_port = cargo_first_value(rp_route, dep_col)
            arrival_port = cargo_first_value(rp_route, arr_col)
            if not cargo_by_report.empty:
                report_times = report_event_time(cargo_by_report)
                temp = cargo_by_report.copy()
                temp["_ReportEventTime"] = report_times
                if pd.isna(pd.to_numeric(pd.Series([cargo_weight]), errors="coerce").iloc[0]) and "Cargo Weight [tons]" in temp.columns:
                    cargo_weight = cargo_departure_value(temp, "Cargo Weight [tons]", voyage_start, voyage_end, datetime_column="_ReportEventTime")
                if "Departure port name" in temp.columns:
                    full_departure = cargo_departure_value(temp, "Departure port name", voyage_start, voyage_end, datetime_column="_ReportEventTime")
                    if pd.notna(full_departure) and str(full_departure).strip():
                        departure_port = full_departure
                if "Arrival port name" in temp.columns:
                    full_arrival = cargo_departure_value(temp, "Arrival port name", voyage_start, voyage_end, datetime_column="_ReportEventTime")
                    if pd.notna(full_arrival) and str(full_arrival).strip():
                        arrival_port = full_arrival

            fuel_voyage = fuel_reports_for_voyage(fuel_reports_all, vessel, voyage_start, voyage_end)
            fuel_totals = voyage_fuel_totals(fuel_voyage)

            rows.append({
                "Vessel": vessel,
                "VoyageId": voyage_id,
                "Start": voyage_start,
                "End": voyage_end,
                "Duration [days]": max((voyage_end - voyage_start).total_seconds() / 86400.0, 0.0),
                "Departure": departure_port,
                "Arrival": arrival_port,
                "Cargo [MT]": cargo_weight,
                "Cargo TEU": cargo_teu,
                "Total Fuel Consumption [MT]": fuel_totals["Total Fuel Consumption [MT]"],
                "Total MGO Consumption [MT]": fuel_totals["MGO Consumption [MT]"],
                "Total HFO Consumption [MT]": fuel_totals["HFO Consumption [MT]"],
                "Total LFO Consumption [MT]": fuel_totals["LFO Consumption [MT]"],
            })

    if not rows:
        return pd.DataFrame()
    result = pd.DataFrame(rows)
    result["VoyageId"] = result["VoyageId"].astype("string").fillna("-").str.strip().replace("", "-")
    for column in VOYAGE_OVERVIEW_NUMERIC_COLUMNS:
        if column in result.columns:
            result[column] = pd.to_numeric(result[column], errors="coerce").round(2)
    return result.sort_values(["Start", "Vessel", "VoyageId"], ascending=[False, True, True]).reset_index(drop=True)


def apply_cargo_overview_excel_formats(worksheet: Any, overview: pd.DataFrame) -> None:
    """Keep voyage export values numeric and display thousands with up to 2 decimals."""
    column_positions = {str(name): idx + 1 for idx, name in enumerate(overview.columns)}
    for column in VOYAGE_OVERVIEW_NUMERIC_COLUMNS:
        col_idx = column_positions.get(column)
        if not col_idx:
            continue
        for row_idx in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                cell.number_format = '#,##0.00'
    for column in ["Start", "End"]:
        col_idx = column_positions.get(column)
        if not col_idx:
            continue
        for row_idx in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                cell.number_format = 'dd/mm/yyyy hh:mm'


def cargo_overview_excel_bytes(overview: pd.DataFrame) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        write_table_sheet(writer, overview, "Voyage Overview", "CargoVoyageOverview")
        apply_cargo_overview_excel_formats(writer.sheets["Voyage Overview"], overview)
    return output.getvalue()


def render_cargo_voyages_workspace(
    username: str,
    password: str,
    token: str,
    auth_method: str,
    api_start_date: date,
    long_df: pd.DataFrame,
    reportdata_metadata: dict[str, Any],
    selected_vessels: list[str],
    selected_start: date,
    selected_end: date,
) -> None:
    st.markdown('<div class="section-title">Voyage Analysis</div>', unsafe_allow_html=True)
    st.caption(
        "Compact voyage overview for the selected vessels and period. Open Voyage Detail only when a specific voyage needs investigation."
    )
    if not selected_vessels:
        st.info("Select at least one vessel in the sidebar.")
        return

    # Load the prepared wide sources once for the complete sidebar selection.
    ship_df, ship_meta = load_wide_source_for_view(
        "shippivots", username, password, token, auth_method, api_start_date, False,
        selected_vessels, selected_start, selected_end,
    )
    rp_df, rp_meta = load_wide_source_for_view(
        "reportpivots", username, password, token, auth_method, api_start_date, False,
        selected_vessels, selected_start, selected_end,
    )
    if ship_meta.get("needs_warmup") or rp_meta.get("needs_warmup"):
        missing = []
        if ship_meta.get("needs_warmup"):
            missing.append("ShipPivots")
        if rp_meta.get("needs_warmup"):
            missing.append("ReportPivots")
        st.info("Voyage Analysis needs prepared snapshots for: " + ", ".join(missing) + ". Run the AtlasFlow warmup first.")
        return

    overview_cache_signature = sha256(
        (
            f"{ship_meta.get('snapshot_generation')}|{rp_meta.get('snapshot_generation')}|"
            f"{reportdata_metadata.get('snapshot_generation')}|{tuple(selected_vessels)}|"
            f"{selected_start.isoformat()}|{selected_end.isoformat()}"
        ).encode("utf-8")
    ).hexdigest()
    cached_overview = st.session_state.get("atlas_cargo_overview_cache")
    if (
        st.session_state.get("atlas_cargo_overview_cache_signature") == overview_cache_signature
        and isinstance(cached_overview, pd.DataFrame)
    ):
        overview = cached_overview
    else:
        overview = build_cargo_voyage_overview(ship_df, rp_df, long_df, selected_vessels)
        st.session_state["atlas_cargo_overview_cache"] = overview
        st.session_state["atlas_cargo_overview_cache_signature"] = overview_cache_signature
    mode_options = ["Voyage Overview", "Voyage Detail"]
    mode = render_text_tab_bar(
        mode_options,
        st.session_state.get("atlas_cargo_mode", "Voyage Overview"),
        param_name="cargo_mode",
        css_class="cargo-mode",
    )
    st.session_state["atlas_cargo_mode"] = mode

    if overview.empty:
        st.info("No VoyageId values were found for the selected vessels in the selected period.")
        return

    if mode == "Voyage Overview":
        # Deliberately compact: summary line + mass-data table first.
        vessel_count = int(overview["Vessel"].nunique())
        st.markdown(
            f'<div class="atlas-pill"><span>Voyages:</span> {len(overview):,} &nbsp; | &nbsp; '
            f'<span>Vessels:</span> {vessel_count:,} &nbsp; | &nbsp; '
            f'<span>Period:</span> {selected_start.strftime("%d/%m/%Y")} → {selected_end.strftime("%d/%m/%Y")}</div>',
            unsafe_allow_html=True,
        )
        st.caption("One row per voyage. Use the existing sidebar Fleet group / Vessel / Period controls to scale from one vessel to the whole selected fleet.")
        overview_display = overview.copy()
        overview_display["VoyageId"] = overview_display["VoyageId"].astype("string").fillna("-").str.strip().replace("", "-")
        st.dataframe(
            overview_display,
            use_container_width=True,
            hide_index=True,
            column_config={
                "VoyageId": st.column_config.TextColumn("VoyageId", width="small"),
                "Start": st.column_config.DatetimeColumn("Start", format="DD/MM/YYYY HH:mm"),
                "End": st.column_config.DatetimeColumn("End", format="DD/MM/YYYY HH:mm"),
                "Duration [days]": st.column_config.NumberColumn("Duration [days]", format="%.2f"),
                "Cargo [MT]": st.column_config.NumberColumn("Cargo [MT]", format="%.2f"),
                "Cargo TEU": st.column_config.NumberColumn("Cargo TEU", format="%.2f"),
                "Total Fuel Consumption [MT]": st.column_config.NumberColumn("Total Fuel Consumption [MT]", format="%.2f"),
                "Total MGO Consumption [MT]": st.column_config.NumberColumn("Total MGO Consumption [MT]", format="%.2f"),
                "Total HFO Consumption [MT]": st.column_config.NumberColumn("Total HFO Consumption [MT]", format="%.2f"),
                "Total LFO Consumption [MT]": st.column_config.NumberColumn("Total LFO Consumption [MT]", format="%.2f"),
            },
        )

        overview_signature = sha256(
            f"{tuple(selected_vessels)}|{selected_start}|{selected_end}|{len(overview)}|{overview['Start'].max()}".encode("utf-8")
        ).hexdigest()
        overview_ready = (
            st.session_state.get("atlas_cargo_overview_export_signature") == overview_signature
            and "atlas_cargo_overview_export_bytes" in st.session_state
        )
        if st.button("Prepare overview Excel", key="atlas_prepare_cargo_overview_excel"):
            st.session_state["atlas_cargo_overview_export_bytes"] = cargo_overview_excel_bytes(overview)
            st.session_state["atlas_cargo_overview_export_signature"] = overview_signature
            overview_ready = True
        if overview_ready:
            st.download_button(
                "Download voyage overview Excel",
                data=st.session_state["atlas_cargo_overview_export_bytes"],
                file_name="atlasflow_cargo_voyage_overview.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="atlas_download_cargo_overview_excel",
            )
        with st.expander("Source freshness", expanded=False):
            st.caption(
                f"ReportData: {reportdata_metadata.get('loaded_at_local', reportdata_metadata.get('loaded_at_utc', '-'))} | "
                f"ReportPivots: {rp_meta.get('loaded_at_local', '-')} | ShipPivots: {ship_meta.get('loaded_at_local', '-')}"
            )
        return

    # Voyage Detail is intentionally opt-in. The user first chooses a vessel,
    # then one voyage from the same mass-data overview universe.
    detail_vessels = sorted(overview["Vessel"].dropna().astype(str).unique().tolist(), key=str.casefold)
    vessel = st.selectbox("Vessel", options=detail_vessels, key="atlas_cargo_detail_vessel")
    vessel_overview = overview[overview["Vessel"].astype(str).eq(vessel)].copy().reset_index(drop=True)
    detail_idx = st.selectbox(
        "Voyage",
        options=vessel_overview.index.tolist(),
        format_func=lambda idx: (
            f"{vessel_overview.loc[idx, 'VoyageId']} | "
            f"{pd.to_datetime(vessel_overview.loc[idx, 'Start'], errors='coerce', utc=True).strftime('%d/%m/%Y %H:%M')} → "
            f"{pd.to_datetime(vessel_overview.loc[idx, 'End'], errors='coerce', utc=True).strftime('%d/%m/%Y %H:%M')}"
        ),
        key="atlas_cargo_detail_voyage_idx",
    )
    voyage = vessel_overview.loc[detail_idx]
    voyage_id = str(voyage["VoyageId"])
    voyage_start = pd.to_datetime(voyage["Start"], errors="coerce", utc=True)
    voyage_end = pd.to_datetime(voyage["End"], errors="coerce", utc=True)

    detail_cache_signature = sha256(
        f"{overview_cache_signature}|{vessel}|{voyage_id}|{voyage_start}|{voyage_end}".encode("utf-8")
    ).hexdigest()
    cached_detail = st.session_state.get("atlas_cargo_detail_cache")
    if (
        st.session_state.get("atlas_cargo_detail_cache_signature") == detail_cache_signature
        and isinstance(cached_detail, dict)
    ):
        ship_voyage = cached_detail["ship_voyage"]
        rp_voyage = cached_detail["rp_voyage"]
        cargo_by_report = cached_detail["cargo_by_report"]
        timeline = cached_detail["timeline"]
        fuel_voyage = cached_detail["fuel_voyage"]
        fuel_totals = cached_detail["fuel_totals"]
    else:
        vessel_ship = _cargo_subset_vessel(ship_df, vessel)
        vessel_rp = _cargo_subset_vessel(rp_df, vessel)
        ship_work = vessel_ship.copy()
        ship_work["DateTime"] = pd.to_datetime(ship_work["DateTime"], errors="coerce", utc=True)
        ship_voyage = ship_work[
            ship_work.get("VoyageId", pd.Series(index=ship_work.index, dtype="string")).astype("string").eq(voyage_id)
        ].copy().sort_values("DateTime")
        rp_voyage = filter_reportpivots_for_voyage(vessel_rp, voyage_start, voyage_end)
        cargo_long = cargo_reportdata_for_voyage(long_df, vessel, voyage_start, voyage_end)
        cargo_by_report = pivot_cargo_reports(cargo_long)
        timeline = cargo_report_timeline(cargo_by_report, rp_voyage)
        selected_fuel_long = fuel_reportdata_selected(long_df, [vessel])
        selected_fuel_reports = pivot_fuel_reports(selected_fuel_long)
        fuel_voyage = fuel_reports_for_voyage(selected_fuel_reports, vessel, voyage_start, voyage_end)
        fuel_totals = voyage_fuel_totals(fuel_voyage)
        st.session_state["atlas_cargo_detail_cache"] = {
            "ship_voyage": ship_voyage,
            "rp_voyage": rp_voyage,
            "cargo_by_report": cargo_by_report,
            "timeline": timeline,
            "fuel_voyage": fuel_voyage,
            "fuel_totals": fuel_totals,
        }
        st.session_state["atlas_cargo_detail_cache_signature"] = detail_cache_signature

    cargo_weight_col = cargo_first_column(rp_voyage, ["CargoWeight", "Cargo Weight", "CargoMT"])
    cargo_teu_col = cargo_first_column(rp_voyage, ["CargoTEU", "Cargo TEU", "TEU"])
    draft_f_col = cargo_first_column(rp_voyage, ["DraftFore", "Draft Forward"])
    draft_a_col = cargo_first_column(rp_voyage, ["DraftAft", "Draft Aft"])
    cargo_weight = voyage.get("Cargo [MT]")
    cargo_teu = voyage.get("Cargo TEU")
    departure_port = voyage.get("Departure")
    arrival_port = voyage.get("Arrival")
    duration_days = max((voyage_end - voyage_start).total_seconds() / 86400.0, 0.0)

    render_metric_cards([
        ("Cargo Weight [MT]", cargo_format_number(cargo_weight, 1), "cargo_weight"),
        ("Cargo TEU", cargo_format_number(cargo_teu, 1), "cargo_teu"),
        ("Voyage Duration", f"{duration_days:,.2f} days", "voyage_duration"),
        ("Reports", f"{len(cargo_by_report):,}", "report_count"),
    ])
    render_metric_cards([
        ("Total Fuel [MT]", cargo_format_number(fuel_totals["Total Fuel Consumption [MT]"], 2), "fuel_total"),
        ("MGO [MT]", cargo_format_number(fuel_totals["MGO Consumption [MT]"], 2), "fuel_grade"),
        ("HFO [MT]", cargo_format_number(fuel_totals["HFO Consumption [MT]"], 2), "fuel_grade"),
        ("LFO [MT]", cargo_format_number(fuel_totals["LFO Consumption [MT]"], 2), "fuel_grade"),
    ])
    st.markdown(
        f'<div class="atlas-pill"><span>Voyage:</span> {escape(voyage_id)} &nbsp; | &nbsp; '
        f'<span>Departure:</span> {escape(str(departure_port if pd.notna(departure_port) else "-"))} &nbsp; | &nbsp; '
        f'<span>Arrival:</span> {escape(str(arrival_port if pd.notna(arrival_port) else "-"))}</div>',
        unsafe_allow_html=True,
    )

    st.markdown('<div class="section-title">Cargo status by report</div>', unsafe_allow_html=True)
    status_columns = [column for column in [
        "ReportId", "ReportType", "StartDateTimeGMT", "EndDateTimeGMT", "StateName",
        "Cargo Weight [tons]", "Total Number Full Units (20 and 40ft)",
        "Total Number Empty Units (20 and 40ft)", "Total Number Reefer Units (20 and 40ft)",
        "Total Number DG Units (20 and 40ft)", "TEU Loaded Units", "TEU Discharged Units",
        "Reefers Loaded Units", "Reefers Discharged Units",
    ] if column in cargo_by_report.columns]
    if status_columns:
        render_preview_table(cargo_by_report[status_columns])
    else:
        st.info("No cargo ReportData rows are available for this voyage in the prepared snapshot.")

    st.markdown('<div class="section-title">Report / port timeline</div>', unsafe_allow_html=True)
    if not timeline.empty:
        render_preview_table(timeline)
    else:
        st.info("No report timeline rows were found for the selected voyage.")

    cargo_evolution = build_cargo_evolution_series(
        rp_voyage,
        voyage_start,
        voyage_end,
        cargo_weight_col,
        cargo_teu_col,
        initial_cargo_weight=cargo_weight,
        initial_cargo_teu=cargo_teu,
    )
    evolution_columns = [
        column
        for column in ["Cargo Weight [MT]", "Cargo TEU"]
        if column in cargo_evolution.columns and pd.to_numeric(cargo_evolution[column], errors="coerce").notna().any()
    ]
    if evolution_columns:
        st.markdown('<div class="section-title">Cargo evolution</div>', unsafe_allow_html=True)
        st.caption(
            "Cargo values are treated as onboard state: the latest reported value is carried forward until a new cargo observation is reported."
        )
        chart_indexed = cargo_evolution.set_index("DateTime")
        if "Cargo Weight [MT]" in evolution_columns:
            st.markdown("**Cargo Weight [MT]**")
            st.line_chart(chart_indexed[["Cargo Weight [MT]"]], use_container_width=True)
        if "Cargo TEU" in evolution_columns:
            st.markdown("**Cargo TEU**")
            st.line_chart(chart_indexed[["Cargo TEU"]], use_container_width=True)

    st.markdown('<div class="section-title">Fuel consumption by report</div>', unsafe_allow_html=True)
    if not fuel_voyage.empty:
        fuel_display_columns = [column for column in [
            "ReportId", "ReportType", "StartDateTimeGMT", "EndDateTimeGMT", "StateName",
            "MGO Consumption [MT]", "HFO Consumption [MT]", "LFO Consumption [MT]",
            "Total Fuel Consumption [MT]",
        ] if column in fuel_voyage.columns]
        render_preview_table(fuel_voyage[fuel_display_columns])
    else:
        st.info("No fuel-consumption ReportData rows are available inside this voyage interval.")

    if not cargo_by_report.empty:
        st.markdown('<div class="section-title">Selected report details</div>', unsafe_allow_html=True)
        report_rows = cargo_by_report.reset_index(drop=True)
        report_idx = st.selectbox(
            "Report",
            options=report_rows.index.tolist(),
            format_func=lambda idx: (
                f"{pd.to_datetime(report_rows.loc[idx, 'StartDateTimeGMT'], errors='coerce', utc=True).strftime('%d/%m/%Y %H:%M') if pd.notna(pd.to_datetime(report_rows.loc[idx, 'StartDateTimeGMT'], errors='coerce', utc=True)) else '-'}"
                f" | {report_rows.loc[idx, 'ReportType']} | Report {report_rows.loc[idx, 'ReportId']}"
            ),
            key="atlas_cargo_report_idx",
        )
        selected_report = report_rows.loc[[report_idx]].copy()
        detail_tabs = ["Cargo Summary", "Containers", "Operations", "Draft & Ballast"]
        detail_tab = render_text_tab_bar(
            detail_tabs,
            st.session_state.get("atlas_cargo_detail_tab", "Cargo Summary"),
            param_name="cargo_detail",
            css_class="cargo-detail",
        )
        st.session_state["atlas_cargo_detail_tab"] = detail_tab
        fields_map = {
            "Cargo Summary": CARGO_SUMMARY_FIELDS,
            "Containers": CARGO_CONTAINER_FIELDS,
            "Operations": CARGO_OPERATION_FIELDS,
            "Draft & Ballast": CARGO_DRAFT_FIELDS,
        }
        fields = [column for column in [*CARGO_REPORT_IDENTITY_COLUMNS, *fields_map[detail_tab]] if column in selected_report.columns]
        detail_df = selected_report[fields].copy()
        identity = [column for column in CARGO_REPORT_IDENTITY_COLUMNS if column in detail_df.columns]
        value_fields = [column for column in detail_df.columns if column not in identity]
        if value_fields:
            key_values = pd.DataFrame({"Field": value_fields, "Value": [detail_df.iloc[0][column] for column in value_fields]})
            key_values = key_values[key_values["Value"].notna()].copy()
            render_preview_table(key_values)
        else:
            st.info("No values are available in this category for the selected report.")

    detail_overview = pd.DataFrame([{
        "ShipName": vessel, "VoyageId": voyage_id, "VoyageStart": voyage_start, "VoyageEnd": voyage_end,
        "DurationDays": duration_days, "DeparturePort": departure_port, "ArrivalPort": arrival_port,
        "CargoWeightMT": cargo_weight, "CargoTEU": cargo_teu,
        "TotalFuelConsumptionMT": fuel_totals["Total Fuel Consumption [MT]"],
        "TotalMGOConsumptionMT": fuel_totals["MGO Consumption [MT]"],
        "TotalHFOConsumptionMT": fuel_totals["HFO Consumption [MT]"],
        "TotalLFOConsumptionMT": fuel_totals["LFO Consumption [MT]"],
        "DraftFore": cargo_latest_value(rp_voyage, draft_f_col), "DraftAft": cargo_latest_value(rp_voyage, draft_a_col),
        "CargoReports": len(cargo_by_report),
    }])
    export_signature = sha256(f"{vessel}|{voyage_id}|{voyage_start}|{voyage_end}|{len(cargo_by_report)}|{len(rp_voyage)}".encode("utf-8")).hexdigest()
    cargo_export_ready = st.session_state.get("atlas_cargo_export_signature") == export_signature and "atlas_cargo_export_bytes" in st.session_state
    if st.button("Prepare voyage Excel", type="primary", key="atlas_prepare_cargo_excel"):
        with st.spinner("Preparing voyage workbook..."):
            st.session_state["atlas_cargo_export_bytes"] = cargo_excel_bytes(detail_overview, timeline, cargo_by_report, rp_voyage, ship_voyage, fuel_voyage)
            st.session_state["atlas_cargo_export_signature"] = export_signature
        cargo_export_ready = True
    if cargo_export_ready:
        st.download_button(
            "Download voyage Excel", data=st.session_state["atlas_cargo_export_bytes"],
            file_name=f"atlasflow_cargo_{normalize_text(vessel)}_{normalize_text(voyage_id)}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="atlas_download_cargo_excel",
        )

# =============================================================================
# Main app
# =============================================================================


def main() -> None:
    run_warmup_if_requested()
    require_dashboard_password()
    apply_custom_css()

    username = read_secret("MARORKA_USERNAME")
    password = read_secret("MARORKA_PASSWORD")
    token = read_secret("MARORKA_TOKEN")
    auth_method = read_secret("MARORKA_AUTH_METHOD", "basic")

    if auth_method.lower() in {"basic", "digest"} and (not username or not password):
        st.info("Add MARORKA_USERNAME and MARORKA_PASSWORD to .streamlit/secrets.toml or Streamlit Cloud Secrets.")
        st.stop()

    api_start_date = API_FULL_START_DATE
    api_end_date = date.today()

    refresh = sidebar_refresh_control()
    selected_group, selected_vessels = selected_vessel_controls()

    if refresh:
        try:
            with st.spinner("Refreshing AtlasFlow APIs one source at a time..."):
                refresh_all_atlasflow_snapshots(
                    username,
                    password,
                    token,
                    auth_method,
                    api_start_date,
                )
        except requests.HTTPError as exc:
            status = exc.response.status_code if exc.response is not None else "unknown"
            st.error(
                f"AtlasFlow refresh failed with status {status}. "
                "Existing session data and the last valid snapshots remain available."
            )
            st.stop()
        except (
            MarorkaConfigError,
            ValueError,
            FileNotFoundError,
            RuntimeError,
            requests.RequestException,
        ) as exc:
            st.error(
                "AtlasFlow refresh failed. Existing session data and the last valid "
                f"snapshots remain available. Details: {exc}"
            )
            st.stop()
        st.success("All AtlasFlow API snapshots refreshed successfully.")
        st.rerun()

    reportdata_signature = atlas_source_signature(
        "reportdata",
        username,
        auth_method,
        api_start_date,
    )
    reportdata_manifest = read_source_manifest("reportdata")
    current_generation = (
        reportdata_manifest.get("generation")
        if isinstance(reportdata_manifest, dict)
        else None
    )
    long_df = st.session_state.get("loaded_long_df")
    metadata = st.session_state.get("loaded_metadata")
    session_signature = st.session_state.get("loaded_request_signature")
    session_generation = st.session_state.get("loaded_reportdata_generation")
    session_ready = (
        isinstance(long_df, pd.DataFrame)
        and isinstance(metadata, dict)
        and source_signature_covers_request(
            session_signature,
            reportdata_signature,
            metadata,
            api_start_date,
        )
        and session_generation == current_generation
    )

    if not session_ready:
        snapshot = load_source_snapshot(
            "reportdata",
            reportdata_signature,
            api_start_date,
        )
        if snapshot is None:
            migrated = ensure_source_snapshot("reportdata", username, auth_method)
            if migrated is not None:
                snapshot = load_source_snapshot(
                    "reportdata",
                    reportdata_signature,
                    api_start_date,
                )
        if snapshot is None:
            st.warning(
                "No prepared ReportData snapshot is available yet. Run the AtlasFlow warmup first; "
                "normal users will not be forced to wait for the large API pull."
            )
            st.code(
                "https://atlas-flow.streamlit.app/?warmup=1&force=1&source=all&token=warmup-atlas-flow",
                language="text",
            )
            st.stop()

        long_df, metadata, snapshot_signature = snapshot
        st.session_state.pop("loaded_raw_df", None)
        st.session_state["loaded_long_df"] = long_df
        st.session_state["loaded_metadata"] = metadata
        st.session_state["loaded_request_signature"] = snapshot_signature
        st.session_state["loaded_prepare_signature"] = source_data_signature("reportdata")
        st.session_state["loaded_reportdata_generation"] = metadata.get("snapshot_generation")

    long_df = st.session_state.get("loaded_long_df")
    metadata = st.session_state.get("loaded_metadata")
    if not isinstance(long_df, pd.DataFrame) or not isinstance(metadata, dict):
        st.error("The prepared ReportData snapshot could not be loaded.")
        st.stop()

    if long_df.empty:
        render_header(selected_group, selected_vessels, [])
        render_api_load_caption(metadata)
        st.warning("No Marorka report values were returned for the loaded API window.")
        st.stop()

    selected_start, selected_end = render_date_slicer(long_df)

    # ReportType is handled with the rest of the displayed-column filters.
    selected_report_types: list[str] = []
    long_filter_signature = sha256(
        (
            f"{metadata.get('snapshot_generation')}|{tuple(selected_vessels)}|"
            f"{selected_start.isoformat()}|{selected_end.isoformat()}|{tuple(selected_report_types)}"
        ).encode("utf-8")
    ).hexdigest()
    cached_filtered_long = st.session_state.get("atlas_filtered_long_cache")
    if (
        st.session_state.get("atlas_filtered_long_cache_signature") == long_filter_signature
        and isinstance(cached_filtered_long, pd.DataFrame)
    ):
        filtered_long_for_options = cached_filtered_long
    else:
        filtered_long_for_options = filter_long_data(
            long_df,
            selected_vessels=selected_vessels,
            selected_report_types=selected_report_types,
            selected_start=selected_start,
            selected_end=selected_end,
        )
        # Keep a single modest selection in memory; wide fleet filters are rebuilt
        # rather than risking Streamlit Cloud memory pressure.
        if dataframe_memory_mb(filtered_long_for_options) <= 96:
            st.session_state["atlas_filtered_long_cache"] = filtered_long_for_options
            st.session_state["atlas_filtered_long_cache_signature"] = long_filter_signature
        else:
            st.session_state.pop("atlas_filtered_long_cache", None)
            st.session_state.pop("atlas_filtered_long_cache_signature", None)

    variable_options = sorted(
        set(available_variables(filtered_long_for_options)).union(DERIVED_VARIABLES),
        key=str.casefold,
    )
    st.sidebar.markdown("### Pivot variables")
    previous_selected_variables = st.session_state.get("atlas_selected_variables", [])
    if not isinstance(previous_selected_variables, list):
        previous_selected_variables = []
    valid_default_variables = [variable for variable in previous_selected_variables if variable in variable_options]
    selected_variables = st.sidebar.multiselect(
        "Variables to include and filter",
        options=variable_options,
        default=valid_default_variables,
        key="atlas_selected_variables",
        help=(
            "Every selected ValueDescription becomes a displayed table column. "
            "The same selected variables are also available below as filters."
        ),
    )

    identity_columns = st.sidebar.multiselect(
        "Pivot rows",
        options=PIVOT_IDENTITY_COLUMNS,
        default=["ShipName"],
        help="Choose the row fields that appear before the selected variable columns.",
    )
    if not identity_columns:
        identity_columns = ["ShipName"]

    workspace_options = [
        "Monthly Comparison",
        "Voyage Analysis",
        "Lub Oil Analysis",
        "Custom Analytics",
        "Noon & Manual Reports",
        "High-Frequency",
        "Descriptive Statistics",
        "Export Center",
        "API Diagnostics",
    ]

    render_header(selected_group, selected_vessels, selected_variables)
    

    # Top navigation keeps the previous tab-like layout while preserving the
    # memory optimization: only the selected workspace loads its heavy data.
    workspace = get_tab_selection(
        "workspace",
        workspace_options,
        st.session_state.get("atlas_workspace", "Monthly Comparison"),
    )
    workspace = render_text_tab_bar(
        workspace_options,
        workspace,
        param_name="workspace",
        reset_params=["preview"],
    )
    st.session_state["atlas_workspace"] = workspace

    active_wide_sources: set[str] = set()
    if workspace == "Voyage Analysis":
        active_wide_sources.update({"reportpivots", "shippivots"})
    elif workspace == "Noon & Manual Reports":
        active_wide_sources.add("reportpivots")
    elif workspace == "High-Frequency":
        active_wide_sources.add("shippivots")
    elif workspace == "Descriptive Statistics":
        descriptive_source_hint = st.session_state.get("atlas_descriptive_source_selector", "Custom Analytics")
        if descriptive_source_hint == "Noon & Manual Reports":
            active_wide_sources.add("reportpivots")
        elif descriptive_source_hint == "High-Frequency":
            active_wide_sources.add("shippivots")
    elif workspace == "Export Center":
        # Wide sources are loaded inside the export button only, not during normal render.
        active_wide_sources = set()
    clear_inactive_wide_sources(active_wide_sources)

    if metadata.get("hit_page_limit"):
        st.warning(
            "The API refresh reached the maximum page safety limit before the feed ended. "
            "The loaded dataset may be incomplete. Check API Diagnostics before using the export."
        )

    if workspace in {"Monthly Comparison", "Voyage Analysis", "Lub Oil Analysis"}:
        # Monthly Comparison reads pre-aggregated summaries. Voyage Analysis
        # builds its own voyage/report views directly from the prepared sources.
        # does not need to build the report-level pivot or its filter controls.
        pivot_df = pd.DataFrame()
        filtered_pivot_df = pd.DataFrame()
        filter_specs: list[dict[str, Any]] = []
        display_columns: list[str] = []
        output_df = pd.DataFrame()
    else:
        if not selected_variables:
            st.info("Select one or more variables from the sidebar to build the AtlasFlow pivot table.")

        pivot_df = build_pivot_table(filtered_long_for_options, tuple(selected_variables))

        filter_column_options = [column for column in [*identity_columns, *selected_variables] if column in pivot_df.columns]
        with st.sidebar.expander("Filters for displayed columns", expanded=False):
            st.caption("Choose columns to filter. Selected variables are already part of the displayed table.")
            previous_filter_columns = st.session_state.get("atlas_columns_to_filter", [])
            if not isinstance(previous_filter_columns, list):
                previous_filter_columns = []
            valid_filter_columns = [column for column in previous_filter_columns if column in filter_column_options]
            if valid_filter_columns != previous_filter_columns:
                st.session_state["atlas_columns_to_filter"] = valid_filter_columns

            columns_to_filter = st.multiselect(
                "Columns to filter",
                options=filter_column_options,
                default=valid_filter_columns,
                key="atlas_columns_to_filter",
            )
            filter_specs = render_column_filters(pivot_df, columns_to_filter)

        filtered_pivot_df = apply_column_filters(pivot_df, filter_specs)

        display_columns = []
        for column in [*identity_columns, *selected_variables]:
            if column in filtered_pivot_df.columns and column not in display_columns:
                display_columns.append(column)
        if not display_columns:
            display_columns = [column for column in DEFAULT_DISPLAY_IDENTITY_COLUMNS if column in filtered_pivot_df.columns]

        output_df = filtered_pivot_df[display_columns].copy()

    # Shared Custom Analytics preview/export configuration. It is only fully rendered in
    # Custom Analytics, but Export Center can reuse the saved choices.
    summary_group_fields: list[str] = []
    summary_value_fields: list[str] = []
    summary_aggregation = st.session_state.get("atlas_export_summary_aggregation", "Average")
    preview_mode = st.session_state.get("atlas_reportdata_preview_mode", "Clean Dataset")
    displayed_table_df = output_df.copy()
    export_sheet_name = "Clean Dataset"
    current_export_signature = sha256(
        "|".join([
            preview_mode,
            ",".join(selected_vessels),
            selected_start.isoformat(),
            selected_end.isoformat(),
            ",".join(display_columns),
            ",".join(selected_variables),
            str(len(output_df)),
        ]).encode("utf-8")
    ).hexdigest()

    if workspace == "Monthly Comparison":
        render_monthly_comparison_workspace(
            username,
            auth_method,
            selected_vessels,
            selected_start,
            selected_end,
        )

    elif workspace == "Voyage Analysis":
        render_cargo_voyages_workspace(
            username,
            password,
            token,
            auth_method,
            api_start_date,
            long_df,
            metadata,
            selected_vessels,
            selected_start,
            selected_end,
        )

    elif workspace == "Lub Oil Analysis":
        render_lubricating_oil_workspace(filtered_long_for_options)

    elif workspace == "Custom Analytics":
        st.markdown('<div class="section-title">Custom Analytics Preview & Export</div>', unsafe_allow_html=True)

        render_api_load_caption(metadata)

        summary_builder_columns = [column for column in output_df.columns]
        summary_value_options = numeric_column_options(output_df)

        st.caption(
            "Choose which table you want to preview and export. The visible table below is the same table prepared for Excel."
        )
        preview_options = ["Clean Dataset", "Summary Analysis", "Source Data"]
        preview_mode = get_tab_selection(
            "preview",
            preview_options,
            st.session_state.get("atlas_reportdata_preview_mode", "Clean Dataset"),
        )
        preview_mode = render_text_tab_bar(
            preview_options,
            preview_mode,
            param_name="preview",
            css_class="compact",
        )
        st.session_state["atlas_reportdata_preview_mode"] = preview_mode

        if preview_mode == "Summary Analysis":
            st.markdown('<div class="section-title">Summary Builder</div>', unsafe_allow_html=True)
            builder_cols = st.columns(3)
            with builder_cols[0]:
                previous_summary_groups = st.session_state.get("atlas_export_summary_groups", [])
                if not isinstance(previous_summary_groups, list):
                    previous_summary_groups = []
                default_summary_groups = [column for column in ["ShipName", "ReportType"] if column in summary_builder_columns]
                valid_summary_group_defaults = [column for column in previous_summary_groups if column in summary_builder_columns]
                if not valid_summary_group_defaults and "atlas_export_summary_groups" not in st.session_state:
                    valid_summary_group_defaults = default_summary_groups
                if valid_summary_group_defaults != previous_summary_groups:
                    st.session_state["atlas_export_summary_groups"] = valid_summary_group_defaults
                summary_group_fields = st.multiselect(
                    "Group by fields",
                    options=summary_builder_columns,
                    default=valid_summary_group_defaults,
                    key="atlas_export_summary_groups",
                    help="Choose the fields that define each summary row.",
                )
            with builder_cols[1]:
                previous_summary_values = st.session_state.get("atlas_export_summary_values", [])
                if not isinstance(previous_summary_values, list):
                    previous_summary_values = []
                valid_summary_value_defaults = [column for column in previous_summary_values if column in summary_value_options]
                if valid_summary_value_defaults != previous_summary_values:
                    st.session_state["atlas_export_summary_values"] = valid_summary_value_defaults
                summary_value_fields = st.multiselect(
                    "Value fields",
                    options=summary_value_options,
                    default=valid_summary_value_defaults,
                    key="atlas_export_summary_values",
                    help="Choose one or more numeric columns to aggregate.",
                )
            with builder_cols[2]:
                summary_aggregation = st.selectbox(
                    "Aggregation",
                    options=["Average", "Sum", "Count", "Minimum", "Maximum", "Median"],
                    index=["Average", "Sum", "Count", "Minimum", "Maximum", "Median"].index(summary_aggregation)
                    if summary_aggregation in ["Average", "Sum", "Count", "Minimum", "Maximum", "Median"] else 0,
                    key="atlas_export_summary_aggregation",
                )
        else:
            summary_group_fields = st.session_state.get("atlas_export_summary_groups", [])
            summary_value_fields = st.session_state.get("atlas_export_summary_values", [])
            if not isinstance(summary_group_fields, list):
                summary_group_fields = []
            if not isinstance(summary_value_fields, list):
                summary_value_fields = []

        summary_can_build = bool(summary_group_fields and summary_value_fields)
        if preview_mode == "Summary Analysis" and summary_can_build:
            displayed_table_df = build_summary_analysis(
                output_df,
                group_fields=summary_group_fields,
                value_fields=summary_value_fields,
                aggregation=summary_aggregation,
            )
            export_sheet_name = "Summary Analysis"
        elif preview_mode == "Summary Analysis":
            displayed_table_df = pd.DataFrame()
            export_sheet_name = "Summary Analysis"
            st.info("Select at least one Group by field and one Value field to preview Summary Analysis.")
        elif preview_mode == "Source Data":
            source_columns = [column for column in [*SOURCE_COLUMNS, "ParsedValue"] if column in filtered_long_for_options.columns]
            displayed_table_df = filtered_long_for_options[source_columns].copy()
            export_sheet_name = "Source Data"
        else:
            displayed_table_df = output_df.copy()
            export_sheet_name = "Clean Dataset"

        render_metric_cards(
            [
                ("Displayed Rows", f"{len(displayed_table_df):,}", "table_eye"),
                ("Selected Variables", f"{len(selected_variables):,}", "checked_columns"),
                ("Source Rows", f"{len(filtered_long_for_options):,}", "database_rows"),
                ("Available Variables", f"{len(variable_options):,}", "columns_plus"),
            ]
        )

        render_preview_table(displayed_table_df)
        if len(displayed_table_df) > TABLE_PREVIEW_ROW_LIMIT:
            st.caption(
                f"Showing first {TABLE_PREVIEW_ROW_LIMIT:,} of {len(displayed_table_df):,} rows. "
                "Excel export includes the full displayed table."
            )

        export_signature_payload = "|".join([
            preview_mode,
            ",".join(selected_vessels),
            selected_start.isoformat(),
            selected_end.isoformat(),
            ",".join(display_columns),
            ",".join(selected_variables),
            str(len(output_df)),
            str(len(displayed_table_df)),
            ",".join(summary_group_fields),
            ",".join(summary_value_fields),
            str(summary_aggregation),
            ",".join(displayed_table_df.columns.astype(str).tolist()) if not displayed_table_df.empty else "empty",
        ])
        current_export_signature = sha256(export_signature_payload.encode("utf-8")).hexdigest()
        clear_stale_export_bytes(current_export_signature)

        export_ready = (
            st.session_state.get("atlas_export_signature") == current_export_signature
            and "atlas_export_bytes" in st.session_state
        )

        if st.button("Prepare displayed table Excel", type="primary", disabled=displayed_table_df.empty):
            with st.spinner("Preparing Excel file..."):
                st.session_state["atlas_export_bytes"] = to_displayed_table_excel_bytes(
                    displayed_table_df,
                    sheet_name=export_sheet_name,
                )
                st.session_state["atlas_summary_analysis_df"] = displayed_table_df if preview_mode == "Summary Analysis" else pd.DataFrame()
                st.session_state["atlas_export_signature"] = current_export_signature
                gc.collect()
            export_ready = True

        if export_ready:
            st.download_button(
                "Download displayed table Excel",
                data=st.session_state["atlas_export_bytes"],
                file_name="atlasflow_displayed_table.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        else:
            st.caption("Excel generation is prepared on demand. The download will contain only the visible table above.")

    elif workspace == "Noon & Manual Reports":
        reportpivots_df, reportpivots_metadata = load_wide_source_for_view(
            "reportpivots", username, password, token, auth_method, api_start_date, refresh,
            selected_vessels, selected_start, selected_end
        )
        if reportpivots_metadata.get("needs_warmup"):
            st.info("No ReportPivots snapshot is available yet. Run the ReportPivots warmup URL first.")
            st.code("https://atlas-flow.streamlit.app/?warmup=1&force=1&source=reportpivots&token=warmup-atlas-flow", language="text")
        else:
            render_wide_source_tab(
                "Noon & Manual Reports",
                reportpivots_df,
                reportpivots_metadata,
                "reportpivots",
                selected_vessels,
                selected_start,
                selected_end,
            )

    elif workspace == "High-Frequency":
        shippivots_df, shippivots_metadata = load_wide_source_for_view(
            "shippivots", username, password, token, auth_method, api_start_date, refresh,
            selected_vessels, selected_start, selected_end
        )
        if shippivots_metadata.get("needs_warmup"):
            st.info("No ShipPivots snapshot is available yet. Run the ShipPivots warmup URL first.")
            st.code("https://atlas-flow.streamlit.app/?warmup=1&force=1&source=shippivots&token=warmup-atlas-flow", language="text")
        else:
            render_wide_source_tab(
                "High-Frequency",
                shippivots_df,
                shippivots_metadata,
                "shippivots",
                selected_vessels,
                selected_start,
                selected_end,
            )

    elif workspace == "Descriptive Statistics":
        st.markdown('<div class="section-title">Descriptive Statistics</div>', unsafe_allow_html=True)
        st.caption("Analyze one source at a time. This avoids loading all large API tables into memory together.")
        selected_source = st.selectbox(
            "Source table",
            options=["Custom Analytics", "Noon & Manual Reports", "High-Frequency"],
            key="atlas_descriptive_source_selector",
        )
        if selected_source == "Custom Analytics":
            analysis_df = output_df.copy()
        elif selected_source == "Noon & Manual Reports":
            source_df, source_metadata = load_wide_source_for_view(
                "reportpivots", username, password, token, auth_method, api_start_date, refresh,
                selected_vessels, selected_start, selected_end
            )
            if source_metadata.get("needs_warmup"):
                st.info("No ReportPivots snapshot is available yet. Run the ReportPivots warmup URL first.")
                st.stop()
            analysis_df = build_wide_source_output_for_export("reportpivots", source_df, selected_vessels, selected_start, selected_end)
        else:
            source_df, source_metadata = load_wide_source_for_view(
                "shippivots", username, password, token, auth_method, api_start_date, refresh,
                selected_vessels, selected_start, selected_end
            )
            if source_metadata.get("needs_warmup"):
                st.info("No ShipPivots snapshot is available yet. Run the ShipPivots warmup URL first.")
                st.stop()
            analysis_df = build_wide_source_output_for_export("shippivots", source_df, selected_vessels, selected_start, selected_end)

        numeric_options = dataframe_numeric_options(analysis_df)
        if not numeric_options:
            st.info("The selected source table has no numeric columns to analyze.")
        else:
            metric_column = st.selectbox("Metric to analyze", options=numeric_options, key="atlas_descriptive_metric")
            group_options = ["None"] + dataframe_categorical_options(analysis_df)
            default_group_index = group_options.index("ShipName") if "ShipName" in group_options else 0
            group_column = st.selectbox("Optional group by", options=group_options, index=default_group_index, key="atlas_descriptive_group")
            stats_df = build_descriptive_statistics(analysis_df, metric_column)
            values = pd.to_numeric(analysis_df[metric_column], errors="coerce")
            render_metric_cards(
                [
                    ("Numeric Values", f"{values.notna().sum():,}", "numeric"),
                    ("Total", f"{values.sum(skipna=True):,.3f}", "total"),
                    ("Average", f"{values.mean(skipna=True):,.3f}", "average"),
                    ("Missing", f"{values.isna().sum():,}", "missing"),
                ]
            )
            st.markdown('<div class="section-title">Overall statistics</div>', unsafe_allow_html=True)
            st.dataframe(format_display_dataframe(stats_df), use_container_width=True, hide_index=True)
            if group_column != "None":
                grouped_df = build_grouped_descriptive_statistics(analysis_df, metric_column, group_column)
                if not grouped_df.empty:
                    st.markdown('<div class="section-title">Grouped statistics</div>', unsafe_allow_html=True)
                    st.dataframe(format_display_dataframe(grouped_df.head(100)), use_container_width=True, hide_index=True)
            datetime_column = detect_analysis_datetime_column(analysis_df)
            if datetime_column:
                trend_df = build_monthly_trend(analysis_df, metric_column, datetime_column)
                if not trend_df.empty:
                    st.markdown('<div class="section-title">Monthly trend</div>', unsafe_allow_html=True)
                    st.dataframe(format_display_dataframe(trend_df), use_container_width=True, hide_index=True)
                    st.line_chart(trend_df.set_index("Month")[["Sum", "Mean"]])
        del analysis_df
        gc.collect()

    elif workspace == "Export Center":
        st.markdown('<div class="section-title">AtlasFlow Export Center</div>', unsafe_allow_html=True)
        st.caption(
            "The full workbook is prepared on demand. ReportPivots and ShipPivots are loaded only while creating the workbook, then released."
        )
        render_metric_cards(
            [
                ("Custom Analytics Rows", f"{len(output_df):,}", "table_eye"),
                ("Noon & Manual Rows", "loaded on demand", "report_rows"),
                ("High-Frequency Rows", "loaded on demand", "time_series_rows"),
            ]
        )

        summary_group_fields = st.session_state.get("atlas_export_summary_groups", [])
        summary_value_fields = st.session_state.get("atlas_export_summary_values", [])
        if not isinstance(summary_group_fields, list):
            summary_group_fields = []
        if not isinstance(summary_value_fields, list):
            summary_value_fields = []
        summary_can_build = bool(summary_group_fields and summary_value_fields)

        multisource_signature_payload = "|".join([
            current_export_signature,
            ",".join(selected_vessels),
            selected_start.isoformat(),
            selected_end.isoformat(),
            ",".join(display_columns),
            ",".join(selected_variables),
            ",".join(summary_group_fields),
            ",".join(summary_value_fields),
            str(summary_aggregation),
        ])
        multisource_signature = sha256(multisource_signature_payload.encode("utf-8")).hexdigest()
        multisource_ready = (
            st.session_state.get("atlas_multisource_export_signature") == multisource_signature
            and "atlas_multisource_export_bytes" in st.session_state
        )

        if st.button("Prepare full AtlasFlow workbook", type="primary"):
            with st.spinner("Loading source snapshots and preparing workbook..."):
                reportpivots_df, reportpivots_metadata = load_wide_source_for_view(
                    "reportpivots", username, password, token, auth_method, api_start_date, refresh=False,
                    selected_vessels=selected_vessels, selected_start=selected_start, selected_end=selected_end
                )
                shippivots_df, shippivots_metadata = load_wide_source_for_view(
                    "shippivots", username, password, token, auth_method, api_start_date, refresh=False,
                    selected_vessels=selected_vessels, selected_start=selected_start, selected_end=selected_end
                )
                reportpivots_output_df = build_wide_source_output_for_export(
                    "reportpivots", reportpivots_df, selected_vessels, selected_start, selected_end
                ) if not reportpivots_metadata.get("needs_warmup") else pd.DataFrame()
                shippivots_output_df = build_wide_source_output_for_export(
                    "shippivots", shippivots_df, selected_vessels, selected_start, selected_end
                ) if not shippivots_metadata.get("needs_warmup") else pd.DataFrame()
                summary_analysis_df = pd.DataFrame()
                if summary_can_build:
                    summary_analysis_df = build_summary_analysis(
                        output_df,
                        group_fields=summary_group_fields,
                        value_fields=summary_value_fields,
                        aggregation=summary_aggregation,
                    )
                st.session_state["atlas_multisource_export_bytes"] = to_multisource_excel_bytes(
                    output_df,
                    summary_analysis_df if not summary_analysis_df.empty else None,
                    reportpivots_output_df,
                    shippivots_output_df,
                )
                st.session_state["atlas_multisource_export_signature"] = multisource_signature
                st.session_state["atlas_summary_analysis_df"] = summary_analysis_df
                st.session_state["atlas_export_signature"] = current_export_signature
                del reportpivots_df, shippivots_df, reportpivots_output_df, shippivots_output_df, summary_analysis_df
                clear_inactive_wide_sources(set())
                gc.collect()
            multisource_ready = True

        if multisource_ready:
            st.download_button(
                "Download full AtlasFlow workbook",
                data=st.session_state["atlas_multisource_export_bytes"],
                file_name="atlasflow_full_export.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        else:
            st.caption("Run the individual warmups first if the workbook is missing Noon & Manual or High-Frequency sheets.")

    elif workspace == "API Diagnostics":
        st.markdown('<div class="section-title">Diagnostics</div>', unsafe_allow_html=True)
        diagnostics = pd.DataFrame(
            {
                "Metric": [
                    "Selected vessels",
                    "API start date",
                    "API end date",
                    "Selected start",
                    "Selected end",
                    "API loaded at",
                    "API loaded local time",
                    "Loaded from snapshot",
                    "Snapshot saved at",
                    "Kept raw rows",
                    "Original API rows scanned",
                    "Discarded rows",
                    "Long rows prepared",
                    "Filtered long rows",
                    "Pivot rows before column filters",
                    "Pivot rows after column filters",
                    "Available variables",
                    "Selected variables",
                    "API pages",
                    "Downloaded MB",
                    "API fetch seconds",
                    "Prepare seconds",
                    "Hit API page limit",
                    "Paging stop reason",
                    "Max page safety limit",
                ],
                "Value": [
                    ", ".join(selected_vessels),
                    api_start_date.isoformat(),
                    api_end_date.isoformat(),
                    selected_start.isoformat(),
                    selected_end.isoformat(),
                    metadata.get("loaded_at_utc", "-"),
                    metadata.get("loaded_at_local", "-"),
                    str(metadata.get("loaded_from_snapshot", False)),
                    metadata.get("snapshot_saved_at_utc", "-"),
                    f"{metadata.get('kept_rows', metadata.get('rows', 0)):,}",
                    f"{metadata.get('scanned_rows', 0):,}",
                    f"{metadata.get('discarded_rows', 0):,}",
                    f"{len(long_df):,}",
                    f"{len(filtered_long_for_options):,}",
                    f"{len(pivot_df):,}",
                    f"{len(output_df):,}",
                    f"{len(variable_options):,}",
                    f"{len(selected_variables):,}",
                    f"{metadata.get('pages', 0):,}",
                    metadata.get("downloaded_mb", "-"),
                    metadata.get("fetch_seconds", "-"),
                    metadata.get("prepare_seconds", "-"),
                    str(metadata.get("hit_page_limit", "-")),
                    metadata.get("paging_stop_reason", "-"),
                    f"{metadata.get('max_pages', MAX_ODATA_PAGES):,}",
                ],
            }
        )
        st.dataframe(diagnostics, use_container_width=True, hide_index=True)

        with st.expander("First API URL", expanded=False):
            st.code(metadata.get("first_url", "-"), language="text")

        st.markdown('<div class="section-title">Memory Audit</div>', unsafe_allow_html=True)
        audit_df = current_memory_audit_rows({
            "local.filtered_long_for_options": filtered_long_for_options,
            "local.pivot_df": pivot_df,
            "local.output_df": output_df,
        })
        st.dataframe(audit_df, use_container_width=True, hide_index=True)
        if st.button("Clear wide-source memory and Excel buffers"):
            clear_inactive_wide_sources(set())
            clear_stale_export_bytes(None)
            st.success("Released inactive wide-source DataFrames and export byte buffers from this session.")
            st.rerun()

        st.markdown('<div class="section-title">Available Variable Counts</div>', unsafe_allow_html=True)
        if st.button("Calculate variable counts"):
            value_counts = (
                filtered_long_for_options.get("ValueDescription", pd.Series(dtype="object"))
                .value_counts(dropna=False)
                .reset_index()
            )
            value_counts.columns = ["ValueDescription", "Rows"]
            st.dataframe(value_counts.head(500), use_container_width=True, hide_index=True)
        else:
            st.caption("Variable counts are calculated on demand so diagnostics do not slow normal loads.")

    # Release the largest temporary views created during this run.
    del pivot_df, filtered_pivot_df
    gc.collect()


if __name__ == "__main__":
    main()
